#!/usr/bin/env python3
"""
Score Validation Script

Compares manually entered scores in Excel against calculated scores from nflreadpy.
Outputs a list of discrepancies for review.

Usage:
    python validate_scores.py --sheet "Week 12" --season 2025 --week 12
    python validate_scores.py --all  # Validate all weeks
"""

import argparse
import re
from typing import List, Tuple, Optional

import openpyxl

from qpfl import QPFLScorer, parse_roster_from_excel
from qpfl.constants import POSITION_ROWS
from qpfl.excel_parser import parse_player_name


def get_excel_scores(filepath: str, sheet_name: str) -> dict:
    """
    Extract manually entered scores from Excel.
    
    Returns:
        Dict mapping (team_name, position, player_name) -> score
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb[sheet_name]
    
    scores = {}
    team_columns = [1, 3, 5, 7, 9, 11, 13, 15, 17, 19]
    
    # Get team names
    team_names = {}
    for col in team_columns:
        team_name = ws.cell(row=2, column=col).value or ""
        team_name = team_name.strip().strip('*')
        if team_name:
            team_names[col] = team_name
    
    # Get player rows for each position
    position_player_rows = {pos: rows for pos, (_, rows) in POSITION_ROWS.items()}
    
    for col, team_name in team_names.items():
        points_col = col + 1
        
        for position, player_rows in position_player_rows.items():
            for row in player_rows:
                player_cell = ws.cell(row=row, column=col)
                score_cell = ws.cell(row=row, column=points_col)
                
                if player_cell.value:
                    # Check if player is started (bold)
                    is_bold = player_cell.font.bold if player_cell.font else False
                    if not is_bold:
                        continue
                    
                    player_name, nfl_team = parse_player_name(str(player_cell.value))
                    excel_score = score_cell.value
                    
                    if player_name and excel_score is not None:
                        key = (team_name, position, player_name, nfl_team)
                        scores[key] = float(excel_score)
    
    wb.close()
    return scores


def validate_week(
    excel_path: str,
    sheet_name: str,
    season: int,
    week: int,
    tolerance: float = 0.0,
) -> Tuple[List[dict], int]:
    """
    Validate scores for a single week.
    
    Args:
        excel_path: Path to Excel file
        sheet_name: Sheet name (e.g., "Week 12")
        season: NFL season year
        week: Week number
        tolerance: Allow differences up to this amount (default 0 = exact match)
    
    Returns:
        Tuple of (discrepancies list, total players checked)
    """
    discrepancies = []
    
    # Get Excel scores
    excel_scores = get_excel_scores(excel_path, sheet_name)
    
    if not excel_scores:
        print(f"  No scored players found in {sheet_name}")
        return discrepancies, 0
    
    total_checked = len(excel_scores)
    
    # Calculate scores
    scorer = QPFLScorer(season, week)
    
    for (team_name, position, player_name, nfl_team), excel_score in excel_scores.items():
        calculated = scorer.score_player(player_name, nfl_team, position)
        
        if not calculated.found_in_stats:
            # Player not found - could be bye week or data issue
            discrepancies.append({
                'week': week,
                'sheet': sheet_name,
                'team': team_name,
                'position': position,
                'player': player_name,
                'nfl_team': nfl_team,
                'excel_score': excel_score,
                'calculated_score': None,
                'difference': None,
                'reason': 'Player not found in stats',
                'breakdown': {},
            })
            continue
        
        diff = excel_score - calculated.total_points
        
        if abs(diff) > tolerance:
            discrepancies.append({
                'week': week,
                'sheet': sheet_name,
                'team': team_name,
                'position': position,
                'player': player_name,
                'nfl_team': nfl_team,
                'excel_score': excel_score,
                'calculated_score': calculated.total_points,
                'difference': diff,
                'reason': 'Score mismatch',
                'breakdown': calculated.breakdown,
            })
    
    return discrepancies, total_checked


def get_available_weeks(excel_path: str) -> List[Tuple[str, int]]:
    """Get list of (sheet_name, week_number) for all week sheets."""
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    weeks = []
    
    for sheet_name in wb.sheetnames:
        match = re.match(r'Week (\d+)', sheet_name)
        if match:
            week_num = int(match.group(1))
            weeks.append((sheet_name, week_num))
    
    wb.close()
    return sorted(weeks, key=lambda x: x[1])


def print_discrepancies(discrepancies: List[dict], total_checked: int, verbose: bool = True):
    """Pretty print discrepancies with stats."""
    matched = total_checked - len(discrepancies)
    pct = (matched / total_checked * 100) if total_checked > 0 else 0
    
    print(f"\n  Checked {total_checked} players: {matched} matched ({pct:.1f}%)")
    
    if not discrepancies:
        print("  ✓ All scores match!")
        return
    
    # Group by type
    mismatches = [d for d in discrepancies if d['reason'] == 'Score mismatch']
    not_found = [d for d in discrepancies if d['reason'] == 'Player not found in stats']
    
    if mismatches:
        print(f"\n  ⚠ Score Mismatches ({len(mismatches)}):")
        print("  " + "-" * 70)
        print(f"  {'Pos':<5} {'Player':<30} {'Excel':>7} {'Calc':>7} {'Diff':>7}")
        print("  " + "-" * 70)
        
        for d in mismatches:
            player_str = f"{d['player']} ({d['nfl_team']})"[:30]
            print(f"  {d['position']:<5} {player_str:<30} {d['excel_score']:>7.1f} {d['calculated_score']:>7.1f} {d['difference']:>+7.1f}")
            if verbose and d['breakdown']:
                breakdown_str = ", ".join(f"{k}: {v}" for k, v in d['breakdown'].items())
                print(f"        └─ {breakdown_str}")
        print()
    
    if not_found:
        print(f"\n  ✗ Players Not Found in Stats ({len(not_found)}):")
        print("  " + "-" * 70)
        print(f"  {'Pos':<5} {'Player':<30} {'Excel':>7} {'Note':<20}")
        print("  " + "-" * 70)
        for d in not_found:
            player_str = f"{d['player']} ({d['nfl_team']})"[:30]
            print(f"  {d['position']:<5} {player_str:<30} {d['excel_score']:>7.1f} (bye/injured/no stats)")
        print()


def main():
    parser = argparse.ArgumentParser(description="Validate QPFL scores against nflreadpy data")
    parser.add_argument(
        "--excel", "-e",
        default="2025 Scores.xlsx",
        help="Path to the Excel file",
    )
    parser.add_argument(
        "--sheet", "-s",
        help="Sheet name to validate (e.g., 'Week 12')",
    )
    parser.add_argument(
        "--season", "-y",
        type=int,
        default=2025,
        help="NFL season year",
    )
    parser.add_argument(
        "--week", "-w",
        type=int,
        help="Week number (required if --sheet is specified)",
    )
    parser.add_argument(
        "--all", "-a",
        action="store_true",
        help="Validate all weeks",
    )
    parser.add_argument(
        "--tolerance", "-t",
        type=float,
        default=0.0,
        help="Allow score differences up to this amount",
    )
    parser.add_argument(
        "--summary", 
        action="store_true",
        help="Only show summary counts, not individual discrepancies",
    )
    
    args = parser.parse_args()
    
    if args.all:
        # Validate all weeks
        weeks = get_available_weeks(args.excel)
        all_discrepancies = []
        total_players = 0
        
        print(f"Validating {len(weeks)} weeks from {args.excel}")
        print("=" * 60)
        
        for sheet_name, week_num in weeks:
            print(f"\n{sheet_name}:")
            discrepancies, checked = validate_week(
                args.excel, sheet_name, args.season, week_num, args.tolerance
            )
            all_discrepancies.extend(discrepancies)
            total_players += checked
            
            if args.summary:
                mismatches = sum(1 for d in discrepancies if d['reason'] == 'Score mismatch')
                not_found = sum(1 for d in discrepancies if d['reason'] == 'Player not found in stats')
                matched = checked - len(discrepancies)
                pct = (matched / checked * 100) if checked > 0 else 0
                print(f"  {checked} players: {matched} matched ({pct:.1f}%), {mismatches} mismatches, {not_found} not found")
            else:
                print_discrepancies(discrepancies, checked, verbose=True)
        
        # Final summary
        print("\n" + "=" * 60)
        print("SUMMARY")
        print("=" * 60)
        total_mismatches = sum(1 for d in all_discrepancies if d['reason'] == 'Score mismatch')
        total_not_found = sum(1 for d in all_discrepancies if d['reason'] == 'Player not found in stats')
        total_matched = total_players - len(all_discrepancies)
        match_pct = (total_matched / total_players * 100) if total_players > 0 else 0
        print(f"Total players checked: {total_players}")
        print(f"Total matched: {total_matched} ({match_pct:.1f}%)")
        print(f"Total mismatches: {total_mismatches}")
        print(f"Total not found: {total_not_found}")
        
    elif args.sheet:
        # Validate single week
        if not args.week:
            # Try to extract week from sheet name
            match = re.match(r'Week (\d+)', args.sheet)
            if match:
                args.week = int(match.group(1))
            else:
                parser.error("--week is required when --sheet is specified")
        
        print(f"Validating {args.sheet} (Week {args.week}, Season {args.season})")
        print("=" * 60)
        
        discrepancies, checked = validate_week(
            args.excel, args.sheet, args.season, args.week, args.tolerance
        )
        print_discrepancies(discrepancies, checked, verbose=not args.summary)
        
    else:
        parser.error("Either --sheet or --all is required")


if __name__ == "__main__":
    main()

