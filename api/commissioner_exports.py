"""Build commissioner-facing Excel snapshots from authoritative JSON data."""

from __future__ import annotations

from collections import Counter
from datetime import datetime, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

TEAM_ORDER = ['GSA', 'WJK', 'RPA', 'S/T', 'CGK', 'AST', 'CWR', 'J/J', 'SLS', 'AYP']
TEAM_COLUMNS = [1, 3, 5, 7, 9, 11, 13, 15, 17, 19]
POSITION_ORDER = ['QB', 'RB', 'WR', 'TE', 'K', 'D/ST', 'HC', 'OL']
POSITION_ROWS = {
    'QB': (6, [7, 8, 9]),
    'RB': (11, [12, 13, 14, 15]),
    'WR': (17, [18, 19, 20, 21, 22]),
    'TE': (24, [25, 26, 27]),
    'K': (29, [30, 31]),
    'D/ST': (33, [34, 35]),
    'HC': (37, [38, 39]),
    'OL': (41, [42, 43]),
}
TAXI_ROWS = [(48, 49), (50, 51), (52, 53), (54, 55)]

NAVY = '172033'
BLUE = '2563EB'
LIGHT_BLUE = 'DBEAFE'
PALE_BLUE = 'EFF6FF'
GOLD = 'F59E0B'
PALE_GOLD = 'FEF3C7'
WHITE = 'FFFFFF'
SLATE = '475569'
LIGHT_SLATE = 'E2E8F0'
PALE_SLATE = 'F8FAFC'
RED = 'B91C1C'
PALE_RED = 'FEE2E2'

THIN_SLATE = Side(style='thin', color=LIGHT_SLATE)
CELL_BORDER = Border(left=THIN_SLATE, right=THIN_SLATE, top=THIN_SLATE, bottom=THIN_SLATE)


def _timestamp(generated_at: datetime | None) -> datetime:
    value = generated_at or datetime.now(timezone.utc)
    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc)


def _team_metadata(teams_data: dict | None) -> dict[str, dict[str, str]]:
    metadata = {team: {'name': team, 'owner': ''} for team in TEAM_ORDER}
    if not isinstance(teams_data, dict):
        return metadata
    for item in teams_data.get('teams', []):
        if not isinstance(item, dict):
            continue
        abbrev = str(item.get('abbrev') or '').strip()
        if abbrev in metadata:
            metadata[abbrev] = {
                'name': str(item.get('name') or abbrev),
                'owner': str(item.get('owner') or ''),
            }
    return metadata


def _team_players(rosters: dict, team: str) -> list[dict]:
    raw = rosters.get(team, [])
    if isinstance(raw, list):
        if any(not isinstance(player, dict) for player in raw):
            raise ValueError(f'Roster for {team} contains a malformed player')
        return [dict(player) for player in raw]
    if not isinstance(raw, dict):
        raise ValueError(f'Roster for {team} is malformed')

    active = raw.get('roster', [])
    taxi = raw.get('taxi_squad', raw.get('taxi', []))
    if not isinstance(active, list) or not isinstance(taxi, list):
        raise ValueError(f'Roster for {team} is malformed')
    if any(not isinstance(player, dict) for player in [*active, *taxi]):
        raise ValueError(f'Roster for {team} contains a malformed player')
    players = [dict(player) for player in active]
    for player in taxi:
        players.append({**player, 'taxi': True})
    return players


def _format_player(player: dict) -> str:
    name = str(player.get('name') or '').strip()
    if not name:
        raise ValueError('A roster player is missing a name')
    nfl_team = str(player.get('nfl_team') or '').strip().upper()
    return f'{name} ({nfl_team})' if nfl_team else name


def _validate_rosters(rosters: dict) -> None:
    if not isinstance(rosters, dict) or not rosters:
        raise ValueError('Rosters file is malformed or empty')
    unknown_teams = sorted(set(rosters) - set(TEAM_ORDER))
    if unknown_teams:
        raise ValueError(f'Rosters contain unknown team(s): {", ".join(unknown_teams)}')

    for team in TEAM_ORDER:
        players = _team_players(rosters, team)
        active = [player for player in players if not player.get('taxi')]
        taxi = [player for player in players if player.get('taxi')]
        unknown_positions = sorted(
            {
                str(player.get('position') or '')
                for player in players
                if str(player.get('position') or '') not in POSITION_ROWS
            }
        )
        if unknown_positions:
            raise ValueError(
                f'{team} has player(s) with invalid position(s): {", ".join(unknown_positions)}'
            )
        for position, (_header_row, rows) in POSITION_ROWS.items():
            count = sum(player.get('position') == position for player in active)
            if count > len(rows):
                raise ValueError(
                    f'{team} has {count} active {position} players; maximum is {len(rows)}'
                )
        if len(taxi) > len(TAXI_ROWS):
            raise ValueError(f'{team} has {len(taxi)} taxi players; maximum is {len(TAXI_ROWS)}')
        duplicate_taxi_positions = sorted(
            position
            for position, count in Counter(player.get('position') for player in taxi).items()
            if count > 1
        )
        if duplicate_taxi_positions:
            raise ValueError(
                f'{team} has multiple taxi players at: {", ".join(duplicate_taxi_positions)}'
            )


def _set_workbook_properties(workbook: Workbook, title: str, generated_at: datetime) -> None:
    workbook.properties.creator = 'QPFL Commissioner Tools'
    workbook.properties.title = title
    workbook.properties.description = 'Generated from the authoritative QPFL JSON data.'
    workbook.properties.created = generated_at.replace(tzinfo=None)
    workbook.properties.modified = generated_at.replace(tzinfo=None)


def _save_workbook(workbook: Workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def build_roster_workbook(
    rosters: dict,
    teams_data: dict | None,
    generated_at: datetime | None = None,
) -> bytes:
    """Return a styled roster workbook compatible with the existing roster importer."""
    _validate_rosters(rosters)
    generated = _timestamp(generated_at)
    teams = _team_metadata(teams_data)

    workbook = Workbook()
    _set_workbook_properties(workbook, 'QPFL Current Rosters', generated)
    sheet = workbook.active
    sheet.title = 'Rosters'
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = 'A5'
    sheet.merge_cells('A1:S1')
    title = sheet['A1']
    title.value = 'QPFL Current Rosters'
    title.fill = PatternFill('solid', fgColor=NAVY)
    title.font = Font(color=WHITE, size=16, bold=True)
    title.alignment = Alignment(horizontal='center', vertical='center')
    sheet.row_dimensions[1].height = 28

    for column in range(1, 20):
        letter = get_column_letter(column)
        sheet.column_dimensions[letter].width = 27 if column in TEAM_COLUMNS else 2.5

    for column, team in zip(TEAM_COLUMNS, TEAM_ORDER, strict=True):
        metadata = teams[team]
        team_cell = sheet.cell(2, column, metadata['name'])
        owner_cell = sheet.cell(3, column, metadata['owner'])
        abbrev_cell = sheet.cell(4, column, team)
        for cell in (team_cell, owner_cell, abbrev_cell):
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = CELL_BORDER
        team_cell.fill = PatternFill('solid', fgColor=BLUE)
        team_cell.font = Font(color=WHITE, bold=True)
        owner_cell.fill = PatternFill('solid', fgColor=LIGHT_BLUE)
        owner_cell.font = Font(color=SLATE, italic=True)
        abbrev_cell.fill = PatternFill('solid', fgColor=NAVY)
        abbrev_cell.font = Font(color=WHITE, bold=True)

        players = _team_players(rosters, team)
        for position in POSITION_ORDER:
            header_row, player_rows = POSITION_ROWS[position]
            header = sheet.cell(header_row, column, position)
            header.fill = PatternFill('solid', fgColor=LIGHT_BLUE)
            header.font = Font(color=NAVY, bold=True)
            header.alignment = Alignment(horizontal='center')
            header.border = CELL_BORDER
            position_players = [
                player
                for player in players
                if player.get('position') == position and not player.get('taxi')
            ]
            for row, player in zip(player_rows, position_players, strict=False):
                cell = sheet.cell(row, column, _format_player(player))
                cell.fill = PatternFill('solid', fgColor=PALE_SLATE)
                cell.border = CELL_BORDER
                cell.alignment = Alignment(wrap_text=True)

        taxi_header = sheet.cell(46, column, 'Taxi Squad')
        taxi_header.fill = PatternFill('solid', fgColor=GOLD)
        taxi_header.font = Font(color=NAVY, bold=True)
        taxi_header.alignment = Alignment(horizontal='center')
        taxi_header.border = CELL_BORDER
        taxi_players = [player for player in players if player.get('taxi')]
        for (position_row, player_row), player in zip(TAXI_ROWS, taxi_players, strict=False):
            position_cell = sheet.cell(position_row, column, player.get('position'))
            position_cell.fill = PatternFill('solid', fgColor=PALE_GOLD)
            position_cell.font = Font(color=NAVY, bold=True)
            position_cell.alignment = Alignment(horizontal='center')
            position_cell.border = CELL_BORDER
            player_cell = sheet.cell(player_row, column, _format_player(player))
            player_cell.fill = PatternFill('solid', fgColor=PALE_SLATE)
            player_cell.border = CELL_BORDER
            player_cell.alignment = Alignment(wrap_text=True)

    sheet.merge_cells('A57:S57')
    generated_cell = sheet['A57']
    generated_cell.value = f'Generated from data/rosters.json on {generated:%Y-%m-%d %H:%M UTC}'
    generated_cell.font = Font(color=SLATE, italic=True, size=9)
    generated_cell.alignment = Alignment(horizontal='right')
    sheet.print_area = 'A1:S57'
    sheet.page_setup.orientation = 'landscape'
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

    return _save_workbook(workbook)


def _draft_pick_map(picks_data: dict, season: int, draft_type: str) -> dict[tuple[int, str], dict]:
    if not isinstance(picks_data, dict) or not isinstance(picks_data.get('picks'), list):
        raise ValueError('Draft picks file is malformed')
    pick_map: dict[tuple[int, str], dict] = {}
    for pick in picks_data['picks']:
        if not isinstance(pick, dict):
            continue
        if str(pick.get('year')) != str(season):
            continue
        if (pick.get('draft_type') or 'offseason') != draft_type:
            continue
        try:
            key = (int(pick.get('round')), str(pick.get('original_team') or ''))
        except (TypeError, ValueError):
            raise ValueError('Draft picks file contains an invalid round') from None
        if key in pick_map:
            raise ValueError(f'Duplicate {draft_type} pick for round {key[0]}, {key[1]}')
        pick_map[key] = pick
    return pick_map


def _draft_order(orders_data: dict, season: int, draft_type: str) -> list[str]:
    if not isinstance(orders_data, dict):
        raise ValueError('Draft orders file is malformed')
    season_orders = orders_data.get(str(season))
    if not isinstance(season_orders, dict):
        raise ValueError(f'No draft order is configured for {season}')
    order = season_orders.get(draft_type)
    if not isinstance(order, list) or not order:
        raise ValueError(f'No {draft_type} order is configured for {season}')
    normalized = [str(team) for team in order]
    if len(normalized) != len(set(normalized)):
        raise ValueError(f'{season} {draft_type} order contains duplicate teams')
    unknown = sorted(set(normalized) - set(TEAM_ORDER))
    if unknown:
        raise ValueError(
            f'{season} {draft_type} order contains unknown team(s): {", ".join(unknown)}'
        )
    return normalized


def _owner_label(pick: dict) -> str:
    owner = str(pick.get('current_owner') or '').strip()
    original = str(pick.get('original_team') or '').strip()
    if not owner:
        raise ValueError(f'{pick.get("year")} R{pick.get("round")} {original} has no current owner')
    if owner == original:
        return owner
    previous = [str(team) for team in pick.get('previous_owners', []) if team]
    lineage = []
    for team in [original, *previous]:
        if team != owner and team not in lineage:
            lineage.append(team)
    return f'{owner} (via {" → ".join(lineage)})' if lineage else owner


def _round_count(pick_map: dict[tuple[int, str], dict]) -> int:
    if not pick_map:
        return 0
    rounds = sorted({round_number for round_number, _team in pick_map})
    expected = list(range(1, max(rounds) + 1))
    if rounds != expected:
        raise ValueError(f'Draft rounds are not contiguous: {rounds}')
    return max(rounds)


def _validate_draft_slots(
    pick_map: dict[tuple[int, str], dict],
    order: list[str],
    round_count: int,
    draft_type: str,
) -> None:
    extra_teams = sorted({team for _round, team in pick_map} - set(order))
    if extra_teams:
        raise ValueError(
            f'{draft_type} picks contain team(s) absent from the order: {", ".join(extra_teams)}'
        )
    invalid_owners = sorted(
        {
            str(pick.get('current_owner') or '')
            for pick in pick_map.values()
            if str(pick.get('current_owner') or '') not in TEAM_ORDER
        }
    )
    if invalid_owners:
        raise ValueError(
            f'{draft_type} picks contain invalid owner(s): {", ".join(invalid_owners)}'
        )
    for round_number in range(1, round_count + 1):
        missing = [team for team in order if (round_number, team) not in pick_map]
        if missing:
            raise ValueError(
                f'Missing {draft_type} round {round_number} pick(s): {", ".join(missing)}'
            )


def _write_round_block(
    sheet,
    start_row: int,
    start_column: int,
    round_number: int,
    label_prefix: str,
    order: list[str],
    pick_map: dict[tuple[int, str], dict],
) -> None:
    header_values = [f'{label_prefix}Round {round_number}', 'Team', 'Add', 'Drop']
    if label_prefix:
        header_values[-1] = None
    for offset, value in enumerate(header_values):
        cell = sheet.cell(start_row, start_column + offset, value)
        cell.fill = PatternFill('solid', fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal='center')

    for slot, original_team in enumerate(order, start=1):
        row = start_row + slot
        pick = pick_map[(round_number, original_team)]
        pick_cell = sheet.cell(row, start_column, slot)
        owner_cell = sheet.cell(row, start_column + 1, _owner_label(pick))
        add_cell = sheet.cell(row, start_column + 2)
        drop_cell = sheet.cell(row, start_column + 3)
        for cell in (pick_cell, owner_cell, add_cell, drop_cell):
            cell.border = CELL_BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.fill = PatternFill('solid', fgColor=WHITE if slot % 2 else PALE_SLATE)
        pick_cell.alignment = Alignment(horizontal='center', vertical='center')
        if pick.get('current_owner') != original_team:
            owner_cell.fill = PatternFill('solid', fgColor=PALE_GOLD)
            owner_cell.font = Font(color=NAVY, bold=True)
        if pick.get('condition'):
            owner_cell.fill = PatternFill('solid', fgColor=PALE_RED)
            owner_cell.font = Font(color=RED, bold=True)
            owner_cell.comment = Comment(
                f'Conditional claim: {pick.get("conditional_claim") or "not specified"}\n'
                f'Condition: {pick["condition"]}',
                'QPFL Commissioner Tools',
            )


def build_draft_board_workbook(
    picks_data: dict,
    orders_data: dict,
    teams_data: dict | None,
    season: int,
    generated_at: datetime | None = None,
) -> bytes:
    """Return an editable draft board whose ownership reflects all recorded trades."""
    generated = _timestamp(generated_at)
    regular_order = _draft_order(orders_data, season, 'offseason')
    taxi_order = _draft_order(orders_data, season, 'offseason_taxi')
    regular_picks = _draft_pick_map(picks_data, season, 'offseason')
    taxi_picks = _draft_pick_map(picks_data, season, 'offseason_taxi')
    regular_rounds = _round_count(regular_picks)
    taxi_rounds = _round_count(taxi_picks)
    if not regular_rounds or not taxi_rounds:
        raise ValueError(f'Draft picks are not configured for {season}')
    _validate_draft_slots(regular_picks, regular_order, regular_rounds, 'offseason')
    _validate_draft_slots(taxi_picks, taxi_order, taxi_rounds, 'offseason_taxi')

    workbook = Workbook()
    _set_workbook_properties(workbook, f'{season} QPFL Draft Board', generated)
    board = workbook.active
    board.title = f'{season} Offseason Draft'
    board.sheet_view.showGridLines = False
    board.merge_cells('A1:B1')
    board['A1'] = f'{season} NFL-Style Offseason Draft'
    board['A1'].font = Font(color=WHITE, bold=True, size=15)
    board['A1'].fill = PatternFill('solid', fgColor=NAVY)
    board['A1'].alignment = Alignment(horizontal='center')
    board['C1'] = f'Generated: {generated:%Y-%m-%d %H:%M UTC}'
    board['C1'].font = Font(color=SLATE, italic=True, size=9)

    for column in (1, 6, 11):
        board.column_dimensions[get_column_letter(column)].width = 15
        board.column_dimensions[get_column_letter(column + 1)].width = 30
        board.column_dimensions[get_column_letter(column + 2)].width = 27
        board.column_dimensions[get_column_letter(column + 3)].width = 24
    for column in (5, 10):
        board.column_dimensions[get_column_letter(column)].width = 3

    for round_number in range(1, regular_rounds + 1):
        group = (round_number - 1) // 3
        position = (round_number - 1) % 3
        start_row = 3 + group * 12
        start_column = 1 + position * 5
        _write_round_block(
            board,
            start_row,
            start_column,
            round_number,
            '',
            regular_order,
            regular_picks,
        )

    taxi_start = 3 + ((regular_rounds + 2) // 3) * 12
    for round_number in range(1, taxi_rounds + 1):
        group = (round_number - 1) // 2
        position = (round_number - 1) % 2
        start_row = taxi_start + group * 12
        start_column = 1 + position * 5
        _write_round_block(
            board,
            start_row,
            start_column,
            round_number,
            'TAXI ',
            taxi_order,
            taxi_picks,
        )

    board.freeze_panes = 'A3'
    board.print_area = f'A1:N{taxi_start + ((taxi_rounds - 1) // 2) * 12 + len(taxi_order)}'
    board.page_setup.orientation = 'landscape'
    board.page_setup.fitToWidth = 1
    board.sheet_properties.pageSetUpPr.fitToPage = True

    ledger = workbook.create_sheet('Pick Ledger')
    ledger.sheet_view.showGridLines = False
    ledger.freeze_panes = 'A2'
    headers = [
        'Draft',
        'Pick',
        'Round',
        'Slot',
        'Original Team',
        'Current Owner',
        'Previous Owners',
        'Conditional Claim',
        'Condition',
    ]
    ledger.append(headers)
    for cell in ledger[1]:
        cell.fill = PatternFill('solid', fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal='center')

    ledger_rows = []
    for label, order, pick_map, rounds in (
        ('Offseason', regular_order, regular_picks, regular_rounds),
        ('Taxi', taxi_order, taxi_picks, taxi_rounds),
    ):
        for round_number in range(1, rounds + 1):
            for slot, original_team in enumerate(order, start=1):
                pick = pick_map[(round_number, original_team)]
                ledger_rows.append(
                    [
                        label,
                        f'{round_number}.{slot:02d}',
                        round_number,
                        slot,
                        original_team,
                        pick.get('current_owner'),
                        ' → '.join(str(team) for team in pick.get('previous_owners', [])),
                        pick.get('conditional_claim'),
                        pick.get('condition'),
                    ]
                )
    for row_number, values in enumerate(ledger_rows, start=2):
        ledger.append(values)
        for cell in ledger[row_number]:
            cell.border = CELL_BORDER
            cell.fill = PatternFill('solid', fgColor=WHITE if row_number % 2 else PALE_SLATE)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        if values[4] != values[5]:
            ledger.cell(row_number, 6).fill = PatternFill('solid', fgColor=PALE_GOLD)
            ledger.cell(row_number, 6).font = Font(color=NAVY, bold=True)
        if values[8]:
            for column in (6, 8, 9):
                ledger.cell(row_number, column).fill = PatternFill('solid', fgColor=PALE_RED)

    widths = [14, 10, 9, 8, 16, 16, 32, 19, 58]
    for column, width in enumerate(widths, start=1):
        ledger.column_dimensions[get_column_letter(column)].width = width
    ledger.auto_filter.ref = f'A1:I{ledger.max_row}'

    directory = workbook.create_sheet('Team Directory')
    directory.sheet_view.showGridLines = False
    directory.append(['Abbreviation', 'Team Name', 'Owner'])
    for cell in directory[1]:
        cell.fill = PatternFill('solid', fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.border = CELL_BORDER
    metadata = _team_metadata(teams_data)
    for team in TEAM_ORDER:
        directory.append([team, metadata[team]['name'], metadata[team]['owner']])
    for row in directory.iter_rows(min_row=2):
        for cell in row:
            cell.border = CELL_BORDER
    directory.column_dimensions['A'].width = 16
    directory.column_dimensions['B'].width = 38
    directory.column_dimensions['C'].width = 34
    directory.freeze_panes = 'A2'

    return _save_workbook(workbook)
