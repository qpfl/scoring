"""Roster synchronization between JSON and Excel.

For 2026+, rosters.json is the source of truth. This module provides the
roster mutation helpers plus the one Excel writer in the codebase, which emits
the same grid layout scripts/init_rosters_from_excel.py reads so the two
round-trip.
"""

import json
from pathlib import Path

import openpyxl

from .constants import (
    ALL_TEAMS,
    POSITION_ORDER,
    POSITION_ROWS,
    ROSTER_SLOTS,
    TAXI_ROWS,
    TAXI_SLOTS,
    TEAM_COLUMNS,
    TEAM_TO_OWNER,
)


def load_rosters_json(rosters_path: str | Path) -> dict[str, list[dict]]:
    """Load rosters from JSON file."""
    rosters_path = Path(rosters_path)
    if not rosters_path.exists():
        return {}

    with open(rosters_path) as f:
        return json.load(f)  # type: ignore[no-any-return]


def save_rosters_json(rosters_path: str | Path, rosters: dict[str, list[dict]]) -> None:
    """Save rosters to JSON file."""
    rosters_path = Path(rosters_path)
    rosters_path.parent.mkdir(parents=True, exist_ok=True)

    with open(rosters_path, 'w') as f:
        json.dump(rosters, f, indent=2)


def format_player_for_excel(player: dict) -> str:
    """Format a player dict as Excel cell value 'Name (TEAM)'."""
    name: str = str(player.get('name', ''))
    nfl_team: str = str(player.get('nfl_team', ''))
    if nfl_team:
        return f'{name} ({nfl_team})'
    return name


def load_team_metadata(teams_path: str | Path | None) -> dict[str, dict[str, str]]:
    """Load {abbrev: {'name': ..., 'owner': ...}} from data/teams.json.

    Falls back to TEAM_TO_OWNER (and the abbrev itself as the team name) for
    any team missing from the file, so the export still works if teams.json
    is absent or incomplete.
    """
    metadata: dict[str, dict[str, str]] = {}

    if teams_path:
        teams_path = Path(teams_path)
        if teams_path.exists():
            with open(teams_path) as f:
                data = json.load(f)
            for team in data.get('teams', []):
                abbrev = str(team.get('abbrev', '')).strip()
                if abbrev:
                    metadata[abbrev] = {
                        'name': str(team.get('name') or abbrev),
                        'owner': str(team.get('owner') or TEAM_TO_OWNER.get(abbrev, '')),
                    }

    for abbrev in ALL_TEAMS:
        metadata.setdefault(abbrev, {'name': abbrev, 'owner': TEAM_TO_OWNER.get(abbrev, '')})

    return metadata


def sync_rosters_to_excel(
    rosters_json_path: str | Path,
    excel_path: str | Path,
    sheet_name: str = 'Rosters',
    teams_path: str | Path | None = None,
) -> bool:
    """Write rosters.json out as a fresh Excel workbook in the QPFL grid layout.

    This is the layout scripts/init_rosters_from_excel.py reads, so the output
    round-trips back to identical JSON:

    - Teams occupy the columns in TEAM_COLUMNS, in ALL_TEAMS order
    - Row 2 = team name, row 3 = owner, row 4 = abbreviation
    - Active players sit at the rows named in POSITION_ROWS, under a header
      cell holding the position label
    - Taxi players sit in TAXI_ROWS as (position label row, player row) pairs

    Only player names are written - no scores, formulas, or formatting. Any
    existing file at excel_path is replaced.

    Args:
        rosters_json_path: Path to rosters.json
        excel_path: Path to the workbook to write (overwritten if it exists)
        sheet_name: Sheet name for the roster grid
        teams_path: Optional path to data/teams.json for team names/owners

    Returns:
        True if the workbook was written
    """
    rosters = load_rosters_json(rosters_json_path)
    if not rosters:
        print('No rosters to sync')
        return False

    excel_path = Path(excel_path)
    teams = load_team_metadata(teams_path)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    written = 0
    skipped = 0

    for col, team_abbrev in zip(TEAM_COLUMNS, ALL_TEAMS, strict=True):
        team = teams[team_abbrev]
        ws.cell(row=2, column=col, value=team['name'])
        ws.cell(row=3, column=col, value=team['owner'])
        ws.cell(row=4, column=col, value=team_abbrev)

        roster = rosters.get(team_abbrev, [])

        # Active roster: position header cell, then one player per slot row.
        for position in POSITION_ORDER:
            header_row, player_rows = POSITION_ROWS[position]
            ws.cell(row=header_row, column=col, value=position)

            players = [p for p in roster if p.get('position') == position and not p.get('taxi')]
            if len(players) > ROSTER_SLOTS[position]:
                print(
                    f'  WARNING: {team_abbrev} has {len(players)} {position} '
                    f'(max {ROSTER_SLOTS[position]}) - '
                    f'{len(players) - ROSTER_SLOTS[position]} not written'
                )

            for row, player in zip(player_rows, players, strict=False):
                ws.cell(row=row, column=col, value=format_player_for_excel(player))

            fit = min(len(players), len(player_rows))
            written += fit
            skipped += len(players) - fit

        # Taxi squad: the position label lives above the player, since taxi
        # slots aren't grouped by position like the active rows are.
        taxi = [p for p in roster if p.get('taxi')]
        if len(taxi) > TAXI_SLOTS:
            print(
                f'  WARNING: {team_abbrev} has {len(taxi)} taxi players '
                f'(max {TAXI_SLOTS}) - {len(taxi) - TAXI_SLOTS} not written'
            )

        taxi_position_counts: dict[str, int] = {}
        for player in taxi:
            position = str(player.get('position', ''))
            taxi_position_counts[position] = taxi_position_counts.get(position, 0) + 1
        for position, count in taxi_position_counts.items():
            if count > 1:
                print(f'  WARNING: {team_abbrev} has {count} taxi {position} (max 1 per position)')

        for (pos_row, player_row), player in zip(TAXI_ROWS, taxi, strict=False):
            ws.cell(row=pos_row, column=col, value=player.get('position', ''))
            ws.cell(row=player_row, column=col, value=format_player_for_excel(player))

        fit = min(len(taxi), len(TAXI_ROWS))
        written += fit
        skipped += len(taxi) - fit

    wb.save(str(excel_path))
    wb.close()

    summary = f'Wrote {written} players to {excel_path}'
    if skipped:
        summary += f' ({skipped} over capacity, not written)'
    print(summary)
    return True


def add_player_to_roster(
    rosters: dict[str, list[dict]],
    team_abbrev: str,
    player: dict,
    is_taxi: bool = False,
) -> dict[str, list[dict]]:
    """Add a player to a team's roster.

    Args:
        rosters: Full rosters dict
        team_abbrev: Team to add player to
        player: Player dict with name, nfl_team, position
        is_taxi: Whether to add to taxi squad

    Returns:
        Updated rosters dict
    """
    if team_abbrev not in rosters:
        rosters[team_abbrev] = []

    new_player = {
        'name': player['name'],
        'nfl_team': player['nfl_team'],
        'position': player['position'],
    }
    if is_taxi:
        new_player['taxi'] = True

    rosters[team_abbrev].append(new_player)
    return rosters


def remove_player_from_roster(
    rosters: dict[str, list[dict]],
    team_abbrev: str,
    player_name: str,
) -> tuple[dict[str, list[dict]], dict | None]:
    """Remove a player from a team's roster.

    Args:
        rosters: Full rosters dict
        team_abbrev: Team to remove player from
        player_name: Name of player to remove

    Returns:
        Tuple of (updated rosters dict, removed player dict or None)
    """
    if team_abbrev not in rosters:
        return rosters, None

    team_roster = rosters[team_abbrev]
    removed_player = None

    for i, player in enumerate(team_roster):
        if player.get('name') == player_name:
            removed_player = team_roster.pop(i)
            break

    return rosters, removed_player


def trade_players(
    rosters: dict[str, list[dict]],
    team1: str,
    team2: str,
    team1_gives: list[str],
    team2_gives: list[str],
) -> dict[str, list[dict]]:
    """Execute a trade between two teams.

    Args:
        rosters: Full rosters dict
        team1: First team abbreviation
        team2: Second team abbreviation
        team1_gives: List of player names team1 is giving
        team2_gives: List of player names team2 is giving

    Returns:
        Updated rosters dict
    """
    # Remove players from each team and collect them
    players_to_team2 = []
    for player_name in team1_gives:
        rosters, player = remove_player_from_roster(rosters, team1, player_name)
        if player:
            players_to_team2.append(player)

    players_to_team1 = []
    for player_name in team2_gives:
        rosters, player = remove_player_from_roster(rosters, team2, player_name)
        if player:
            players_to_team1.append(player)

    # Add players to new teams
    for player in players_to_team2:
        is_taxi = player.get('taxi', False)
        rosters = add_player_to_roster(rosters, team2, player, is_taxi)

    for player in players_to_team1:
        is_taxi = player.get('taxi', False)
        rosters = add_player_to_roster(rosters, team1, player, is_taxi)

    return rosters


def sync_pick_trade_to_json(
    draft_picks_path: str | Path,
    from_team: str,
    to_team: str,
    season: str,
    round_num: int,
    pick_type: str = 'offseason',
) -> bool:
    """Record a traded pick in the draft_picks.json file.

    Args:
        draft_picks_path: Path to data/draft_picks.json
        from_team: Team giving away the pick
        to_team: Team receiving the pick
        season: Season year as string (e.g., "2026")
        round_num: Round number
        pick_type: Type of pick (offseason, waiver, offseason_taxi, waiver_taxi)

    Returns:
        True if sync was successful
    """
    draft_picks_path = Path(draft_picks_path)

    if not draft_picks_path.exists():
        print(f'Warning: Draft picks file not found: {draft_picks_path}')
        return False

    with open(draft_picks_path) as f:
        data = json.load(f)

    picks = data.get('picks', {})

    # Ensure teams exist in picks
    if from_team not in picks:
        picks[from_team] = {}
    if to_team not in picks:
        picks[to_team] = {}

    # Ensure season exists for both teams
    if season not in picks[from_team]:
        picks[from_team][season] = {}
    if season not in picks[to_team]:
        picks[to_team][season] = {}

    # Ensure pick_type exists for both teams
    if pick_type not in picks[from_team][season]:
        picks[from_team][season][pick_type] = []
    if pick_type not in picks[to_team][season]:
        picks[to_team][season][pick_type] = []

    # Remove the pick from from_team
    from_picks = picks[from_team][season][pick_type]
    pick_index = None
    for i, p in enumerate(from_picks):
        # The pick might be from any original owner
        if p.get('round') == round_num:
            pick_index = i
            break

    if pick_index is not None:
        removed_pick = from_picks.pop(pick_index)
        original_owner = removed_pick.get('from', from_team)
    else:
        # Pick not found in from_team's list - they might be trading their own
        original_owner = from_team

    # Add the pick to to_team
    picks[to_team][season][pick_type].append(
        {'round': round_num, 'from': original_owner, 'own': original_owner == to_team}
    )

    # Save back
    data['picks'] = picks
    with open(draft_picks_path, 'w') as f:
        json.dump(data, f, indent=2)

    print(
        f'Pick trade recorded in JSON: {from_team} -> {to_team}, {season} Round {round_num} ({pick_type})'
    )
    return True
