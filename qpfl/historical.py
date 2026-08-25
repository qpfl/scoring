"""Read the official weekly records from the retired Excel scoring workbooks."""

import re
from pathlib import Path
from typing import Any

import openpyxl

from .constants import POSITION_ROWS

WEEK_WORDS = {
    'one': 1,
    'two': 2,
    'three': 3,
    'four': 4,
    'five': 5,
    'six': 6,
    'seven': 7,
    'eight': 8,
    'nine': 9,
    'ten': 10,
    'eleven': 11,
    'twelve': 12,
    'thirteen': 13,
    'fourteen': 14,
    'fifteen': 15,
}

TEAM_CODE_ALIASES = {
    'SPY': 'AYP',
    'T/S': 'S/T',
}

HISTORICAL_POSITION_ROWS = {
    2020: {
        'QB': (7, [8, 9, 10]),
        'RB': (12, [13, 14, 15, 16]),
        'WR': (18, [19, 20, 21, 22]),
        'TE': (24, [25, 26, 27]),
        'K': (29, [30, 31]),
        'D/ST': (33, [34, 35]),
        'HC': (37, [38, 39]),
    },
    2021: {
        'QB': (7, [8, 9, 10]),
        'RB': (12, [13, 14, 15, 16]),
        'WR': (18, [19, 20, 21, 22]),
        'TE': (24, [25, 26, 27]),
        'K': (29, [30, 31]),
        'D/ST': (33, [34, 35]),
        'HC': (37, [38, 39]),
    },
    2022: {
        'QB': (6, [7, 8, 9]),
        'RB': (11, [12, 13, 14, 15]),
        'WR': (17, [18, 19, 20, 21]),
        'TE': (23, [24, 25, 26]),
        'K': (28, [29, 30]),
        'D/ST': (32, [33, 34]),
        'HC': (36, [37, 38]),
    },
    2023: {
        'QB': (6, [7, 8, 9]),
        'RB': (11, [12, 13, 14, 15]),
        'WR': (17, [18, 19, 20, 21]),
        'TE': (23, [24, 25, 26]),
        'K': (28, [29, 30]),
        'D/ST': (32, [33, 34]),
        'HC': (36, [37, 38]),
    },
    2024: {
        'QB': (6, [7, 8, 9]),
        'RB': (11, [12, 13, 14, 15]),
        'WR': (17, [18, 19, 20, 21]),
        'TE': (23, [24, 25, 26]),
        'K': (28, [29, 30]),
        'D/ST': (32, [33, 34]),
        'HC': (36, [37, 38]),
        'OL': (40, [41, 42]),
    },
}

HISTORICAL_TAXI_ROWS = {
    2020: [(43, 44), (45, 46), (47, 48), (49, 50)],
    2021: [(42, 43), (44, 45), (46, 47), (48, 49)],
    2022: [(43, 44), (45, 46), (47, 48), (49, 50)],
    2023: [(43, 44), (45, 46), (47, 48), (49, 50)],
    2024: [(47, 48), (49, 50), (51, 52), (53, 54)],
}

SCORE_ROWS = {
    2022: 40,
    2023: 40,
    2024: 44,
    2025: 45,
}

PLAYER_LABEL_RE = re.compile(r'^(?P<name>.+?)\s*\((?P<team>[A-Z0-9]{2,3})\)$')
PLAYOFF_SEED_RE = re.compile(r'^\(\d+\)\s*')


def normalize_historical_team_code(value: Any) -> str:
    code = str(value or '').strip()
    return TEAM_CODE_ALIASES.get(code, code)


def strip_playoff_seed(name: str) -> str:
    """Remove a playoff seed prefix such as ``(3) `` from a team name."""
    return PLAYOFF_SEED_RE.sub('', name)


def historical_team_columns(season: int) -> list[int]:
    team_count = 8 if season <= 2021 else 10
    return [1 + 2 * index for index in range(team_count)]


def historical_team_info_rows(season: int) -> tuple[int, int, int]:
    if season <= 2021:
        return 3, 4, 5
    return 2, 3, 4


def position_rows_for_season(season: int | None) -> dict[str, tuple[int, list[int]]]:
    if season in HISTORICAL_POSITION_ROWS:
        return HISTORICAL_POSITION_ROWS[season]
    return POSITION_ROWS


def taxi_rows_for_season(season: int | None) -> list[tuple[int, int]]:
    if season in HISTORICAL_TAXI_ROWS:
        return HISTORICAL_TAXI_ROWS[season]
    return [(48, 49), (50, 51), (52, 53), (54, 55)]


def historical_week_sheets(workbook, season: int) -> list[tuple[int, str]]:
    weeks: dict[int, tuple[int, str]] = {}

    for sheet_name in workbook.sheetnames:
        week = None
        priority = 1

        if sheet_name in {'Playoffs, Week 1', 'Playoffs, Round 1'}:
            week = 15
        elif sheet_name in {'Semi-Finals', 'Super Bowl Week'}:
            week = 16
        elif sheet_name in {'Championship', 'Championship Week'}:
            week = 17
            priority = 2
        elif sheet_name == 'Championship Week 2.0':
            week = 17
            priority = 3
        else:
            numeric_match = re.fullmatch(r'Week (\d+)', sheet_name)
            written_match = re.fullmatch(r'Week (\w+)', sheet_name, re.IGNORECASE)
            embedded_match = re.search(r'Week (\d+)', sheet_name)

            if numeric_match:
                week = int(numeric_match.group(1))
            elif written_match and written_match.group(1).lower() in WEEK_WORDS:
                week = WEEK_WORDS[written_match.group(1).lower()]
            elif embedded_match:
                week = int(embedded_match.group(1))

        if week is None:
            continue
        if week not in weeks or priority >= weeks[week][0]:
            weeks[week] = (priority, sheet_name)

    return [(week, weeks[week][1]) for week in sorted(weeks)]


def parse_historical_player_label(value: Any) -> tuple[str, str]:
    label = str(value or '').strip()
    match = PLAYER_LABEL_RE.match(label)
    if match:
        return match.group('name').strip(), match.group('team')
    return label, ''


def _is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def official_team_score(worksheet, column: int, season: int) -> float:
    if season <= 2021:
        for row in (40, 41):
            value = worksheet.cell(row=row, column=column + 1).value
            if _is_number(value):
                return float(value)
        score_row = 42
    else:
        score_row = SCORE_ROWS[season]
        value = worksheet.cell(row=score_row, column=column + 1).value
        if _is_number(value):
            return float(value)

    total = 0.0
    for _position, (_header_row, player_rows) in position_rows_for_season(season).items():
        for row in player_rows:
            player_cell = worksheet.cell(row=row, column=column)
            score = worksheet.cell(row=row, column=column + 1).value
            if player_cell.value and player_cell.font.bold and _is_number(score):
                total += float(score)

    if score_row:
        return total
    raise ValueError(f'Could not read an official score from column {column}')


def load_historical_workbook(path: str | Path, season: int) -> dict[int, dict[str, Any]]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    name_row, owner_row, abbrev_row = historical_team_info_rows(season)
    weeks: dict[int, dict[str, Any]] = {}

    for week, sheet_name in historical_week_sheets(workbook, season):
        worksheet = workbook[sheet_name]
        teams: dict[str, dict[str, Any]] = {}

        for column in historical_team_columns(season):
            abbrev = normalize_historical_team_code(worksheet.cell(abbrev_row, column).value)
            if not abbrev:
                continue

            roster = []
            for position, (_header_row, player_rows) in position_rows_for_season(season).items():
                for row in player_rows:
                    player_cell = worksheet.cell(row=row, column=column)
                    name, nfl_team = parse_historical_player_label(player_cell.value)
                    if not name or name.lower() == 'v':
                        continue
                    score = worksheet.cell(row=row, column=column + 1).value
                    roster.append(
                        {
                            'name': name,
                            'nfl_team': nfl_team,
                            'position': position,
                            'score': float(score) if _is_number(score) else None,
                            'starter': bool(player_cell.font.bold),
                        }
                    )

            teams[abbrev] = {
                'name': strip_playoff_seed(
                    str(worksheet.cell(name_row, column).value or '').strip().strip('*')
                ),
                'owner': str(worksheet.cell(owner_row, column).value or '').strip(),
                'abbrev': abbrev,
                'roster': roster,
                'total_score': official_team_score(worksheet, column, season),
            }

        scores = [team['total_score'] for team in teams.values()]
        for team in teams.values():
            team['score_rank'] = 1 + sum(score > team['total_score'] for score in scores)

        weeks[week] = {'sheet_name': sheet_name, 'teams': teams}

    workbook.close()
    return weeks


def load_official_scores(path: str | Path, season: int) -> dict[tuple[int, str], float]:
    weeks = load_historical_workbook(path, season)
    return {
        (week, abbrev): team['total_score']
        for week, week_data in weeks.items()
        for abbrev, team in week_data['teams'].items()
    }
