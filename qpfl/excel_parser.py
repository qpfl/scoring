"""Excel roster parsing utilities."""

import re

import openpyxl

from .constants import POSITION_ROWS, TEAM_COLUMNS
from .models import FantasyTeam


def parse_player_name(cell_value: str) -> tuple[str, str]:
    """
    Parse player name from Excel format "Player Name (TEAM)" to (name, team_abbrev).

    Examples:
        "Patrick Mahomes II (KC)" -> ("Patrick Mahomes II", "KC")
        "San Francisco 49ers (SF)" -> ("San Francisco 49ers", "SF")
    """
    if not cell_value:
        return '', ''

    match = re.match(r'^(.+?)\s*\(([A-Z]{2,3})\)$', cell_value.strip())
    if match:
        return match.group(1).strip(), match.group(2)
    return cell_value.strip(), ''


def parse_roster_from_excel(filepath: str, sheet_name: str = 'Week 13') -> list[FantasyTeam]:
    """
    Parse fantasy team rosters from Excel file.

    Args:
        filepath: Path to the Excel file
        sheet_name: Name of the sheet to read

    Returns:
        List of FantasyTeam objects
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb[sheet_name]

    teams = []

    # Parse team headers (rows 2-4)
    for col in TEAM_COLUMNS:
        team_name_cell = ws.cell(row=2, column=col)
        team_name = team_name_cell.value or ''
        team_name = team_name.strip().strip('*')  # Remove bold markers

        owner = ws.cell(row=3, column=col).value or ''
        abbrev = ws.cell(row=4, column=col).value or ''

        if team_name:
            team = FantasyTeam(
                name=team_name,
                owner=owner,
                abbreviation=abbrev,
                column_index=col,
                players={},
            )
            teams.append(team)

    # Parse players for each team
    for team in teams:
        col = team.column_index

        for position, (_header_row, player_rows) in POSITION_ROWS.items():
            team.players[position] = []

            for row in player_rows:
                cell = ws.cell(row=row, column=col)
                cell_value = cell.value

                if cell_value:
                    is_bold = cell.font.bold if cell.font else False
                    player_name, nfl_team = parse_player_name(str(cell_value))

                    if player_name:
                        team.players[position].append((player_name, nfl_team, is_bold))

    wb.close()
    return teams


def update_excel_scores(
    excel_path: str,
    sheet_name: str,
    teams: list[FantasyTeam],
    results: dict,
):
    """
    Update the Excel file with calculated scores for STARTERS ONLY.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Sheet to update
        teams: List of FantasyTeam objects
        results: Dict mapping team name to (total_score, position_scores)
                position_scores is Dict[position, List[(PlayerScore, is_starter)]]
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name]

    # Get player rows for each position
    position_player_rows = {pos: rows for pos, (_, rows) in POSITION_ROWS.items()}

    for team in teams:
        if team.name not in results:
            continue

        total, scores = results[team.name]
        points_col = team.column_index + 1

        for position, player_rows in position_player_rows.items():
            if position not in scores:
                continue

            # Only process STARTERS
            for player_name, _nfl_team, is_started in team.players.get(position, []):
                if not is_started:
                    continue

                # Find the score for this player
                player_score = None
                for ps, _ in scores[position]:  # scores is now List[(PlayerScore, is_starter)]
                    if ps.name == player_name:
                        player_score = ps
                        break

                if player_score is None:
                    continue

                # Find the row for this player
                for row in player_rows:
                    cell = ws.cell(row=row, column=team.column_index)
                    if cell.value:
                        parsed_name, _ = parse_player_name(str(cell.value))
                        if parsed_name == player_name:
                            score_cell = ws.cell(row=row, column=points_col)
                            score_cell.value = player_score.total_points
                            break

    wb.save(excel_path)
    print(f'\nScores saved to {excel_path}')
