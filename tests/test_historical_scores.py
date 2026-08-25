import json
from pathlib import Path

import openpyxl
import pytest

from qpfl.historical import load_historical_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
COMPLETED_SEASONS = range(2020, 2026)
CHAMPIONSHIP_SCORE_ANCHORS = {
    2020: {'GSA': 75.0, 'CWR': 78.0},
    2021: {'RPA': 105.0, 'GSA': 106.0},
    2022: {'GSA': 86.0, 'CGK': 79.0},
    2023: {'CGK': 84.0, 'S/T': 100.0},
    2024: {'GSA': 110.0, 'CGK': 74.0},
    2025: {'CGK': 102.0, 'S/T': 92.0},
}
TEAM_CODE_ALIASES = {'SPY': 'AYP', 'T/S': 'S/T'}
SCORE_ROWS = {
    2020: (40, 41),
    2021: (40, 41),
    2022: (40,),
    2023: (40,),
    2024: (44,),
    2025: (45,),
}
ABBREV_ROWS = {2020: 5, 2021: 5, 2022: 4, 2023: 4, 2024: 4, 2025: 4}
STARTER_ROWS_2025 = [
    7,
    8,
    9,
    12,
    13,
    14,
    15,
    18,
    19,
    20,
    21,
    22,
    25,
    26,
    27,
    30,
    31,
    34,
    35,
    38,
    39,
    42,
    43,
]


def load_json(path: Path) -> dict:
    with open(path) as file:
        return json.load(file)


def official_sheet_names(season: int) -> dict[int, str]:
    if season <= 2021:
        playoff = 'Playoffs, Week 1' if season == 2020 else 'Playoffs, Round 1'
        return {
            **{week: f'Week {week}' for week in range(1, 15)},
            15: playoff,
            16: 'Super Bowl Week',
        }
    if season == 2022:
        written_weeks = {
            1: 'Week One',
            2: 'Week Two',
            3: 'Week Three',
            4: 'Week Four',
            5: 'Rivalry Week 5',
        }
        return {
            **written_weeks,
            **{week: f'Week {week}' for week in range(6, 16)},
            16: 'Semi-Finals',
            17: 'Championship Week 2.0',
        }
    if season <= 2024:
        return {
            **{week: f'Week {week}' for week in range(1, 16)},
            16: 'Semi-Finals',
            17: 'Championship',
        }
    return {week: f'Week {week}' for week in range(1, 18)}


def read_official_scores(season: int) -> dict[int, dict[str, float]]:
    workbook = openpyxl.load_workbook(
        PROJECT_ROOT / 'previous_seasons' / f'{season} Scores.xlsx', data_only=True
    )
    official = {}

    for week, sheet_name in official_sheet_names(season).items():
        worksheet = workbook[sheet_name]
        scores = {}
        team_count = 8 if season <= 2021 else 10
        for column in range(1, team_count * 2, 2):
            abbrev = str(worksheet.cell(ABBREV_ROWS[season], column).value or '').strip()
            abbrev = TEAM_CODE_ALIASES.get(abbrev, abbrev)
            if not abbrev:
                continue
            score = None
            for row in SCORE_ROWS[season]:
                value = worksheet.cell(row, column + 1).value
                if isinstance(value, (int, float)):
                    score = value
                    break
            if score is None:
                score = sum(
                    float(worksheet.cell(row, column + 1).value or 0)
                    for row in STARTER_ROWS_2025
                    if worksheet.cell(row, column).value and worksheet.cell(row, column).font.bold
                )
            scores[abbrev] = float(score)
        official[week] = scores

    workbook.close()
    return official


@pytest.fixture(scope='module')
def official_seasons():
    return {season: read_official_scores(season) for season in COMPLETED_SEASONS}


@pytest.mark.parametrize('season', COMPLETED_SEASONS)
def test_displayed_scores_match_official_workbooks(official_seasons, season):
    official = official_seasons[season]
    weeks_dir = PROJECT_ROOT / 'web' / 'data' / 'seasons' / str(season) / 'weeks'
    displayed = {
        int(path.stem.split('_')[1]): load_json(path) for path in weeks_dir.glob('week_*.json')
    }

    assert set(displayed) == set(official)
    for week, official_scores in official.items():
        displayed_teams = {team['abbrev']: team for team in displayed[week]['teams']}
        assert set(displayed_teams) == set(official_scores)
        ranks = {
            abbrev: 1 + sum(other > score for other in official_scores.values())
            for abbrev, score in official_scores.items()
        }

        for abbrev, official_score in official_scores.items():
            team = displayed_teams[abbrev]
            assert team['total_score'] == official_score, (
                f'{season} Week {week} {abbrev} does not match the official workbook'
            )
            assert team['score_rank'] == ranks[abbrev]
            assert (
                sum(
                    float(player.get('score') or 0.0)
                    for player in team.get('roster', [])
                    if player.get('starter')
                )
                == official_score
            )

        for matchup in displayed[week]['matchups']:
            for key in ('team1', 'team2'):
                team = matchup[key]
                assert team['total_score'] == displayed_teams[team['abbrev']]['total_score']
                assert team['score_rank'] == displayed_teams[team['abbrev']]['score_rank']


@pytest.mark.parametrize('season', COMPLETED_SEASONS)
def test_legacy_and_displayed_historical_scores_cannot_diverge(season):
    legacy = load_json(PROJECT_ROOT / 'web' / f'data_{season}.json')
    legacy_scores = {
        (week['week'], team['abbrev']): (team['total_score'], team['score_rank'])
        for week in legacy['weeks']
        for team in week['teams']
    }
    split_scores = {}
    weeks_dir = PROJECT_ROOT / 'web' / 'data' / 'seasons' / str(season) / 'weeks'
    for path in weeks_dir.glob('week_*.json'):
        week = load_json(path)
        split_scores.update(
            {
                (week['week'], team['abbrev']): (team['total_score'], team['score_rank'])
                for team in week['teams']
            }
        )

    assert split_scores == legacy_scores


def test_2024_displayed_rosters_match_the_workbook_layout():
    official = load_historical_workbook(
        PROJECT_ROOT / 'previous_seasons' / '2024 Scores.xlsx', 2024
    )
    for week, source_week in official.items():
        displayed = load_json(
            PROJECT_ROOT / 'web' / 'data' / 'seasons' / '2024' / 'weeks' / f'week_{week}.json'
        )
        displayed_teams = {team['abbrev']: team for team in displayed['teams']}

        for abbrev, source_team in source_week['teams'].items():
            source_positions = [player['position'] for player in source_team['roster']]
            displayed_positions = [
                player['position'] for player in displayed_teams[abbrev]['roster']
            ]
            assert displayed_positions == source_positions


def test_official_workbooks_are_anchored_to_known_championship_scores(official_seasons):
    for season, expected in CHAMPIONSHIP_SCORE_ANCHORS.items():
        championship_week = max(official_seasons[season])
        scores = official_seasons[season][championship_week]
        assert {abbrev: scores[abbrev] for abbrev in expected} == expected
