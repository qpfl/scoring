"""Tests for exporting data/rosters.json back out to the QPFL Excel grid.

The important property is that the writer in qpfl/roster_sync.py and the reader
in scripts/init_rosters_from_excel.py agree on the layout, so a roster survives
a JSON -> Excel -> JSON round trip.
"""

import json

import openpyxl

from qpfl.roster_sync import sync_rosters_to_excel
from scripts.init_rosters_from_excel import init_rosters_from_excel


def _write_rosters(path, rosters):
    path.write_text(json.dumps(rosters))
    return path


def _write_teams(path, teams):
    path.write_text(json.dumps({'teams': teams}))
    return path


def _normalize(rosters):
    """Compare on content only - the reader adds status and orders by position."""
    return {
        team: sorted(
            (p['name'], p.get('position'), p.get('nfl_team'), bool(p.get('taxi'))) for p in players
        )
        for team, players in rosters.items()
        if players
    }


def test_round_trip_preserves_rosters(tmp_path):
    rosters = {
        'GSA': [
            {'name': 'Patrick Mahomes', 'nfl_team': 'KC', 'position': 'QB'},
            {'name': 'Bijan Robinson', 'nfl_team': 'ATL', 'position': 'RB'},
            {'name': 'Justin Jefferson', 'nfl_team': 'MIN', 'position': 'WR'},
            {'name': 'Brock Bowers', 'nfl_team': 'LV', 'position': 'TE'},
            {'name': 'Harrison Butker', 'nfl_team': 'KC', 'position': 'K'},
            {'name': 'Denver Broncos', 'nfl_team': 'DEN', 'position': 'D/ST'},
            {'name': 'Andy Reid', 'nfl_team': 'KC', 'position': 'HC'},
            {'name': 'Philadelphia Eagles', 'nfl_team': 'PHI', 'position': 'OL'},
            {'name': 'Taxi Guy', 'nfl_team': 'BUF', 'position': 'WR', 'taxi': True},
        ],
        'CGK': [
            {'name': 'Josh Allen', 'nfl_team': 'BUF', 'position': 'QB'},
            {'name': 'Saquon Barkley', 'nfl_team': 'PHI', 'position': 'RB'},
        ],
    }
    rosters_path = _write_rosters(tmp_path / 'rosters.json', rosters)
    excel_path = tmp_path / 'Rosters_current.xlsx'
    output_path = tmp_path / 'roundtrip.json'

    assert sync_rosters_to_excel(rosters_path, excel_path) is True
    assert init_rosters_from_excel(excel_path, output_path) is True

    assert _normalize(json.loads(output_path.read_text())) == _normalize(rosters)


def test_taxi_position_label_round_trips(tmp_path):
    """Taxi position comes from the label cell above the player, not a position block."""
    rosters = {
        'GSA': [
            {'name': 'Taxi One', 'nfl_team': 'BUF', 'position': 'WR', 'taxi': True},
            {'name': 'Taxi Two', 'nfl_team': 'MIA', 'position': 'RB', 'taxi': True},
        ]
    }
    rosters_path = _write_rosters(tmp_path / 'rosters.json', rosters)
    excel_path = tmp_path / 'Rosters_current.xlsx'
    output_path = tmp_path / 'roundtrip.json'

    sync_rosters_to_excel(rosters_path, excel_path)
    init_rosters_from_excel(excel_path, output_path)

    parsed = {p['name']: p for p in json.loads(output_path.read_text())['GSA']}
    assert parsed['Taxi One']['position'] == 'WR'
    assert parsed['Taxi One']['taxi'] is True
    assert parsed['Taxi Two']['position'] == 'RB'
    assert parsed['Taxi Two']['nfl_team'] == 'MIA'


def test_team_name_owner_and_abbrev_headers(tmp_path):
    rosters_path = _write_rosters(
        tmp_path / 'rosters.json',
        {'GSA': [{'name': 'Patrick Mahomes', 'nfl_team': 'KC', 'position': 'QB'}]},
    )
    teams_path = _write_teams(
        tmp_path / 'teams.json',
        [{'abbrev': 'GSA', 'name': 'No Kings Except Henry', 'owner': 'Griffin Ansel'}],
    )
    excel_path = tmp_path / 'Rosters_current.xlsx'

    sync_rosters_to_excel(rosters_path, excel_path, teams_path=teams_path)

    ws = openpyxl.load_workbook(excel_path).active
    # GSA is first in ALL_TEAMS, so it lands in column 1.
    assert ws.cell(row=2, column=1).value == 'No Kings Except Henry'
    assert ws.cell(row=3, column=1).value == 'Griffin Ansel'
    assert ws.cell(row=4, column=1).value == 'GSA'
    assert ws.cell(row=6, column=1).value == 'QB'
    assert ws.cell(row=7, column=1).value == 'Patrick Mahomes (KC)'


def test_missing_teams_json_falls_back_to_constants(tmp_path):
    rosters_path = _write_rosters(
        tmp_path / 'rosters.json',
        {'GSA': [{'name': 'Patrick Mahomes', 'nfl_team': 'KC', 'position': 'QB'}]},
    )
    excel_path = tmp_path / 'Rosters_current.xlsx'

    assert sync_rosters_to_excel(rosters_path, excel_path, teams_path=tmp_path / 'nope.json')

    ws = openpyxl.load_workbook(excel_path).active
    assert ws.cell(row=4, column=1).value == 'GSA'
    assert ws.cell(row=3, column=1).value == 'Griffin'  # TEAM_TO_OWNER fallback


def test_over_capacity_position_warns(tmp_path, capsys):
    """ROSTER_SLOTS['QB'] is 3 - a fourth QB has nowhere to go."""
    rosters_path = _write_rosters(
        tmp_path / 'rosters.json',
        {'GSA': [{'name': f'QB {i}', 'nfl_team': 'KC', 'position': 'QB'} for i in range(4)]},
    )

    sync_rosters_to_excel(rosters_path, tmp_path / 'Rosters_current.xlsx')

    out = capsys.readouterr().out
    assert 'WARNING' in out
    assert 'GSA has 4 QB' in out
    assert 'over capacity' in out


def test_too_many_taxi_players_warns(tmp_path, capsys):
    rosters_path = _write_rosters(
        tmp_path / 'rosters.json',
        {
            'GSA': [
                {'name': f'Taxi {i}', 'nfl_team': 'KC', 'position': 'WR', 'taxi': True}
                for i in range(5)
            ]
        },
    )

    sync_rosters_to_excel(rosters_path, tmp_path / 'Rosters_current.xlsx')

    out = capsys.readouterr().out
    assert 'GSA has 5 taxi players' in out
    assert 'max 1 per position' in out


def test_missing_rosters_json_writes_nothing(tmp_path):
    excel_path = tmp_path / 'Rosters_current.xlsx'

    assert sync_rosters_to_excel(tmp_path / 'nope.json', excel_path) is False
    assert not excel_path.exists()
