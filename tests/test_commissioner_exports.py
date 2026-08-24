"""Tests for commissioner roster and draft workbook generation."""

import json
from io import BytesIO
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from api.commissioner_exports import build_draft_board_workbook, build_roster_workbook
from scripts.sync_drafts_from_excel import parse_draft_sheet

PROJECT_ROOT = Path(__file__).resolve().parent.parent


def test_roster_workbook_uses_importer_compatible_grid():
    rosters = {
        'GSA': [
            {'name': 'Patrick Mahomes', 'position': 'QB', 'nfl_team': 'KC'},
            {'name': 'Taxi Receiver', 'position': 'WR', 'nfl_team': 'BUF', 'taxi': True},
        ]
    }
    teams = {
        'teams': [
            {
                'abbrev': 'GSA',
                'name': 'No Kings Except Henry',
                'owner': 'Griffin Ansel',
            }
        ]
    }

    workbook = load_workbook(BytesIO(build_roster_workbook(rosters, teams)))
    sheet = workbook['Rosters']

    assert sheet['A2'].value == 'No Kings Except Henry'
    assert sheet['A3'].value == 'Griffin Ansel'
    assert sheet['A4'].value == 'GSA'
    assert sheet['A6'].value == 'QB'
    assert sheet['A7'].value == 'Patrick Mahomes (KC)'
    assert sheet['A48'].value == 'WR'
    assert sheet['A49'].value == 'Taxi Receiver (BUF)'
    assert sheet.freeze_panes == 'A5'
    workbook.close()


def test_roster_workbook_rejects_data_that_would_be_silently_truncated():
    rosters = {
        'GSA': [
            {'name': f'Quarterback {index}', 'position': 'QB', 'nfl_team': 'KC'}
            for index in range(4)
        ]
    }

    with pytest.raises(ValueError, match='GSA has 4 active QB players'):
        build_roster_workbook(rosters, None)


def test_current_draft_board_has_trade_adjusted_2026_slots_and_ledger():
    picks = json.loads((PROJECT_ROOT / 'data' / 'draft_picks.json').read_text())
    orders = json.loads((PROJECT_ROOT / 'data' / 'draft_orders.json').read_text())
    teams = json.loads((PROJECT_ROOT / 'data' / 'teams.json').read_text())

    content = build_draft_board_workbook(picks, orders, teams, season=2026)
    workbook = load_workbook(BytesIO(content))
    board = workbook['2026 Offseason Draft']

    assert board['F3'].value == 'Round 2'
    assert board['F10'].value == 7
    assert board['G10'].value == 'AYP (via SLS → S/T)'
    assert board['F12'].value == 9
    assert board['G12'].value == 'S/T'
    assert board['A27'].value == 'TAXI Round 1'
    assert board['F39'].value == 'TAXI Round 4'

    ledger = workbook['Pick Ledger']
    ledger_rows = list(ledger.iter_rows(min_row=2, values_only=True))
    pick_207 = next(row for row in ledger_rows if row[0] == 'Offseason' and row[1] == '2.07')
    pick_209 = next(row for row in ledger_rows if row[0] == 'Offseason' and row[1] == '2.09')
    assert pick_207[4:7] == ('SLS', 'AYP', 'SLS → S/T')
    assert pick_209[4:7] == ('S/T', 'S/T', None)
    assert len(ledger_rows) == 100
    workbook.close()

    draft_frame = pd.read_excel(
        BytesIO(content),
        sheet_name='2026 Offseason Draft',
        header=None,
    )
    parsed = parse_draft_sheet(draft_frame, '2026 Offseason Draft')
    assert len(parsed) == 1
    assert [round_data['round'] for round_data in parsed[0]['rounds']] == [
        '1',
        '2',
        '3',
        '4',
        '5',
        '6',
        'TAXI Round 1',
        'TAXI Round 2',
        'TAXI Round 3',
        'TAXI Round 4',
    ]
    assert all(len(round_data['picks']) == 10 for round_data in parsed[0]['rounds'])


def test_draft_board_rejects_a_missing_pick_slot():
    orders = {
        '2026': {
            'offseason': ['GSA', 'CGK'],
            'offseason_taxi': ['GSA', 'CGK'],
        }
    }
    picks = {
        'picks': [
            {
                'year': '2026',
                'round': 1,
                'draft_type': 'offseason',
                'original_team': 'GSA',
                'current_owner': 'GSA',
            },
            {
                'year': '2026',
                'round': 1,
                'draft_type': 'offseason_taxi',
                'original_team': 'GSA',
                'current_owner': 'GSA',
            },
            {
                'year': '2026',
                'round': 1,
                'draft_type': 'offseason_taxi',
                'original_team': 'CGK',
                'current_owner': 'CGK',
            },
        ]
    }

    with pytest.raises(ValueError, match=r'Missing offseason round 1 pick\(s\): CGK'):
        build_draft_board_workbook(picks, orders, None, season=2026)
