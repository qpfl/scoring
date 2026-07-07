"""Tests for scripts/init_rosters_from_excel.py taxi-squad parsing (docs/ROADMAP_2026.md P2.2)."""

import json

import openpyxl

from scripts.init_rosters_from_excel import init_rosters_from_excel


def _build_workbook(path, taxi_entries):
    """taxi_entries: list of (row_offset_index, position, player_cell) for team column 1 (GSA)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=2, column=1, value='Team A')
    ws.cell(row=3, column=1, value='Owner A')
    ws.cell(row=4, column=1, value='GSA')
    ws.cell(row=6, column=1, value='QB')
    ws.cell(row=7, column=1, value='Patrick Mahomes (KC)')

    taxi_rows = [(48, 49), (50, 51), (52, 53), (54, 55)]
    for i, (position, player_cell) in enumerate(taxi_entries):
        pos_row, player_row = taxi_rows[i]
        ws.cell(row=pos_row, column=1, value=position)
        ws.cell(row=player_row, column=1, value=player_cell)

    wb.save(path)


def test_taxi_players_marked_and_parsed(tmp_path):
    excel_path = tmp_path / 'Rosters.xlsx'
    output_path = tmp_path / 'rosters.json'
    _build_workbook(excel_path, [('WR', 'Taxi Guy (BUF)')])

    assert init_rosters_from_excel(excel_path, output_path) is True

    rosters = json.loads(output_path.read_text())
    taxi_guy = next(p for p in rosters['GSA'] if p['name'] == 'Taxi Guy')
    assert taxi_guy['taxi'] is True
    assert taxi_guy['position'] == 'WR'
    assert taxi_guy['nfl_team'] == 'BUF'

    active = next(p for p in rosters['GSA'] if p['name'] == 'Patrick Mahomes')
    assert 'taxi' not in active


def test_taxi_one_per_position_warns(tmp_path, capsys):
    excel_path = tmp_path / 'Rosters.xlsx'
    output_path = tmp_path / 'rosters.json'
    _build_workbook(excel_path, [('WR', 'Taxi One (BUF)'), ('WR', 'Taxi Two (MIA)')])

    init_rosters_from_excel(excel_path, output_path)

    out = capsys.readouterr().out
    assert 'WARNING' in out
    assert 'max 1 per position' in out
