#!/usr/bin/env python3
"""
Export historical season data from Excel files to JSON.

This script creates IMMUTABLE historical data files. Once created, these files
should NEVER be modified. All historical data comes from the official Excel
files in previous_seasons/.

Usage:
    python scripts/export_historical.py 2024  # Export single season
    python scripts/export_historical.py --all  # Export all historical seasons
"""

import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))



def parse_excel_scores(excel_path: Path, season: int) -> dict:
    """Parse scores from an Excel file for a season."""
    wb = load_workbook(excel_path, data_only=True)

    weeks = []
    teams_data = {}

    # Get all week sheets
    week_sheets = []
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith('Week '):
            try:
                week_num = int(sheet_name.replace('Week ', ''))
                week_sheets.append((week_num, sheet_name))
            except ValueError:
                pass
        elif sheet_name == 'Semi-Finals':
            week_sheets.append((16, sheet_name))
        elif sheet_name == 'Championship':
            week_sheets.append((17, sheet_name))

    week_sheets.sort(key=lambda x: x[0])

    for week_num, sheet_name in week_sheets:
        ws = wb[sheet_name]
        week_data = parse_week_sheet(ws, week_num)
        weeks.append(week_data)

        # Accumulate team stats
        for team in week_data.get('teams', []):
            abbrev = team['abbrev']
            if abbrev not in teams_data:
                teams_data[abbrev] = {
                    'abbrev': abbrev,
                    'team_name': team.get('team_name', ''),
                    'owner': team.get('owner', ''),
                    'total_points': 0,
                    'wins': 0,
                    'losses': 0,
                }
            teams_data[abbrev]['total_points'] += team.get('total_score', 0) or 0

    # Calculate standings based on regular season (weeks 1-15)
    standings = calculate_standings(weeks, teams_data, season)

    return {
        'season': season,
        'weeks': weeks,
        'standings': standings,
        'teams': list(teams_data.values()),
        'exported_at': datetime.now().isoformat(),
        'source': str(excel_path.name),
        'immutable': True,  # Mark as immutable
    }


def parse_week_sheet(ws, week_num: int) -> dict:
    """Parse a single week's sheet."""
    matchups = []
    teams = []

    # The sheet structure varies, but we need to find team columns and score rows
    # This is a simplified parser - may need adjustment per season format

    # Find team abbreviations in row 4 typically
    team_abbrevs = []
    for col in range(1, ws.max_column + 1, 2):  # Teams are in odd columns
        cell = ws.cell(row=4, column=col)
        if cell.value and len(str(cell.value)) <= 5:  # Abbrevs are short
            team_abbrevs.append((col, str(cell.value)))

    # Parse each team's roster
    for col_idx, abbrev in team_abbrevs:
        team_data = parse_team_column(ws, col_idx, abbrev)
        if team_data:
            teams.append(team_data)

    # Create matchups from pairs of teams
    for i in range(0, len(teams), 2):
        if i + 1 < len(teams):
            matchup = {
                'team1': teams[i],
                'team2': teams[i + 1],
            }
            matchups.append(matchup)

    return {
        'week': week_num,
        'matchups': matchups,
        'teams': teams,
    }


def parse_team_column(ws, col_idx: int, abbrev: str) -> dict:
    """Parse a team's data from a column."""
    # Get team name from row 2
    team_name = ws.cell(row=2, column=col_idx).value or ''
    owner = ws.cell(row=3, column=col_idx).value or ''

    roster = []
    total_score = 0

    # Position mapping based on typical QPFL sheet layout
    position_rows = {
        'QB': [7, 8],
        'RB': [12, 13, 14, 15],
        'WR': [18, 19, 20, 21],
        'TE': [24],
        'K': [29, 30],
        'D/ST': [33],
        'HC': [37, 38],
        'OL': [41, 42],
    }

    score_col = col_idx + 1  # Scores are in adjacent column

    for pos, rows in position_rows.items():
        for row in rows:
            player_cell = ws.cell(row=row, column=col_idx)
            score_cell = ws.cell(row=row, column=score_col)

            if player_cell.value:
                player_name = str(player_cell.value)
                # Extract NFL team from parentheses
                nfl_team = ''
                if '(' in player_name and ')' in player_name:
                    nfl_team = player_name[player_name.rfind('(')+1:player_name.rfind(')')]
                    player_name = player_name[:player_name.rfind('(')].strip()

                score = score_cell.value
                if isinstance(score, (int, float)):
                    score = float(score)
                else:
                    score = 0.0

                roster.append({
                    'name': player_name,
                    'position': pos,
                    'nfl_team': nfl_team,
                    'score': score,
                })

                total_score += score

    # Try to get total from the Score: row (usually row 44)
    score_label = ws.cell(row=44, column=col_idx).value
    if score_label and 'Score:' in str(score_label):
        official_score = ws.cell(row=44, column=col_idx + 1).value
        if isinstance(official_score, (int, float)):
            total_score = float(official_score)

    return {
        'abbrev': abbrev,
        'team_name': team_name,
        'owner': owner,
        'roster': roster,
        'total_score': total_score,
    }


def calculate_standings(weeks: list, teams_data: dict, season: int) -> list:
    """Calculate final standings based on regular season (weeks 1-15)."""
    standings = {}

    for week in weeks:
        week_num = week['week']
        if week_num > 15:  # Only regular season
            continue

        # Get all scores for the week to calculate ranks
        week_scores = []
        for team in week.get('teams', []):
            week_scores.append((team['abbrev'], team.get('total_score', 0) or 0))

        week_scores.sort(key=lambda x: -x[1])

        # Assign rank points
        for rank, (abbrev, score) in enumerate(week_scores, 1):
            if abbrev not in standings:
                standings[abbrev] = {
                    'abbrev': abbrev,
                    'rank_points': 0,
                    'wins': 0,
                    'losses': 0,
                    'points_for': 0,
                    'top_half': 0,
                }

            standings[abbrev]['points_for'] += score

            # Rank points: 10 - rank + 1 = 11 - rank
            standings[abbrev]['rank_points'] += (11 - rank) / 10

            # Top half bonus
            if rank <= 5:
                standings[abbrev]['top_half'] += 1
                standings[abbrev]['rank_points'] += 0.5

        # Calculate wins/losses from matchups
        for matchup in week.get('matchups', []):
            t1 = matchup.get('team1', {})
            t2 = matchup.get('team2', {})
            s1 = t1.get('total_score', 0) or 0
            s2 = t2.get('total_score', 0) or 0

            if s1 > s2:
                standings[t1['abbrev']]['wins'] += 1
                standings[t2['abbrev']]['losses'] += 1
            elif s2 > s1:
                standings[t2['abbrev']]['wins'] += 1
                standings[t1['abbrev']]['losses'] += 1

    # Sort by rank_points, then wins, then points_for
    result = sorted(
        standings.values(),
        key=lambda x: (-x['rank_points'], -x['wins'], -x['points_for'])
    )

    return result


def export_season(season: int, output_dir: Path) -> None:
    """Export a single season to JSON."""
    excel_path = Path(f'previous_seasons/{season} Scores.xlsx')

    if not excel_path.exists():
        print(f"ERROR: Excel file not found: {excel_path}")
        return

    print(f"Exporting {season} season from {excel_path}...")

    try:
        data = parse_excel_scores(excel_path, season)

        output_file = output_dir / f'{season}.json'
        with open(output_file, 'w') as f:
            json.dump(data, f, indent=2)

        print(f"  -> Wrote {output_file}")

        # Verify key scores
        for week in data['weeks']:
            if week['week'] == 17:  # Championship
                print("  Championship scores:")
                for matchup in week['matchups'][:2]:
                    t1 = matchup['team1']
                    t2 = matchup['team2']
                    print(f"    {t1['abbrev']} ({t1['total_score']}) vs {t2['abbrev']} ({t2['total_score']})")

    except Exception as e:
        print(f"ERROR exporting {season}: {e}")
        import traceback
        traceback.print_exc()


def main():
    if len(sys.argv) < 2:
        print("Usage: python scripts/export_historical.py <year> | --all")
        sys.exit(1)

    # Create historical data directory
    output_dir = Path('web/data/historical')
    output_dir.mkdir(parents=True, exist_ok=True)

    if sys.argv[1] == '--all':
        for season in range(2020, 2025):
            export_season(season, output_dir)
    else:
        try:
            season = int(sys.argv[1])
            export_season(season, output_dir)
        except ValueError:
            print(f"Invalid season: {sys.argv[1]}")
            sys.exit(1)


if __name__ == '__main__':
    main()

