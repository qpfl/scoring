#!/usr/bin/env python3
"""Export Excel scores to JSON for web display."""

import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import nflreadpy as nfl
import openpyxl

from qpfl.constants import POSITION_ROWS, TAXI_ROWS, TEAM_COLUMNS

# Trade deadline week
TRADE_DEADLINE_WEEK = 12

# Owner name to team code mapping
OWNER_TO_CODE = {
    'Griffin': 'GSA',
    'Bill': 'WJK',
    'Ryan': 'RPA',
    'Spencer/Tim': 'S/T',
    'Kaminska': 'CGK',
    'Anagh': 'AST',
    'Connor': 'CWR',
    'Joe/Joe': 'J/J',
    'Stephen': 'SLS',
    'Arnav': 'AYP',
}

# All team codes
ALL_TEAMS = ['GSA', 'WJK', 'RPA', 'S/T', 'CGK', 'AST', 'CWR', 'J/J', 'SLS', 'AYP']

# Team code aliases (for parsing variations)
TEAM_ALIASES = {
    'T/S': 'S/T',
    'SPY': 'AYP',
}

# Global cache for canonical player names from rosters.json
_CANONICAL_NAMES: dict[str, str] = {}  # lowercase normalized -> canonical name


def _normalize_for_matching(name: str) -> str:
    """Normalize a name for fuzzy matching by removing suffixes and lowercasing."""
    # Remove common suffixes
    normalized = re.sub(r'\s+(Sr\.?|Jr\.?|II|III|IV|V)$', '', name.strip(), flags=re.IGNORECASE)
    return normalized.lower()


def _load_canonical_names() -> dict[str, str]:
    """Load canonical player names from rosters.json."""
    global _CANONICAL_NAMES
    if _CANONICAL_NAMES:
        return _CANONICAL_NAMES

    script_dir = Path(__file__).parent
    rosters_path = script_dir.parent / 'data' / 'rosters.json'

    if not rosters_path.exists():
        return {}

    try:
        with open(rosters_path) as f:
            rosters = json.load(f)

        for _team_abbrev, players in rosters.items():
            for player in players:
                canonical_name = player.get('name', '')
                if canonical_name:
                    # Map the normalized version to the canonical name
                    normalized = _normalize_for_matching(canonical_name)
                    _CANONICAL_NAMES[normalized] = canonical_name
    except Exception:
        pass

    return _CANONICAL_NAMES


def _match_canonical_name(name: str) -> str:
    """Match a player name to its canonical version from rosters.json."""
    canonical_names = _load_canonical_names()
    if not canonical_names:
        return name

    normalized = _normalize_for_matching(name)

    # Try exact match on normalized name
    if normalized in canonical_names:
        return canonical_names[normalized]

    # Try matching by last name only if first name/initial matches
    # This handles cases like "J. Cook" -> "James Cook III"
    name_parts = normalized.split()
    if len(name_parts) >= 2:
        first_part = name_parts[0].rstrip('.')  # Remove trailing dot from initials
        last_name = name_parts[-1]

        for canonical_normalized, canonical_name in canonical_names.items():
            canonical_parts = canonical_normalized.split()
            if len(canonical_parts) >= 2:
                canonical_first = canonical_parts[0]
                canonical_last = canonical_parts[-1]

                # Last names must match
                if canonical_last != last_name:
                    continue

                # First name must match or be an initial of the canonical first name
                if first_part == canonical_first:
                    return canonical_name
                if len(first_part) == 1 and canonical_first.startswith(first_part):
                    return canonical_name

    # No match found, return original
    return name


# Schedule data (hardcoded for 2025 season - future seasons use schedule.txt)
SCHEDULE = [
    # Week 1-15 matchups as (team1, team2) tuples using owner names
    [
        ('Griffin', 'Bill'),
        ('Ryan', 'Spencer/Tim'),
        ('Kaminska', 'Anagh'),
        ('Connor', 'Joe/Joe'),
        ('Stephen', 'Arnav'),
    ],
    [
        ('Griffin', 'Anagh'),
        ('Ryan', 'Kaminska'),
        ('Connor', 'Bill'),
        ('Stephen', 'Joe/Joe'),
        ('Spencer/Tim', 'Arnav'),
    ],
    [
        ('Griffin', 'Joe/Joe'),
        ('Ryan', 'Arnav'),
        ('Kaminska', 'Bill'),
        ('Connor', 'Stephen'),
        ('Spencer/Tim', 'Anagh'),
    ],
    [
        ('Griffin', 'Stephen'),
        ('Ryan', 'Joe/Joe'),
        ('Kaminska', 'Spencer/Tim'),
        ('Connor', 'Anagh'),
        ('Bill', 'Arnav'),
    ],
    [
        ('Griffin', 'Ryan'),
        ('Connor', 'Kaminska'),
        ('Bill', 'Joe/Joe'),
        ('Arnav', 'Anagh'),
        ('Spencer/Tim', 'Stephen'),
    ],  # Rivalry Week
    [
        ('Griffin', 'Arnav'),
        ('Ryan', 'Anagh'),
        ('Kaminska', 'Joe/Joe'),
        ('Connor', 'Spencer/Tim'),
        ('Stephen', 'Bill'),
    ],
    [
        ('Griffin', 'Kaminska'),
        ('Ryan', 'Stephen'),
        ('Connor', 'Arnav'),
        ('Spencer/Tim', 'Bill'),
        ('Joe/Joe', 'Anagh'),
    ],
    [
        ('Griffin', 'Connor'),
        ('Ryan', 'Bill'),
        ('Kaminska', 'Arnav'),
        ('Stephen', 'Anagh'),
        ('Spencer/Tim', 'Joe/Joe'),
    ],
    [
        ('Griffin', 'Spencer/Tim'),
        ('Ryan', 'Connor'),
        ('Kaminska', 'Stephen'),
        ('Joe/Joe', 'Arnav'),
        ('Anagh', 'Bill'),
    ],
    [
        ('Griffin', 'Stephen'),
        ('Ryan', 'Kaminska'),
        ('Connor', 'Spencer/Tim'),
        ('Joe/Joe', 'Bill'),
        ('Anagh', 'Arnav'),
    ],
    [
        ('Griffin', 'Connor'),
        ('Ryan', 'Arnav'),
        ('Kaminska', 'Bill'),
        ('Stephen', 'Joe/Joe'),
        ('Spencer/Tim', 'Anagh'),
    ],
    [
        ('Griffin', 'Arnav'),
        ('Ryan', 'Anagh'),
        ('Kaminska', 'Connor'),
        ('Stephen', 'Bill'),
        ('Spencer/Tim', 'Joe/Joe'),
    ],
    [
        ('Griffin', 'Ryan'),
        ('Kaminska', 'Joe/Joe'),
        ('Connor', 'Bill'),
        ('Stephen', 'Anagh'),
        ('Spencer/Tim', 'Arnav'),
    ],
    [
        ('Griffin', 'Kaminska'),
        ('Ryan', 'Spencer/Tim'),
        ('Connor', 'Joe/Joe'),
        ('Stephen', 'Arnav'),
        ('Anagh', 'Bill'),
    ],
    [
        ('Griffin', 'Bill'),
        ('Ryan', 'Stephen'),
        ('Kaminska', 'Spencer/Tim'),
        ('Connor', 'Arnav'),
        ('Joe/Joe', 'Anagh'),
    ],
]

# Playoff bracket structure for weeks 16-17
# Week 16: Semifinals - matchups based on final regular season standings
# Week 17: Finals - matchups based on week 16 results
PLAYOFF_STRUCTURE = {
    16: {
        'is_playoffs': True,
        'round': 'Semifinals',
        'matchups': [
            # Playoffs (1-4 seeds)
            {'seed1': 1, 'seed2': 4, 'bracket': 'playoffs', 'game': 'semi_1'},
            {'seed1': 2, 'seed2': 3, 'bracket': 'playoffs', 'game': 'semi_2'},
            # Mid Bowl (two-week total points matchup)
            {
                'seed1': 5,
                'seed2': 6,
                'bracket': 'mid_bowl',
                'game': 'mid_bowl_week1',
                'two_week': True,
            },
            # Sewer Series (7-10 seeds)
            {'seed1': 7, 'seed2': 10, 'bracket': 'sewer_series', 'game': 'sewer_1'},
            {'seed1': 8, 'seed2': 9, 'bracket': 'sewer_series', 'game': 'sewer_2'},
        ],
    },
    17: {
        'is_playoffs': True,
        'round': 'Finals',
        'matchups': [
            # Championship: winners of semi_1 and semi_2
            {
                'from_games': ['semi_1', 'semi_2'],
                'take': 'winners',
                'bracket': 'championship',
                'game': 'championship',
            },
            # Consolation Cup: losers of semi_1 and semi_2
            {
                'from_games': ['semi_1', 'semi_2'],
                'take': 'losers',
                'bracket': 'consolation_cup',
                'game': 'consolation_cup',
            },
            # Mid Bowl Week 2 (continuation of two-week matchup)
            {
                'seed1': 5,
                'seed2': 6,
                'bracket': 'mid_bowl',
                'game': 'mid_bowl_week2',
                'two_week': True,
            },
            # Toilet Bowl: losers of sewer_1 and sewer_2
            {
                'from_games': ['sewer_1', 'sewer_2'],
                'take': 'losers',
                'bracket': 'toilet_bowl',
                'game': 'toilet_bowl',
            },
        ],
    },
}


def get_playoff_matchups(
    standings: list[dict], week_num: int, week_16_results: dict = None
) -> list[dict]:
    """Generate playoff matchups based on standings and week 16 results.

    Args:
        standings: List of team standings sorted by rank (index 0 = seed 1)
        week_num: Week number (16 or 17)
        week_16_results: Dict of game_id -> {'winner': abbrev, 'loser': abbrev} for week 17

    Returns:
        List of matchup dicts with team1, team2, and playoff metadata
    """
    if week_num not in PLAYOFF_STRUCTURE:
        return []

    playoff_info = PLAYOFF_STRUCTURE[week_num]
    matchups = []

    # Create seed to team mapping
    seed_to_team = {i + 1: team['abbrev'] for i, team in enumerate(standings)}

    for game in playoff_info['matchups']:
        matchup = {
            'bracket': game['bracket'],
            'game': game['game'],
        }

        if 'seed1' in game:
            # Seeded matchup (week 16)
            matchup['team1'] = seed_to_team.get(game['seed1'])
            matchup['team2'] = seed_to_team.get(game['seed2'])
            matchup['seed1'] = game['seed1']
            matchup['seed2'] = game['seed2']
        elif 'from_games' in game and week_16_results:
            # Bracket-based matchup (week 17)
            teams = []
            for from_game in game['from_games']:
                if from_game in week_16_results:
                    team = week_16_results[from_game].get(
                        game['take'][:-1]
                    )  # 'winners' -> 'winner'
                    if team:
                        teams.append(team)

            if len(teams) == 2:
                matchup['team1'] = teams[0]
                matchup['team2'] = teams[1]
            else:
                matchup['team1'] = 'TBD'
                matchup['team2'] = 'TBD'
        else:
            matchup['team1'] = 'TBD'
            matchup['team2'] = 'TBD'

        matchups.append(matchup)

    return matchups


def adjust_standings_for_playoffs_json(
    standings: list[dict], season: int, weeks: list[dict]
) -> list[dict]:
    """Adjust standings to reflect final playoff positions.

    Reorders standings based on playoff results:
    - 1st: Championship winner
    - 2nd: Championship loser
    - 3rd: Consolation cup winner
    - 4th: Consolation cup loser
    - 5th+: Based on regular season (non-playoff teams)
    """
    if not weeks:
        return standings

    # Find the final week (championship week)
    final_week_num = 16 if season <= 2021 else 17
    final_week = next((w for w in weeks if w.get('week') == final_week_num), None)

    if not final_week:
        return standings

    # Find championship and consolation cup matchups
    championship_matchup = None
    consolation_matchup = None

    for matchup in final_week.get('matchups', []):
        game = matchup.get('game', '')
        if game == 'championship':
            championship_matchup = matchup
        elif game == 'consolation_cup':
            consolation_matchup = matchup

    if not championship_matchup:
        return standings

    # Determine winners and losers
    t1 = championship_matchup.get('team1', {})
    t2 = championship_matchup.get('team2', {})

    if isinstance(t1, str) or isinstance(t2, str):
        return standings  # TBD teams

    s1 = t1.get('total_score', 0)
    s2 = t2.get('total_score', 0)

    if s1 == 0 and s2 == 0:
        return standings  # No scores yet

    champ_abbrev = t1.get('abbrev') if s1 > s2 else t2.get('abbrev')
    runner_up_abbrev = t2.get('abbrev') if s1 > s2 else t1.get('abbrev')

    third_abbrev = None
    fourth_abbrev = None
    if consolation_matchup:
        t1 = consolation_matchup.get('team1', {})
        t2 = consolation_matchup.get('team2', {})
        if isinstance(t1, dict) and isinstance(t2, dict):
            s1 = t1.get('total_score', 0)
            s2 = t2.get('total_score', 0)
            if s1 > 0 or s2 > 0:
                third_abbrev = t1.get('abbrev') if s1 > s2 else t2.get('abbrev')
                fourth_abbrev = t2.get('abbrev') if s1 > s2 else t1.get('abbrev')

    # Create a mapping of abbrev to desired position
    playoff_positions = {
        champ_abbrev: 1,
        runner_up_abbrev: 2,
    }
    if third_abbrev:
        playoff_positions[third_abbrev] = 3
    if fourth_abbrev:
        playoff_positions[fourth_abbrev] = 4

    # Separate playoff teams from non-playoff teams
    playoff_teams = [s for s in standings if s.get('abbrev') in playoff_positions]
    non_playoff_teams = [s for s in standings if s.get('abbrev') not in playoff_positions]

    # Sort playoff teams by their playoff position
    playoff_teams.sort(key=lambda x: playoff_positions.get(x.get('abbrev'), 999))

    # Combine: playoff teams first (in playoff order), then non-playoff teams
    return playoff_teams + non_playoff_teams


def get_schedule_data(standings: list[dict] = None, weeks: list[dict] = None) -> list[dict]:
    """Convert schedule to JSON format with team codes."""
    schedule_data = []

    # Regular season weeks 1-15
    for week_num, matchups in enumerate(SCHEDULE, 1):
        week_matchups = []
        for owner1, owner2 in matchups:
            week_matchups.append(
                {
                    'team1': OWNER_TO_CODE.get(owner1, owner1),
                    'team2': OWNER_TO_CODE.get(owner2, owner2),
                }
            )
        schedule_data.append(
            {
                'week': week_num,
                'is_rivalry': week_num == 5,
                'is_playoffs': False,
                'matchups': week_matchups,
            }
        )

    # Calculate week 16 results for week 17 matchups
    week_16_results = {}
    if weeks:
        for week_data in weeks:
            if week_data.get('week') == 16:
                for matchup in week_data.get('matchups', []):
                    t1 = matchup.get('team1', {})
                    t2 = matchup.get('team2', {})
                    game_id = matchup.get('game')

                    if (
                        game_id
                        and t1.get('total_score') is not None
                        and t2.get('total_score') is not None
                    ):
                        s1, s2 = t1['total_score'], t2['total_score']
                        if s1 > s2:
                            week_16_results[game_id] = {
                                'winner': t1['abbrev'],
                                'loser': t2['abbrev'],
                            }
                        elif s2 > s1:
                            week_16_results[game_id] = {
                                'winner': t2['abbrev'],
                                'loser': t1['abbrev'],
                            }
                        # If tied, don't set winner/loser (TBD)
                break

    # Playoff weeks 16-17
    if standings:
        for week_num in [16, 17]:
            playoff_info = PLAYOFF_STRUCTURE[week_num]
            week_matchups = get_playoff_matchups(
                standings, week_num, week_16_results if week_num == 17 else None
            )

            schedule_data.append(
                {
                    'week': week_num,
                    'is_rivalry': False,
                    'is_playoffs': True,
                    'playoff_round': playoff_info['round'],
                    'matchups': week_matchups,
                }
            )

    return schedule_data


def normalize_team_code(team: str) -> str:
    """Normalize team code variations."""
    team = team.strip()
    return TEAM_ALIASES.get(team, team)


def parse_player_name(cell_value: str) -> tuple[str, str]:
    """Parse 'Player Name (TEAM)' into (name, team)."""
    if not cell_value:
        return '', ''
    match = re.match(r'^(.+?)\s*\(([A-Z]{2,3})\)$', cell_value.strip())
    if match:
        name = match.group(1).strip()
        team = match.group(2)
    else:
        name = cell_value.strip()
        team = ''

    # Apply fuzzy matching to get canonical name from rosters.json
    name = _match_canonical_name(name)
    return name, team


def parse_fa_pool(ws) -> list[dict]:
    """Parse the FA pool from the Excel worksheet.

    NOTE: FA pool is now loaded from data/fa_pool.json instead.
    This function is kept for backwards compatibility but returns empty.
    """
    # FA pool constants are no longer defined - load from JSON instead
    return []


def export_week(ws, week_num: int, bench_scores: dict = None) -> dict[str, Any]:
    """Export a single week's data to dict format.

    Args:
        ws: Excel worksheet
        week_num: Week number
        bench_scores: Optional dict mapping (team_abbrev, player_name) -> score for bench players
    """
    matchups = []
    teams_data = []

    # Get all team info
    for _i, col in enumerate(TEAM_COLUMNS):
        team_name = ws.cell(row=2, column=col).value
        if not team_name:
            continue

        team_name = str(team_name).strip().strip('*')
        owner = ws.cell(row=3, column=col).value or ''
        abbrev = ws.cell(row=4, column=col).value or ''

        # Get all players and scores
        roster = []
        total_score = 0.0

        for position, (_header_row, player_rows) in POSITION_ROWS.items():
            for row in player_rows:
                player_cell = ws.cell(row=row, column=col)
                score_cell = ws.cell(row=row, column=col + 1)

                if player_cell.value:
                    player_name, nfl_team = parse_player_name(str(player_cell.value))
                    is_starter = player_cell.font.bold if player_cell.font else False
                    # Handle non-numeric score values like "BYE"
                    try:
                        excel_score = float(score_cell.value) if score_cell.value else 0.0
                    except (ValueError, TypeError):
                        excel_score = 0.0

                    # For bench players, use calculated score if available
                    if is_starter:
                        score = excel_score
                    elif bench_scores and (abbrev, player_name) in bench_scores:
                        score = bench_scores[(abbrev, player_name)]
                    else:
                        score = excel_score

                    roster.append(
                        {
                            'name': player_name,
                            'nfl_team': nfl_team,
                            'position': position,
                            'score': score,
                            'starter': is_starter,
                        }
                    )

                    if is_starter:
                        total_score += excel_score  # Always use Excel score for total

        # Get taxi squad players with scores
        taxi_squad = []
        for pos_row, player_row in TAXI_ROWS:
            pos_cell = ws.cell(row=pos_row, column=col)
            player_cell = ws.cell(row=player_row, column=col)

            if pos_cell.value and player_cell.value:
                position = str(pos_cell.value).strip()
                player_name, nfl_team = parse_player_name(str(player_cell.value))
                if player_name:
                    # Get score from bench_scores if available
                    score = 0.0
                    if bench_scores and (abbrev, player_name) in bench_scores:
                        score = bench_scores[(abbrev, player_name)]

                    taxi_squad.append(
                        {
                            'name': player_name,
                            'nfl_team': nfl_team,
                            'position': position,
                            'score': score,
                        }
                    )

        teams_data.append(
            {
                'name': team_name,
                'owner': owner,
                'abbrev': abbrev,
                'roster': roster,
                'taxi_squad': taxi_squad,
                'total_score': round(total_score, 1),
            }
        )

    # Calculate score_rank from total_scores (1 = highest score)
    sorted_by_score = sorted(teams_data, key=lambda t: t['total_score'], reverse=True)
    for rank, team in enumerate(sorted_by_score, 1):
        team['score_rank'] = rank

    # Group into matchups (teams are paired: 0v1, 2v3, etc.)
    for i in range(0, len(teams_data), 2):
        if i + 1 < len(teams_data):
            matchups.append(
                {
                    'team1': teams_data[i],
                    'team2': teams_data[i + 1],
                }
            )

    # Check if week has valid scores (at least one non-zero score)
    has_scores = any(t['total_score'] > 0 for t in teams_data)

    return {
        'week': week_num,
        'matchups': matchups,
        'teams': teams_data,
        'has_scores': has_scores,
    }


def get_current_nfl_week() -> int:
    """Get the current NFL week from nflreadpy."""
    return nfl.get_current_week()


def get_game_times(season: int = 2025) -> dict[int, dict[str, str]]:
    """Get game kickoff times for each team by week.

    Returns:
        Dict mapping week -> {team_abbrev -> kickoff_datetime_iso}
    """
    try:
        schedule = nfl.load_schedules(seasons=season)
        game_times = {}

        for week in range(1, 19):
            week_games = schedule.filter(schedule['week'] == week)
            if week_games.height == 0:
                continue

            game_times[week] = {}

            for row in week_games.iter_rows(named=True):
                # Get game datetime
                game_date = row.get('gameday', '')
                game_time = row.get('gametime', '')

                if game_date and game_time:
                    # Combine date and time into ISO format
                    # gametime is typically in "HH:MM" format (ET)
                    try:
                        dt_str = f'{game_date} {game_time}'
                        # Parse and convert to ISO format with timezone
                        dt = datetime.strptime(dt_str, '%Y-%m-%d %H:%M')
                        # NFL times are Eastern, add timezone info
                        # Store as ISO string (frontend will handle timezone)
                        kickoff_iso = dt.strftime('%Y-%m-%dT%H:%M:00-05:00')

                        home_team = row.get('home_team', '')
                        away_team = row.get('away_team', '')

                        if home_team:
                            game_times[week][home_team] = kickoff_iso
                        if away_team:
                            game_times[week][away_team] = kickoff_iso
                    except (ValueError, TypeError):
                        pass

        return game_times
    except Exception as e:
        print(f'Warning: Could not load game times: {e}')
        return {}


def calculate_team_stats(weeks: list, standings: list) -> dict:
    """Calculate comprehensive team statistics from weekly data.

    Args:
        weeks: List of week data with matchups
        standings: List of standings entries

    Returns:
        Dict with team stats keyed by team abbreviation
    """
    import statistics

    team_stats = {}

    # Initialize stats for each team
    for standing in standings:
        abbrev = standing['abbrev']
        team_stats[abbrev] = {
            'abbrev': abbrev,
            'name': standing.get('name', abbrev),
            'points_for': [],
            'points_against': [],
            'margins': [],
            'weekly_ranks': [],
            'wins': 0,
            'losses': 0,
            'ties': 0,
            'streak': {'type': None, 'count': 0},
            'current_streak': [],
        }

    # Process each week's matchups
    for week_data in weeks:
        matchups = week_data.get('matchups', [])

        # Calculate weekly scores for ranking
        weekly_scores = []
        for matchup in matchups:
            if matchup.get('team1') and matchup.get('team2'):
                t1 = matchup['team1']
                t2 = matchup['team2']
                s1 = t1.get('total_score') or t1.get('score')
                s2 = t2.get('total_score') or t2.get('score')
                if s1 is not None:
                    weekly_scores.append((t1['abbrev'], s1))
                if s2 is not None:
                    weekly_scores.append((t2['abbrev'], s2))

        # Sort by score descending for ranking
        weekly_scores.sort(key=lambda x: x[1], reverse=True)
        rank_map = {abbrev: rank + 1 for rank, (abbrev, _) in enumerate(weekly_scores)}

        # Process each matchup
        for matchup in matchups:
            if not matchup.get('team1') or not matchup.get('team2'):
                continue

            t1 = matchup['team1']
            t2 = matchup['team2']

            # Get scores - support both 'total_score' and 'score' keys
            s1 = t1.get('total_score') or t1.get('score')
            s2 = t2.get('total_score') or t2.get('score')

            if s1 is None or s2 is None:
                continue

            # Team 1 stats
            if t1['abbrev'] in team_stats:
                stats = team_stats[t1['abbrev']]
                stats['points_for'].append(s1)
                stats['points_against'].append(s2)
                margin = s1 - s2
                stats['margins'].append(margin)
                if t1['abbrev'] in rank_map:
                    stats['weekly_ranks'].append(rank_map[t1['abbrev']])

                if margin > 0:
                    stats['wins'] += 1
                    stats['current_streak'].append('W')
                elif margin < 0:
                    stats['losses'] += 1
                    stats['current_streak'].append('L')
                else:
                    stats['ties'] += 1
                    stats['current_streak'].append('T')

            # Team 2 stats
            if t2['abbrev'] in team_stats:
                stats = team_stats[t2['abbrev']]
                stats['points_for'].append(s2)
                stats['points_against'].append(s1)
                margin = s2 - s1
                stats['margins'].append(margin)
                if t2['abbrev'] in rank_map:
                    stats['weekly_ranks'].append(rank_map[t2['abbrev']])

                if margin > 0:
                    stats['wins'] += 1
                    stats['current_streak'].append('W')
                elif margin < 0:
                    stats['losses'] += 1
                    stats['current_streak'].append('L')
                else:
                    stats['ties'] += 1
                    stats['current_streak'].append('T')

    # Calculate derived stats for each team
    for _abbrev, stats in team_stats.items():
        pf = stats['points_for']
        pa = stats['points_against']
        margins = stats['margins']
        ranks = stats['weekly_ranks']

        games_played = len(pf)
        if games_played == 0:
            continue

        # Basic totals
        stats['total_points_for'] = sum(pf)
        stats['total_points_against'] = sum(pa)
        stats['point_differential'] = stats['total_points_for'] - stats['total_points_against']

        # Averages
        stats['ppg'] = stats['total_points_for'] / games_played
        stats['ppg_against'] = stats['total_points_against'] / games_played
        stats['avg_margin'] = sum(margins) / games_played
        stats['avg_rank'] = sum(ranks) / len(ranks) if ranks else 0

        # Standard deviation
        stats['std_dev'] = statistics.stdev(pf) if len(pf) > 1 else 0

        # Best/Worst weeks
        stats['best_week'] = max(pf)
        stats['worst_week'] = min(pf)
        stats['best_week_num'] = pf.index(max(pf)) + 1
        stats['worst_week_num'] = pf.index(min(pf)) + 1

        # Largest win/loss margins
        stats['largest_win'] = max(margins) if margins else 0
        stats['largest_loss'] = min(margins) if margins else 0

        # Win percentage
        total_games = stats['wins'] + stats['losses'] + stats['ties']
        stats['win_pct'] = (
            (stats['wins'] + 0.5 * stats['ties']) / total_games if total_games > 0 else 0
        )

        # Current streak
        streak = stats['current_streak']
        if streak:
            last_result = streak[-1]
            streak_count = 0
            for result in reversed(streak):
                if result == last_result:
                    streak_count += 1
                else:
                    break
            stats['streak'] = {'type': last_result, 'count': streak_count}

        # Games above/below .500
        stats['games_above_500'] = stats['wins'] - stats['losses']

        # Record string
        stats['record'] = f'{stats["wins"]}-{stats["losses"]}'
        if stats['ties'] > 0:
            stats['record'] += f'-{stats["ties"]}'

        # Calculate OPR: (5*avg_points + 2*(high_score + low_score) + 3*win%)/10
        # Note: win_pct is 0-1, so multiply by 100 for the formula
        avg_points = stats['ppg']
        high_score = stats['best_week']
        low_score = stats['worst_week']
        win_pct_100 = stats['win_pct'] * 100
        stats['opr'] = (5 * avg_points + 2 * (high_score + low_score) + 3 * win_pct_100) / 10

        # Cleanup temporary lists
        del stats['points_for']
        del stats['points_against']
        del stats['margins']
        del stats['weekly_ranks']
        del stats['current_streak']

    # Calculate league average OPR and Adjusted OPR
    opr_values = [s['opr'] for s in team_stats.values() if 'opr' in s]
    league_avg_opr = sum(opr_values) / len(opr_values) if opr_values else 1

    for _abbrev, stats in team_stats.items():
        if 'opr' in stats:
            stats['adjusted_opr'] = stats['opr'] / league_avg_opr if league_avg_opr > 0 else 0
            stats['league_avg_opr'] = league_avg_opr

    return team_stats


def calculate_bench_scores(excel_path: str, sheet_name: str, week_num: int) -> dict:
    """Calculate scores for bench players and taxi squad players using the scorer.

    Returns:
        Dict mapping (team_abbrev, player_name) -> score
    """
    import sys

    import openpyxl

    # Ensure parent directory is in path for qpfl import
    script_dir = Path(__file__).parent
    project_dir = script_dir.parent
    if str(project_dir) not in sys.path:
        sys.path.insert(0, str(project_dir))

    try:
        from qpfl import QPFLScorer
        from qpfl.excel_parser import parse_roster_from_excel
    except ImportError:
        return {}

    try:
        teams = parse_roster_from_excel(excel_path, sheet_name)
        scorer = QPFLScorer(2025, week_num)

        bench_scores = {}
        for team in teams:
            for position, players in team.players.items():
                for player_name, nfl_team, is_started in players:
                    if not is_started:  # Only calculate for bench players
                        try:
                            result = scorer.score_player(player_name, nfl_team, position)
                            bench_scores[(team.abbreviation, player_name)] = result.total_points
                        except Exception:
                            pass  # Skip if scoring fails

        # Also calculate scores for taxi squad players
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            ws = wb[sheet_name]

            for _i, col in enumerate(TEAM_COLUMNS):
                abbrev = ws.cell(row=4, column=col).value
                if not abbrev:
                    continue
                abbrev = str(abbrev).strip()

                for pos_row, player_row in TAXI_ROWS:
                    pos_cell = ws.cell(row=pos_row, column=col)
                    player_cell = ws.cell(row=player_row, column=col)

                    if pos_cell.value and player_cell.value:
                        position = str(pos_cell.value).strip()
                        player_name, nfl_team = parse_player_name(str(player_cell.value))
                        if player_name:
                            try:
                                result = scorer.score_player(player_name, nfl_team, position)
                                bench_scores[(abbrev, player_name)] = result.total_points
                            except Exception:
                                pass  # Skip if scoring fails
        except Exception as e:
            print(f'Warning: Could not calculate taxi scores: {e}')

        return bench_scores
    except Exception as e:
        print(f'Warning: Could not calculate bench scores for week {week_num}: {e}')
        return {}


def merge_json_lineup(week_data: dict, lineup_file: Path, week_num: int) -> dict:
    """Merge JSON lineup data into Excel week data.

    This allows teams using the website to submit lineups that get merged
    with the Excel data for other teams.
    """
    try:
        with open(lineup_file) as f:
            lineup_data = json.load(f)
    except Exception as e:
        print(f'Warning: Could not read lineup file {lineup_file}: {e}')
        return week_data

    json_teams = set(lineup_data.get('lineups', {}).keys())
    if not json_teams:
        return week_data

    # Filter out teams with empty lineups (they use Excel, not website)
    active_json_teams = set()
    for team_code, starters in lineup_data.get('lineups', {}).items():
        total_starters = sum(len(v) for v in starters.values())
        if total_starters > 0:
            active_json_teams.add(team_code)

    if active_json_teams:
        print(f'  Merging JSON lineups for Week {week_num}: {", ".join(sorted(active_json_teams))}')

    # Update starter flags in roster based on JSON lineup data
    for team in week_data.get('teams', []):
        abbrev = team.get('abbrev')
        if abbrev not in active_json_teams:
            continue

        json_starters = lineup_data['lineups'][abbrev]

        # Update starter flags in roster
        for player in team.get('roster', []):
            position = player.get('position')
            player_name = player.get('name')

            # Check if this player is a starter according to JSON
            position_starters = json_starters.get(position, [])
            player['starter'] = player_name in position_starters

        # Recalculate total score from starters
        team['total_score'] = sum(p['score'] for p in team.get('roster', []) if p.get('starter'))

    # Rebuild matchups with updated team data
    teams_by_abbrev = {t['abbrev']: t for t in week_data.get('teams', [])}

    new_matchups = []
    if week_num <= len(SCHEDULE):
        # Regular season - use SCHEDULE for matchups
        for owner1, owner2 in SCHEDULE[week_num - 1]:
            t1_abbrev = OWNER_TO_CODE.get(owner1)
            t2_abbrev = OWNER_TO_CODE.get(owner2)

            t1 = teams_by_abbrev.get(t1_abbrev)
            t2 = teams_by_abbrev.get(t2_abbrev)

            if t1 and t2:
                new_matchups.append({'team1': t1, 'team2': t2})
        week_data['matchups'] = new_matchups
    else:
        # Playoff weeks - update existing matchups in place with updated team data
        for matchup in week_data.get('matchups', []):
            t1_abbrev = matchup['team1']['abbrev']
            t2_abbrev = matchup['team2']['abbrev']
            if t1_abbrev in teams_by_abbrev:
                matchup['team1'] = teams_by_abbrev[t1_abbrev]
            if t2_abbrev in teams_by_abbrev:
                matchup['team2'] = teams_by_abbrev[t2_abbrev]

    # Recalculate has_scores
    week_data['has_scores'] = any(t.get('total_score', 0) > 0 for t in week_data.get('teams', []))

    # Recalculate score_rank
    sorted_by_score = sorted(
        week_data.get('teams', []), key=lambda t: t.get('total_score', 0), reverse=True
    )
    for rank, team in enumerate(sorted_by_score, 1):
        team['score_rank'] = rank

    return week_data


def add_playoff_metadata_to_week(weeks: list[dict], standings: list[dict], week_num: int):
    """Add playoff metadata (game, bracket) to week 16 matchups based on standings.

    This allows us to determine week 17 matchups from week 16 results.
    """
    if week_num not in PLAYOFF_STRUCTURE:
        return

    # Find the week data
    week_data = None
    for w in weeks:
        if w.get('week') == week_num:
            week_data = w
            break

    if not week_data:
        return

    # Create seed to team and team to seed mappings
    team_to_seed = {}
    for i, team in enumerate(standings):
        team_to_seed[team['abbrev']] = i + 1

    # Get expected matchups from playoff structure (for reference)
    _playoff_info = PLAYOFF_STRUCTURE[week_num]

    # Match actual matchups to expected playoff matchups
    # We need to be flexible - match by seed RANGE (1-4 = playoffs, 5-6 = mid bowl, 7-10 = sewer)
    semi_game_counter = 0
    sewer_game_counter = 0

    for matchup in week_data.get('matchups', []):
        t1 = matchup.get('team1', {})
        t2 = matchup.get('team2', {})
        t1_abbrev = t1.get('abbrev') if isinstance(t1, dict) else t1
        t2_abbrev = t2.get('abbrev') if isinstance(t2, dict) else t2

        t1_seed = team_to_seed.get(t1_abbrev, 99)
        t2_seed = team_to_seed.get(t2_abbrev, 99)

        # Determine bracket by seed ranges
        seeds = sorted([t1_seed, t2_seed])

        if seeds[0] <= 4 and seeds[1] <= 4:
            # Playoff matchup (seeds 1-4)
            semi_game_counter += 1
            matchup['game'] = f'semi_{semi_game_counter}'
            matchup['bracket'] = 'playoffs'
        elif seeds[0] == 5 and seeds[1] == 6:
            # Mid bowl
            matchup['game'] = 'mid_bowl_week1'
            matchup['bracket'] = 'mid_bowl'
            matchup['two_week'] = True
        elif seeds[0] >= 7 and seeds[1] >= 7:
            # Sewer series (seeds 7-10)
            sewer_game_counter += 1
            matchup['game'] = f'sewer_{sewer_game_counter}'
            matchup['bracket'] = 'sewer_series'


def export_all_weeks(excel_path: str) -> dict[str, Any]:
    """Export all weeks from Excel to JSON format."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    weeks = []
    # Use team code (abbrev) as unique identifier
    standings = {}  # abbrev -> {rank_points, wins, losses, ties, points_for, points_against, ...}

    # Find all week sheets (including playoff sheets with special names)
    week_sheets = []
    playoff_sheet_names = {
        'Semi-Finals': 16,
        'Championship': 17,
    }
    for sheet_name in wb.sheetnames:
        match = re.match(r'^Week (\d+)$', sheet_name)
        if match:
            week_sheets.append((int(match.group(1)), sheet_name))
        elif sheet_name in playoff_sheet_names:
            week_sheets.append((playoff_sheet_names[sheet_name], sheet_name))

    # Sort by week number
    week_sheets.sort(key=lambda x: x[0])

    # Check for JSON lineup files to merge
    script_dir = Path(__file__).parent
    project_dir = script_dir.parent
    lineups_dir = project_dir / 'data' / 'lineups' / '2025'

    # Load team name overrides
    team_names_path = project_dir / 'data' / 'team_names.json'
    team_name_overrides = {}
    if team_names_path.exists():
        with open(team_names_path) as f:
            team_name_overrides = json.load(f).get('team_names', {})

    # Export all weeks first
    for week_num, sheet_name in week_sheets:
        ws = wb[sheet_name]

        # Calculate bench scores for weeks with data
        bench_scores = calculate_bench_scores(excel_path, sheet_name, week_num)
        if bench_scores:
            print(f'  Calculated {len(bench_scores)} bench scores for Week {week_num}')

        week_data = export_week(ws, week_num, bench_scores)

        # Check for JSON lineup file and merge if present
        lineup_file = lineups_dir / f'week_{week_num}.json'
        if lineup_file.exists():
            week_data = merge_json_lineup(week_data, lineup_file, week_num)

        # Apply team name overrides for this week
        if team_name_overrides:
            for team in week_data.get('teams', []):
                team['name'] = get_team_name_for_week(
                    team['abbrev'], week_num, team_name_overrides, team.get('name', team['abbrev'])
                )
            # Also update matchups
            for matchup in week_data.get('matchups', []):
                for team_key in ['team1', 'team2']:
                    team = matchup.get(team_key, {})
                    if isinstance(team, dict) and 'abbrev' in team:
                        team['name'] = get_team_name_for_week(
                            team['abbrev'],
                            week_num,
                            team_name_overrides,
                            team.get('name', team['abbrev']),
                        )

        weeks.append(week_data)

    # Determine which weeks to include in standings
    # Only include completed weeks (before current NFL week)
    current_nfl_week = get_current_nfl_week()

    print(f'Current NFL week: {current_nfl_week}, standings include weeks 1-{current_nfl_week - 1}')

    for week_data in weeks:
        # Skip weeks without scores for standings calculation
        if not week_data.get('has_scores', False):
            continue

        # Only include regular season weeks (1-15) for standings
        # Playoff weeks (16+) don't affect regular season standings
        if week_data['week'] > 15:
            continue

        # Update standings using team code as unique ID
        for matchup in week_data['matchups']:
            t1, t2 = matchup['team1'], matchup['team2']

            for team in [t1, t2]:
                abbrev = team['abbrev']
                if abbrev not in standings:
                    standings[abbrev] = {
                        'name': team['name'],
                        'owner': team['owner'],
                        'abbrev': abbrev,
                        'rank_points': 0.0,
                        'wins': 0,
                        'losses': 0,
                        'ties': 0,
                        'top_half': 0,
                        'points_for': 0.0,
                        'points_against': 0.0,
                    }
                else:
                    # Update name/owner to latest (they may change)
                    standings[abbrev]['name'] = team['name']
                    standings[abbrev]['owner'] = team['owner']

            # Get scores
            s1 = t1['total_score']
            s2 = t2['total_score']

            # Update points for/against
            standings[t1['abbrev']]['points_for'] += s1
            standings[t1['abbrev']]['points_against'] += s2
            standings[t2['abbrev']]['points_for'] += s2
            standings[t2['abbrev']]['points_against'] += s1

            # Calculate rank points for matchup result
            # Win = 1 point, Tie = 0.5 points each
            if s1 > s2:
                standings[t1['abbrev']]['rank_points'] += 1.0
                standings[t1['abbrev']]['wins'] += 1
                standings[t2['abbrev']]['losses'] += 1
            elif s2 > s1:
                standings[t2['abbrev']]['rank_points'] += 1.0
                standings[t2['abbrev']]['wins'] += 1
                standings[t1['abbrev']]['losses'] += 1
            else:
                standings[t1['abbrev']]['rank_points'] += 0.5
                standings[t2['abbrev']]['rank_points'] += 0.5
                standings[t1['abbrev']]['ties'] += 1
                standings[t2['abbrev']]['ties'] += 1

        # Calculate top 5 bonus for each team based on their score_rank
        # Group teams by score to handle ties
        teams_by_score = sorted(week_data['teams'], key=lambda x: x['total_score'], reverse=True)

        # Assign ranks handling ties (teams with same score share the rank)
        current_rank = 1
        i = 0
        while i < len(teams_by_score):
            # Find all teams with the same score
            current_score = teams_by_score[i]['total_score']
            tied_teams = []
            while i < len(teams_by_score) and teams_by_score[i]['total_score'] == current_score:
                tied_teams.append(teams_by_score[i])
                i += 1

            # Check if any of these tied positions are in top 5
            tied_positions = list(range(current_rank, current_rank + len(tied_teams)))
            positions_in_top5 = [p for p in tied_positions if p <= 5]

            if positions_in_top5:
                # Calculate points: 0.5 points shared among tied teams that span top 5
                # If some positions are in top 5 and some aren't, split proportionally
                points_per_team = (0.5 * len(positions_in_top5)) / len(tied_teams)

                for team in tied_teams:
                    standings[team['abbrev']]['rank_points'] += points_per_team
                    standings[team['abbrev']]['top_half'] += 1

            current_rank += len(tied_teams)

    # Sort standings by: 1) rank_points, 2) wins (tiebreaker), 3) points_for (second tiebreaker)
    sorted_standings = sorted(
        standings.values(),
        key=lambda x: (x['rank_points'], x['wins'], x['points_for']),
        reverse=True,
    )

    wb.close()

    # Add playoff metadata to week 16 matchups
    add_playoff_metadata_to_week(weeks, sorted_standings, 16)

    # Use nflreadpy's current week - don't cap so offseason trading logic works (week 18+)
    # Frontend handles display capping at 17
    current_week = get_current_nfl_week()
    display_week = min(current_week, 17)  # For team names and lineup loading

    # Apply team name overrides to canonical teams
    teams_data = load_teams()
    current_teams_data = apply_team_name_overrides(teams_data, display_week, team_name_overrides)

    # Load current week lineups for pending matchups display
    current_lineups = load_current_lineups(display_week)

    return {
        'updated_at': datetime.now(timezone.utc).isoformat().replace('+00:00', 'Z'),
        'season': 2025,
        'current_week': current_week,
        'teams': current_teams_data,  # Canonical team info (with current week names)
        'rosters': load_rosters(),  # Full roster for each team
        'weeks': weeks,
        'standings': sorted_standings,
        'schedule': get_schedule_data(sorted_standings, weeks),
        'game_times': get_game_times(2025),
        'team_stats': calculate_team_stats(weeks, sorted_standings),
        'fa_pool': parse_fa_pool(wb[week_sheets[-1][1]]) if week_sheets else [],
        'pending_trades': load_pending_trades(),
        'trade_deadline_week': TRADE_DEADLINE_WEEK,
        'lineups': current_lineups,  # Current week lineup submissions
        'trade_blocks': load_trade_blocks(),
    }


def load_pending_trades() -> list[dict]:
    """Load pending trades from JSON file."""
    pending_trades_path = Path(__file__).parent.parent / 'data' / 'pending_trades.json'
    if pending_trades_path.exists():
        with open(pending_trades_path) as f:
            return json.load(f).get('trades', [])
    return []


def load_trade_blocks() -> dict:
    """Load trade blocks from JSON file."""
    trade_blocks_path = Path(__file__).parent.parent / 'data' / 'trade_blocks.json'
    if trade_blocks_path.exists():
        with open(trade_blocks_path) as f:
            return json.load(f)
    return {}


def load_current_lineups(week: int) -> dict:
    """Load current week lineups from JSON file for pending matchups display."""
    lineups_path = Path(__file__).parent.parent / 'data' / 'lineups' / '2025' / f'week_{week}.json'
    if lineups_path.exists():
        with open(lineups_path) as f:
            return json.load(f).get('lineups', {})
    return {}


def load_teams() -> list[dict]:
    """Load canonical team info from teams.json."""
    teams_path = Path(__file__).parent.parent / 'data' / 'teams.json'
    if teams_path.exists():
        with open(teams_path) as f:
            return json.load(f).get('teams', [])
    return []


def load_rosters() -> dict[str, list[dict]]:
    """Load full rosters from rosters.json."""
    rosters_path = Path(__file__).parent.parent / 'data' / 'rosters.json'
    if rosters_path.exists():
        with open(rosters_path) as f:
            return json.load(f)
    return {}


def parse_transactions(doc_path: str) -> list[dict]:
    """Parse transactions document into structured seasons/weeks with indentation.

    NOTE: Transactions are now loaded from data/transaction_log.json instead.
    This function is deprecated.
    """
    try:
        import docx as docx_module
    except ImportError:
        return []

    doc = docx_module.Document(doc_path)
    seasons = []
    current_season = None
    current_week = None
    current_transaction = None

    # Indentation thresholds (in EMUs: 914400 = 1 inch)
    level_1 = 400000  # ~0.44 inch - transaction header
    level_2 = 800000  # ~0.87 inch - sub-header (date, "To X:")
    level_3 = 1200000  # ~1.31 inch - list items

    def get_indent_level(para):
        """Get indentation level (0-3) based on left indent."""
        left_indent = para.paragraph_format.left_indent
        if left_indent is None:
            return 0
        if left_indent >= level_3:
            return 3
        if left_indent >= level_2:
            return 2
        if left_indent >= level_1:
            return 1
        return 0

    def save_transaction():
        nonlocal current_transaction
        if current_transaction and current_week:
            current_week['transactions'].append(current_transaction)
            current_transaction = None

    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue

        style = para.style.name if para.style else ''

        if style == 'Title':
            continue
        elif style == 'Heading 1':
            # New season
            save_transaction()
            current_season = {'season': text, 'weeks': []}
            seasons.append(current_season)
            current_week = None
        elif style == 'Heading 2':
            # New week/event
            save_transaction()
            if current_season:
                current_week = {'title': text, 'transactions': []}
                current_season['weeks'].append(current_week)
        else:
            # Transaction content with indentation
            if text.lower() == 'none':
                continue

            indent = get_indent_level(para)

            if indent <= 1:
                # New transaction block
                save_transaction()
                current_transaction = {'title': text, 'items': []}
            elif indent == 2:
                # Sub-header within transaction
                if current_transaction:
                    current_transaction['items'].append({'type': 'header', 'text': text})
            else:
                # List item (indent level 3)
                if current_transaction:
                    current_transaction['items'].append({'type': 'item', 'text': text})

    # Save any remaining transaction
    save_transaction()

    # Filter out empty weeks
    for season in seasons:
        season['weeks'] = [w for w in season['weeks'] if w['transactions']]

    return seasons


def load_transaction_log() -> list[dict]:
    """Load all transactions from the unified JSON log file.

    This is now the single source of truth for all transactions (historical and recent).
    """
    log_path = Path(__file__).parent.parent / 'data' / 'transaction_log.json'
    if log_path.exists():
        with open(log_path) as f:
            return json.load(f).get('transactions', [])
    return []


def format_player_for_display(player: dict | str) -> str:
    """Format a player with position and team info.

    Handles both old format (just player name string) and new format (dict with name/position/nfl_team).
    """
    if isinstance(player, str):
        return player  # Old format - just the name

    # New format - dict with full info
    name = player.get('name', 'Unknown')
    position = player.get('position', '')
    nfl_team = player.get('nfl_team', '')

    if position and nfl_team:
        return f'{position} {name} ({nfl_team})'
    elif position:
        return f'{position} {name}'
    elif nfl_team:
        return f'{name} ({nfl_team})'
    return name


def format_transaction_for_display(tx: dict) -> dict:
    """Format a JSON transaction into the display format."""
    items = []

    if tx['type'] == 'trade':
        # Format trade transaction
        items.append({'type': 'header', 'text': tx.get('timestamp', '')[:10].replace('-', '/')})
        items.append({'type': 'header', 'text': f'To {tx.get("partner", "Unknown")}:'})
        for player in tx.get('proposer_gives', {}).get('players', []):
            items.append({'type': 'item', 'text': format_player_for_display(player)})
        for pick in tx.get('proposer_gives', {}).get('picks', []):
            items.append({'type': 'item', 'text': pick})
        items.append({'type': 'header', 'text': f'To {tx.get("proposer", "Unknown")}:'})
        for player in tx.get('proposer_receives', {}).get('players', []):
            items.append({'type': 'item', 'text': format_player_for_display(player)})
        for pick in tx.get('proposer_receives', {}).get('picks', []):
            items.append({'type': 'item', 'text': pick})

        return {
            'title': f'Trade between {tx.get("proposer", "?")} and {tx.get("partner", "?")}',
            'items': items,
        }

    elif tx['type'] == 'taxi_activation':
        items.append({'type': 'header', 'text': tx.get('timestamp', '')[:10].replace('-', '/')})
        activated = format_player_for_display(tx.get('activated', 'Unknown'))
        released = format_player_for_display(tx.get('released', 'Unknown'))
        items.append(
            {
                'type': 'header',
                'text': f'Activated {activated} from taxi squad, released {released}',
            }
        )
        return {'title': tx.get('team', 'Unknown'), 'items': items}

    elif tx['type'] == 'fa_activation':
        items.append({'type': 'header', 'text': tx.get('timestamp', '')[:10].replace('-', '/')})
        added = format_player_for_display(tx.get('added', 'Unknown'))
        released = format_player_for_display(tx.get('released', 'Unknown'))
        items.append({'type': 'header', 'text': f'Added {added} from FA Pool, released {released}'})
        return {'title': tx.get('team', 'Unknown'), 'items': items}

    return {'title': 'Unknown Transaction', 'items': items}


def merge_transaction_log(doc_transactions: list[dict]) -> list[dict]:
    """Merge JSON log transactions with document transactions."""
    json_transactions = load_transaction_log()

    if not json_transactions:
        return doc_transactions

    # Group JSON transactions by week
    week_transactions = {}
    for tx in json_transactions:
        week = tx.get('week', 0)
        if week not in week_transactions:
            week_transactions[week] = []
        week_transactions[week].append(format_transaction_for_display(tx))

    # Find or create current season (2025 Season)
    current_season = None
    for season in doc_transactions:
        if '2025' in season.get('season', ''):
            current_season = season
            break

    if not current_season:
        current_season = {'season': '2025 Season', 'weeks': []}
        doc_transactions.insert(0, current_season)

    # Add JSON transactions to appropriate weeks
    for week_num, txs in week_transactions.items():
        # Find existing week
        week_title = f'Week {week_num}'
        existing_week = None
        for w in current_season['weeks']:
            if f'Week {week_num}' in w.get('title', ''):
                existing_week = w
                break

        if existing_week:
            # Add to existing week (at the beginning - newest first)
            existing_week['transactions'] = txs + existing_week['transactions']
        else:
            # Create new week
            new_week = {'title': week_title, 'transactions': txs}
            # Insert in order (higher week numbers first for most recent)
            inserted = False
            for i, w in enumerate(current_season['weeks']):
                # Extract week number from title
                try:
                    existing_week_num = int(
                        ''.join(filter(str.isdigit, w['title'].split()[0:2][1])) or 0
                    )
                    if week_num > existing_week_num:
                        current_season['weeks'].insert(i, new_week)
                        inserted = True
                        break
                except (ValueError, IndexError, KeyError):
                    continue
            if not inserted:
                current_season['weeks'].append(new_week)

    return doc_transactions


def main():
    """Main export function."""
    # Get paths relative to script location
    script_dir = Path(__file__).parent
    project_dir = script_dir.parent
    excel_path = project_dir / '2025 Scores.xlsx'
    output_path = project_dir / 'web' / 'data.json'
    web_dir = project_dir / 'web'

    # Ensure output directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)

    print(f'Exporting {excel_path} to {output_path}...')

    data = export_all_weeks(str(excel_path))

    # Parse additional documents if available (check docs folder first, then root)
    data_dir = project_dir / 'data'
    shared_dir = web_dir / 'data' / 'shared'

    # Load static data from JSON files (no more Word/Excel parsing)

    # Constitution
    constitution_path = shared_dir / 'constitution.json'
    if constitution_path.exists():
        print('Loading constitution from JSON...')
        with open(constitution_path) as f:
            const_data = json.load(f)
        data['constitution'] = const_data.get('articles', [])

    # Hall of Fame
    hof_json_path = shared_dir / 'hall_of_fame.json'
    if hof_json_path.exists():
        print('Loading Hall of Fame from JSON...')
        with open(hof_json_path) as f:
            hof_stats = json.load(f)
        data['hall_of_fame'] = {
            'finishes_by_year': hof_stats.get('finishes_by_year', []),
            'owner_stats': hof_stats.get('owner_stats', []),
            'player_records': hof_stats.get('player_records', {}),
            'team_records': hof_stats.get('team_records', {}),
            'fun_stats': hof_stats.get('fun_stats', []),
            'rivalry_records': hof_stats.get('rivalry_records', {}),
        }

    # Banners - use existing images
    banners_dir = web_dir / 'images' / 'banners'
    existing_banners = (
        sorted([f.name for f in banners_dir.glob('*_banner.png')]) if banners_dir.exists() else []
    )
    if existing_banners:
        print(f'Using {len(existing_banners)} existing banner images...')
        data['banners'] = existing_banners

    # Draft picks from JSON
    draft_picks_path = data_dir / 'draft_picks.json'
    if draft_picks_path.exists():
        print('Loading draft picks from JSON...')
        with open(draft_picks_path) as f:
            picks_data = json.load(f)
        data['draft_picks'] = picks_data.get('picks', {})

    # Load transactions from unified transaction log (single source of truth)
    print('Loading transactions...')
    data['transactions'] = load_transaction_log()

    # Drafts from JSON
    drafts_path = data_dir / 'drafts.json'
    if drafts_path.exists():
        print('Loading drafts from JSON...')
        with open(drafts_path) as f:
            drafts_data = json.load(f)
        data['drafts'] = drafts_data.get('drafts', [])

    with open(output_path, 'w') as f:
        json.dump(data, f, indent=2)

    print(f'Exported {len(data["weeks"])} weeks')
    print(f'Standings: {len(data["standings"])} teams')
    print(f'Updated at: {data["updated_at"]}')


def get_team_name_for_week(
    abbrev: str, week: int, team_name_overrides: dict, default_name: str
) -> str:
    """Get the team name for a specific week, applying any overrides."""
    if abbrev not in team_name_overrides:
        return default_name

    # Find the most recent name that's effective for this week
    name_entries = team_name_overrides[abbrev]
    current_name = default_name

    for entry in name_entries:
        if entry.get('effective_week', 1) <= week:
            current_name = entry.get('name', default_name)

    return current_name


def apply_team_name_overrides(teams_data: list, week: int, team_name_overrides: dict) -> list:
    """Apply team name overrides for a specific week."""
    if not team_name_overrides:
        return teams_data

    updated_teams = []
    for team in teams_data:
        team_copy = team.copy()
        team_copy['name'] = get_team_name_for_week(
            team['abbrev'], week, team_name_overrides, team.get('name', team['abbrev'])
        )
        updated_teams.append(team_copy)

    return updated_teams


def export_from_json(data_dir: Path, season: int = 2025) -> dict[str, Any]:
    """Export data from JSON files instead of Excel.

    This reads from:
    - data/teams.json - team info
    - data/rosters.json - player rosters
    - data/lineups/{season}/week_X.json - weekly lineups

    And uses the scorer to calculate player scores.
    """
    import sys

    script_dir = Path(__file__).parent
    project_dir = script_dir.parent
    if str(project_dir) not in sys.path:
        sys.path.insert(0, str(project_dir))

    # Load teams
    with open(data_dir / 'teams.json') as f:
        teams_data = json.load(f)['teams']

    # Load team name overrides
    team_names_path = data_dir / 'team_names.json'
    team_name_overrides = {}
    if team_names_path.exists():
        with open(team_names_path) as f:
            team_name_overrides = json.load(f).get('team_names', {})

    teams_by_abbrev = {t['abbrev']: t for t in teams_data}

    # Load rosters
    with open(data_dir / 'rosters.json') as f:
        rosters = json.load(f)

    # First, look for pre-exported week files with full historical roster data
    script_dir = Path(__file__).parent
    web_dir = script_dir.parent / 'web'
    weeks_dir = web_dir / 'data' / 'seasons' / str(season) / 'weeks'
    week_json_files = (
        sorted(weeks_dir.glob('week_*.json'), key=lambda p: int(p.stem.split('_')[1]))
        if weeks_dir.exists()
        else []
    )

    # Find all lineup files (for weeks that don't have pre-exported data)
    lineups_dir = data_dir / 'lineups' / str(season)
    lineup_files = sorted(lineups_dir.glob('week_*.json'), key=lambda p: int(p.stem.split('_')[1]))

    if not lineup_files and not week_json_files:
        print('No lineup or week files found')
        return {}

    # Import scorer
    try:
        from qpfl import QPFLScorer

        scorer_available = True
    except ImportError:
        scorer_available = False
        print('Warning: QPFLScorer not available, scores will be 0')

    weeks = []
    standings = {}
    current_nfl_week = get_current_nfl_week()

    print(f'Current NFL week: {current_nfl_week}')

    # Build a map of which weeks have pre-exported data
    exported_weeks = {int(f.stem.split('_')[1]): f for f in week_json_files}

    # Process all weeks, preferring pre-exported data when available
    all_week_nums = set()
    for f in lineup_files:
        all_week_nums.add(int(f.stem.split('_')[1]))
    for f in week_json_files:
        all_week_nums.add(int(f.stem.split('_')[1]))

    for week_num in sorted(all_week_nums):
        # Use pre-exported week data if available (it has historical roster)
        if week_num in exported_weeks:
            with open(exported_weeks[week_num]) as f:
                week_data = json.load(f)

            # Extract teams from matchups
            teams_for_week = week_data.get('teams', [])
            if not teams_for_week:
                # Fallback: extract from matchups
                for matchup in week_data.get('matchups', []):
                    t1 = matchup.get('team1')
                    t2 = matchup.get('team2')
                    if isinstance(t1, dict):
                        teams_for_week.append(t1)
                    if isinstance(t2, dict):
                        teams_for_week.append(t2)

            weeks.append(
                {
                    'week': week_num,
                    'matchups': week_data.get('matchups', []),
                    'teams': teams_for_week,
                    'has_scores': week_data.get('has_scores', False),
                }
            )

            # Update standings for completed regular season weeks only (not playoffs)
            # Regular season is weeks 1-15 for 2022+
            is_regular_season = week_num <= 15
            if week_data.get('has_scores') and week_num < current_nfl_week and is_regular_season:
                for matchup in week_data.get('matchups', []):
                    t1, t2 = matchup.get('team1'), matchup.get('team2')
                    if isinstance(t1, dict) and isinstance(t2, dict):
                        for team in [t1, t2]:
                            abbrev = team.get('abbrev')
                            if abbrev not in standings:
                                standings[abbrev] = {
                                    'name': team.get('name', abbrev),
                                    'owner': team.get('owner', ''),
                                    'abbrev': abbrev,
                                    'rank_points': 0.0,
                                    'wins': 0,
                                    'losses': 0,
                                    'ties': 0,
                                    'top_half': 0,
                                    'points_for': 0.0,
                                    'points_against': 0.0,
                                }
                            standings[abbrev]['points_for'] += team.get('total_score', 0)

                        s1, s2 = t1.get('total_score', 0), t2.get('total_score', 0)
                        if s1 > s2:
                            standings[t1['abbrev']]['wins'] += 1
                            standings[t2['abbrev']]['losses'] += 1
                            standings[t1['abbrev']]['rank_points'] += 1.0
                        elif s2 > s1:
                            standings[t2['abbrev']]['wins'] += 1
                            standings[t1['abbrev']]['losses'] += 1
                            standings[t2['abbrev']]['rank_points'] += 1.0
                        else:
                            standings[t1['abbrev']]['ties'] += 1
                            standings[t2['abbrev']]['ties'] += 1
                            standings[t1['abbrev']]['rank_points'] += 0.5
                            standings[t2['abbrev']]['rank_points'] += 0.5

                        standings[t1['abbrev']]['points_against'] += s2
                        standings[t2['abbrev']]['points_against'] += s1

                # Add top-half bonus points for this week (0.5 RP for finishing in top half)
                week_teams = [(t.get('abbrev'), t.get('total_score', 0)) for t in teams_for_week]
                week_teams.sort(key=lambda x: x[1], reverse=True)
                num_teams = len(week_teams)
                top_half_cutoff = num_teams // 2

                # Handle ties in scoring for top-half determination
                current_rank = 1
                i = 0
                while i < len(week_teams):
                    current_score = week_teams[i][1]
                    tied_teams = []
                    while i < len(week_teams) and week_teams[i][1] == current_score:
                        tied_teams.append(week_teams[i][0])
                        i += 1

                    tied_positions = list(range(current_rank, current_rank + len(tied_teams)))
                    positions_in_top_half = [p for p in tied_positions if p <= top_half_cutoff]

                    if positions_in_top_half:
                        # 0.5 RP per top-half position, divided among tied teams
                        points_per_team = (0.5 * len(positions_in_top_half)) / len(tied_teams)
                        for abbrev in tied_teams:
                            if abbrev in standings:
                                standings[abbrev]['rank_points'] += points_per_team
                                standings[abbrev]['top_half'] += len(positions_in_top_half) / len(
                                    tied_teams
                                )

                    current_rank += len(tied_teams)

            continue

        # Fall back to reconstructing from lineup + roster for weeks without exported data
        lineup_file = lineups_dir / f'week_{week_num}.json'
        if not lineup_file.exists():
            continue

        with open(lineup_file) as f:
            lineup_data = json.load(f)

        # Create scorer for this week
        scorer = QPFLScorer(season, week_num) if scorer_available else None

        teams_for_week = []

        for abbrev, starters in lineup_data['lineups'].items():
            team_info = teams_by_abbrev.get(abbrev, {})
            roster = rosters.get(abbrev, [])

            # Build roster with scores
            roster_with_scores = []
            total_score = 0.0

            for player in roster:
                player_name = player['name']
                position = player['position']
                nfl_team = player['nfl_team']

                # Check if player is starting
                is_starter = player_name in starters.get(position, [])

                # Calculate score
                score = 0.0
                if scorer:
                    try:
                        result = scorer.score_player(player_name, nfl_team, position)
                        score = result.total_points
                    except Exception:
                        pass  # Score stays 0

                roster_with_scores.append(
                    {
                        'name': player_name,
                        'nfl_team': nfl_team,
                        'position': position,
                        'score': score,
                        'starter': is_starter,
                    }
                )

                if is_starter:
                    total_score += score

            # Apply team name override for this week
            team_name = get_team_name_for_week(
                abbrev, week_num, team_name_overrides, team_info.get('name', abbrev)
            )

            teams_for_week.append(
                {
                    'name': team_name,
                    'owner': team_info.get('owner', ''),
                    'abbrev': abbrev,
                    'roster': roster_with_scores,
                    'total_score': round(total_score, 1),
                }
            )

        # Calculate score_rank
        sorted_by_score = sorted(teams_for_week, key=lambda t: t['total_score'], reverse=True)
        for rank, team in enumerate(sorted_by_score, 1):
            team['score_rank'] = rank

        # Check if week has scores
        has_scores = any(t['total_score'] > 0 for t in teams_for_week)

        # Create matchups based on schedule
        week_matchups = []
        if week_num <= len(SCHEDULE):
            for owner1, owner2 in SCHEDULE[week_num - 1]:
                t1_abbrev = OWNER_TO_CODE.get(owner1)
                t2_abbrev = OWNER_TO_CODE.get(owner2)

                t1 = next((t for t in teams_for_week if t['abbrev'] == t1_abbrev), None)
                t2 = next((t for t in teams_for_week if t['abbrev'] == t2_abbrev), None)

                if t1 and t2:
                    week_matchups.append({'team1': t1, 'team2': t2})

        weeks.append(
            {
                'week': week_num,
                'matchups': week_matchups,
                'teams': teams_for_week,
                'has_scores': has_scores,
            }
        )

        # Update standings only for completed regular season weeks (not playoffs)
        # Regular season is weeks 1-15 for 2022+
        is_regular_season = week_num <= 15
        if has_scores and week_num < current_nfl_week and is_regular_season:
            for matchup in week_matchups:
                t1, t2 = matchup['team1'], matchup['team2']

                for team in [t1, t2]:
                    abbrev = team['abbrev']
                    if abbrev not in standings:
                        standings[abbrev] = {
                            'name': team['name'],
                            'owner': team['owner'],
                            'abbrev': abbrev,
                            'rank_points': 0.0,
                            'wins': 0,
                            'losses': 0,
                            'ties': 0,
                            'top_half': 0,
                            'points_for': 0.0,
                            'points_against': 0.0,
                        }

                # Determine winner and award rank points
                # Win = 1 point, Tie = 0.5 points each
                s1, s2 = t1['total_score'], t2['total_score']
                if s1 > s2:
                    standings[t1['abbrev']]['rank_points'] += 1.0
                    standings[t1['abbrev']]['wins'] += 1
                    standings[t2['abbrev']]['losses'] += 1
                elif s2 > s1:
                    standings[t2['abbrev']]['rank_points'] += 1.0
                    standings[t2['abbrev']]['wins'] += 1
                    standings[t1['abbrev']]['losses'] += 1
                else:
                    standings[t1['abbrev']]['rank_points'] += 0.5
                    standings[t2['abbrev']]['rank_points'] += 0.5
                    standings[t1['abbrev']]['ties'] += 1
                    standings[t2['abbrev']]['ties'] += 1

                standings[t1['abbrev']]['points_for'] += s1
                standings[t1['abbrev']]['points_against'] += s2
                standings[t2['abbrev']]['points_for'] += s2
                standings[t2['abbrev']]['points_against'] += s1

            # Calculate top 5 bonus for each team based on their score_rank
            # Group teams by score to handle ties
            teams_by_score = sorted(teams_for_week, key=lambda x: x['total_score'], reverse=True)

            # Assign ranks handling ties (teams with same score share the rank)
            current_rank = 1
            i = 0
            while i < len(teams_by_score):
                current_score = teams_by_score[i]['total_score']
                tied_teams = []
                while i < len(teams_by_score) and teams_by_score[i]['total_score'] == current_score:
                    tied_teams.append(teams_by_score[i])
                    i += 1

                # Check if any of these tied positions are in top 5
                tied_positions = list(range(current_rank, current_rank + len(tied_teams)))
                positions_in_top5 = [p for p in tied_positions if p <= 5]

                if positions_in_top5:
                    points_per_team = (0.5 * len(positions_in_top5)) / len(tied_teams)

                    for team in tied_teams:
                        abbrev = team['abbrev']
                        if abbrev in standings:
                            standings[abbrev]['rank_points'] += points_per_team
                            standings[abbrev]['top_half'] += 1

                current_rank += len(tied_teams)

    # Sort standings
    standings_list = sorted(
        standings.values(),
        key=lambda s: (s['rank_points'], s['wins'], s['points_for']),
        reverse=True,
    )

    # Populate playoff matchups for weeks 16 and 17 ONLY if they don't already have valid matchups
    # (Pre-exported week JSON files already have correct playoff matchups)
    for week_data in weeks:
        week_num = week_data['week']
        if week_num not in [16, 17]:
            continue

        # Check if week already has valid playoff matchups from source JSON
        existing_matchups = week_data.get('matchups', [])
        has_valid_matchups = (
            len(existing_matchups) > 0
            and any(m.get('bracket') for m in existing_matchups)
            and all(isinstance(m.get('team1'), dict) for m in existing_matchups)
        )

        if has_valid_matchups:
            # Already have correct matchups from source JSON, skip regeneration
            continue

        teams_by_abbrev = {t['abbrev']: t for t in week_data.get('teams', [])}

        # Get playoff matchups from structure
        week_16_results = {}
        if week_num == 17:
            # Calculate week 16 results for determining week 17 matchups
            for w in weeks:
                if w['week'] == 16:
                    for matchup in w.get('matchups', []):
                        game_id = matchup.get('game')
                        t1 = matchup.get('team1', {})
                        t2 = matchup.get('team2', {})
                        s1 = t1.get('total_score', 0) if isinstance(t1, dict) else 0
                        s2 = t2.get('total_score', 0) if isinstance(t2, dict) else 0
                        t1_abbrev = t1.get('abbrev') if isinstance(t1, dict) else t1
                        t2_abbrev = t2.get('abbrev') if isinstance(t2, dict) else t2

                        if game_id and s1 is not None and s2 is not None:
                            if s1 > s2:
                                week_16_results[game_id] = {'winner': t1_abbrev, 'loser': t2_abbrev}
                            elif s2 > s1:
                                week_16_results[game_id] = {'winner': t2_abbrev, 'loser': t1_abbrev}
                    break

        playoff_matchups = get_playoff_matchups(
            standings_list, week_num, week_16_results if week_num == 17 else None
        )

        # Convert abbreviated matchups to full team data
        week_matchups = []
        for pm in playoff_matchups:
            t1_abbrev = pm.get('team1')
            t2_abbrev = pm.get('team2')

            t1 = teams_by_abbrev.get(t1_abbrev)
            t2 = teams_by_abbrev.get(t2_abbrev)

            if t1 and t2:
                matchup = {
                    'team1': t1,
                    'team2': t2,
                    'bracket': pm.get('bracket'),
                    'game': pm.get('game'),
                }
                if 'seed1' in pm:
                    matchup['seed1'] = pm['seed1']
                    matchup['seed2'] = pm['seed2']
                week_matchups.append(matchup)

        week_data['matchups'] = week_matchups

    # Adjust standings for playoff results (1st-4th based on playoffs, rest by regular season)
    standings_list = adjust_standings_for_playoffs_json(standings_list, season, weeks)

    # Determine latest week with data (for team names, lineups, etc.)
    latest_week = max(w['week'] for w in weeks) if weeks else 1

    # Use actual NFL week for current_week so offseason trading logic works (week 18+)
    try:
        nfl_week = nfl.get_current_week()
    except Exception:
        nfl_week = latest_week
    # Use whichever is higher - NFL week or data week
    current_week = max(nfl_week, latest_week)

    # Load FA pool from JSON file
    fa_pool_path = data_dir / 'fa_pool.json'
    fa_pool = []
    if fa_pool_path.exists():
        with open(fa_pool_path) as f:
            fa_pool = json.load(f).get('players', [])

    # Load pending trades
    pending_trades_path = data_dir / 'pending_trades.json'
    pending_trades = []
    if pending_trades_path.exists():
        with open(pending_trades_path) as f:
            pending_trades = json.load(f).get('trades', [])

    # Load trade blocks
    trade_blocks_path = data_dir / 'trade_blocks.json'
    trade_blocks = {}
    if trade_blocks_path.exists():
        with open(trade_blocks_path) as f:
            trade_blocks = json.load(f)

    # Load current week lineups for pending matchups display
    current_lineups = {}
    current_week_lineup_path = lineups_dir / f'week_{latest_week}.json'
    if current_week_lineup_path.exists():
        with open(current_week_lineup_path) as f:
            current_lineups = json.load(f).get('lineups', {})

    # Apply team name overrides to canonical teams (using current week)
    current_teams_data = apply_team_name_overrides(teams_data, latest_week, team_name_overrides)

    return {
        'updated_at': datetime.now(timezone.utc).isoformat(),
        'season': season,
        'current_week': current_week,  # Uses NFL week so offseason trading works (week 18+)
        'teams': current_teams_data,  # Canonical team info (with current week names)
        'rosters': rosters,  # Full roster for each team
        'weeks': weeks,
        'standings': standings_list,
        'schedule': get_schedule_data(standings_list, weeks),
        'game_times': get_game_times(season),
        'team_stats': calculate_team_stats(weeks, standings_list),
        'fa_pool': fa_pool,
        'pending_trades': pending_trades,
        'trade_deadline_week': TRADE_DEADLINE_WEEK,
        'lineups': current_lineups,  # Current week lineup submissions
        'trade_blocks': trade_blocks,
    }


def main_json():
    """Main function using JSON-based data."""
    script_dir = Path(__file__).parent
    project_dir = script_dir.parent
    data_dir = project_dir / 'data'
    web_dir = project_dir / 'web'
    output_path = web_dir / 'data.json'
    shared_dir = web_dir / 'data' / 'shared'

    print('Exporting from JSON files...')
    data = export_from_json(data_dir)

    # Load static data from JSON files (no more Word/Excel parsing)

    # Constitution
    constitution_path = shared_dir / 'constitution.json'
    if constitution_path.exists():
        print('Loading constitution from JSON...')
        with open(constitution_path) as f:
            const_data = json.load(f)
        data['constitution'] = const_data.get('articles', [])

    # Hall of Fame
    hof_json_path = shared_dir / 'hall_of_fame.json'
    if hof_json_path.exists():
        print('Loading Hall of Fame from JSON...')
        with open(hof_json_path) as f:
            hof_stats = json.load(f)
        data['hall_of_fame'] = {
            'finishes_by_year': hof_stats.get('finishes_by_year', []),
            'owner_stats': hof_stats.get('owner_stats', []),
            'player_records': hof_stats.get('player_records', {}),
            'team_records': hof_stats.get('team_records', {}),
            'fun_stats': hof_stats.get('fun_stats', []),
            'rivalry_records': hof_stats.get('rivalry_records', {}),
        }

    # Banners - use existing images
    banners_dir = web_dir / 'images' / 'banners'
    existing_banners = (
        sorted([f.name for f in banners_dir.glob('*_banner.png')]) if banners_dir.exists() else []
    )
    if existing_banners:
        print(f'Using {len(existing_banners)} existing banner images...')
        data['banners'] = existing_banners

    # Draft picks from JSON
    draft_picks_path = data_dir / 'draft_picks.json'
    if draft_picks_path.exists():
        print('Loading draft picks from JSON...')
        with open(draft_picks_path) as f:
            picks_data = json.load(f)
        data['draft_picks'] = picks_data.get('picks', {})

    # Load transactions from unified transaction log (single source of truth)
    print('Loading transactions...')
    data['transactions'] = load_transaction_log()

    # Drafts from JSON
    drafts_path = data_dir / 'drafts.json'
    if drafts_path.exists():
        print('Loading drafts from JSON...')
        with open(drafts_path) as f:
            drafts_data = json.load(f)
        data['drafts'] = drafts_data.get('drafts', [])

    with open(output_path, 'w') as f:
        json.dump(data, f, indent=2)

    print(f'Exported {len(data["weeks"])} weeks')
    print(f'Standings: {len(data["standings"])} teams')
    print(f'Updated at: {data["updated_at"]}')


def export_historical_season(excel_path: str, season: int) -> dict[str, Any]:
    """Export a historical season from Excel to JSON format.

    This is a simplified version for past seasons where:
    - All weeks are completed
    - No live game times needed
    - No lineup merging from JSON files
    - No FA pool or pending trades
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    weeks = []
    standings = {}

    # Find all week sheets (including playoff sheets with special names)
    week_sheets = []
    playoff_sheet_names = {
        'Semi-Finals': 16,
        'Championship': 17,
    }
    for sheet_name in wb.sheetnames:
        match = re.match(r'^Week (\d+)$', sheet_name)
        if match:
            week_sheets.append((int(match.group(1)), sheet_name))
        elif sheet_name in playoff_sheet_names:
            week_sheets.append((playoff_sheet_names[sheet_name], sheet_name))

    # Sort by week number
    week_sheets.sort(key=lambda x: x[0])

    # Export all weeks
    for week_num, sheet_name in week_sheets:
        ws = wb[sheet_name]
        week_data = export_week(ws, week_num, bench_scores=None)
        weeks.append(week_data)

    # Calculate standings from all weeks (all are completed for historical seasons)
    for week_data in weeks:
        if not week_data.get('has_scores', False):
            continue

        # Skip playoff weeks for standings calculation
        if week_data['week'] > 15:
            continue

        for matchup in week_data['matchups']:
            t1, t2 = matchup['team1'], matchup['team2']

            for team in [t1, t2]:
                abbrev = team['abbrev']
                if abbrev not in standings:
                    standings[abbrev] = {
                        'name': team['name'],
                        'owner': team['owner'],
                        'abbrev': abbrev,
                        'rank_points': 0.0,
                        'wins': 0,
                        'losses': 0,
                        'ties': 0,
                        'top_half': 0,
                        'points_for': 0.0,
                        'points_against': 0.0,
                    }
                else:
                    standings[abbrev]['name'] = team['name']
                    standings[abbrev]['owner'] = team['owner']

            s1 = t1['total_score']
            s2 = t2['total_score']

            standings[t1['abbrev']]['points_for'] += s1
            standings[t1['abbrev']]['points_against'] += s2
            standings[t2['abbrev']]['points_for'] += s2
            standings[t2['abbrev']]['points_against'] += s1

            if s1 > s2:
                standings[t1['abbrev']]['rank_points'] += 1.0
                standings[t1['abbrev']]['wins'] += 1
                standings[t2['abbrev']]['losses'] += 1
            elif s2 > s1:
                standings[t2['abbrev']]['rank_points'] += 1.0
                standings[t2['abbrev']]['wins'] += 1
                standings[t1['abbrev']]['losses'] += 1
            else:
                standings[t1['abbrev']]['rank_points'] += 0.5
                standings[t2['abbrev']]['rank_points'] += 0.5
                standings[t1['abbrev']]['ties'] += 1
                standings[t2['abbrev']]['ties'] += 1

        # Calculate top 5 bonus
        teams_by_score = sorted(week_data['teams'], key=lambda x: x['total_score'], reverse=True)

        current_rank = 1
        i = 0
        while i < len(teams_by_score):
            current_score = teams_by_score[i]['total_score']
            tied_teams = []
            while i < len(teams_by_score) and teams_by_score[i]['total_score'] == current_score:
                tied_teams.append(teams_by_score[i])
                i += 1

            tied_positions = list(range(current_rank, current_rank + len(tied_teams)))
            positions_in_top5 = [p for p in tied_positions if p <= 5]

            if positions_in_top5:
                points_per_team = (0.5 * len(positions_in_top5)) / len(tied_teams)

                for team in tied_teams:
                    standings[team['abbrev']]['rank_points'] += points_per_team
                    standings[team['abbrev']]['top_half'] += 1

            current_rank += len(tied_teams)

    # Sort standings by: 1) rank_points, 2) wins (tiebreaker), 3) points_for (second tiebreaker)
    sorted_standings = sorted(
        standings.values(),
        key=lambda x: (x['rank_points'], x['wins'], x['points_for']),
        reverse=True,
    )

    wb.close()

    # Get the final week number
    final_week = max(w['week'] for w in weeks) if weeks else 17

    # Extract team info from the final week's data
    teams_data = []
    if weeks:
        final_week_data = weeks[-1]
        for team in final_week_data.get('teams', []):
            teams_data.append(
                {
                    'name': team['name'],
                    'owner': team['owner'],
                    'abbrev': team['abbrev'],
                }
            )

    return {
        'updated_at': datetime.now(timezone.utc).isoformat().replace('+00:00', 'Z'),
        'season': season,
        'current_week': final_week,
        'is_historical': True,
        'teams': teams_data,
        'weeks': weeks,
        'standings': sorted_standings,
        'schedule': [],  # No schedule needed for historical seasons
        'game_times': {},  # No game times for historical
        'team_stats': calculate_team_stats(weeks, sorted_standings),
    }


def export_historical(season: int):
    """Export a historical season to JSON."""
    script_dir = Path(__file__).parent
    project_dir = script_dir.parent

    excel_path = project_dir / 'previous_seasons' / f'{season} Scores.xlsx'
    output_path = project_dir / 'web' / f'data_{season}.json'

    if not excel_path.exists():
        print(f'Error: {excel_path} not found')
        return

    print(f'Exporting historical season {season}...')
    data = export_historical_season(str(excel_path), season)

    with open(output_path, 'w') as f:
        json.dump(data, f, indent=2)

    print(f'Exported {len(data["weeks"])} weeks to {output_path}')
    print(f'Standings: {len(data["standings"])} teams')


def update_historical_team_stats(season: int):
    """Update team_stats in an existing historical season JSON file.

    This preserves the existing data (which was carefully curated) and only
    recalculates the team_stats field. This is safer than re-exporting from
    Excel, which may have different formats for different seasons.
    """
    script_dir = Path(__file__).parent
    project_dir = script_dir.parent

    json_path = project_dir / 'web' / f'data_{season}.json'

    if not json_path.exists():
        print(f'Warning: {json_path} not found, skipping team_stats update')
        return False

    with open(json_path) as f:
        data = json.load(f)

    weeks = data.get('weeks', [])
    standings = data.get('standings', [])

    if not weeks or not standings:
        print(f'Warning: {season} has no weeks or standings, skipping team_stats update')
        return False

    # Calculate and update team_stats
    data['team_stats'] = calculate_team_stats(weeks, standings)
    data['updated_at'] = datetime.now(timezone.utc).isoformat().replace('+00:00', 'Z')

    with open(json_path, 'w') as f:
        json.dump(data, f, indent=2)

    print(f'Updated team_stats for {season}: {len(data["team_stats"])} teams')
    return True


def export_all_seasons():
    """Export current season and update team_stats for all historical seasons.

    Historical seasons are NOT re-exported from Excel because they have different
    formats. Instead, we update the team_stats field in the existing JSON files.
    """
    script_dir = Path(__file__).parent
    project_dir = script_dir.parent

    # Export current season (2025)
    print('=== Exporting 2025 (current season) ===')
    main()

    # Update team_stats for all historical seasons (without re-parsing Excel)
    historical_seasons = [2020, 2021, 2022, 2023, 2024]
    for season in historical_seasons:
        json_path = project_dir / 'web' / f'data_{season}.json'
        if json_path.exists():
            print(f'\n=== Updating team_stats for {season} ===')
            update_historical_team_stats(season)


if __name__ == '__main__':
    import sys

    if '--json' in sys.argv:
        main_json()
    elif '--all' in sys.argv:
        export_all_seasons()
    elif '--reexport-historical' in sys.argv:
        # Force re-export historical seasons from Excel (use with caution!)
        # This is only needed if the historical Excel files have been fixed
        try:
            idx = sys.argv.index('--reexport-historical')
            season = int(sys.argv[idx + 1])
            print(f'WARNING: Re-exporting {season} from Excel (may break if format differs)')
            export_historical(season)
        except (IndexError, ValueError):
            print('Usage: python export_for_web.py --reexport-historical YEAR')
            sys.exit(1)
    elif '--season' in sys.argv:
        try:
            idx = sys.argv.index('--season')
            season = int(sys.argv[idx + 1])
            if season == 2025:
                main()
            else:
                # For historical seasons, just update team_stats (safer)
                update_historical_team_stats(season)
        except (IndexError, ValueError):
            print('Usage: python export_for_web.py --season YEAR')
            sys.exit(1)
    else:
        main()
