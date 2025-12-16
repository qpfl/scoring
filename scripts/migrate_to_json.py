#!/usr/bin/env python3
"""Migrate Excel data to JSON format."""

import json
import re
from pathlib import Path
import openpyxl

# Position rows in Excel (header_row, player_rows)
POSITION_ROWS = {
    'QB': (6, [7, 8, 9]),
    'RB': (11, [12, 13, 14, 15]),
    'WR': (17, [18, 19, 20, 21, 22]),
    'TE': (24, [25, 26, 27]),
    'K': (29, [30, 31]),
    'D/ST': (33, [34, 35]),
    'HC': (37, [38, 39]),
    'OL': (41, [42, 43]),
}

# Taxi squad rows: (position_row, player_row) pairs
TAXI_ROWS = [(48, 49), (50, 51), (52, 53), (54, 55)]

TEAM_COLUMNS = [1, 3, 5, 7, 9, 11, 13, 15, 17, 19]


def parse_player_name(cell_value: str) -> tuple[str, str]:
    """Parse 'Player Name (TEAM)' into (name, team)."""
    if not cell_value:
        return "", ""
    match = re.match(r'^(.+?)\s*\(([A-Z]{2,3})\)$', cell_value.strip())
    if match:
        return match.group(1).strip(), match.group(2)
    return cell_value.strip(), ""


def migrate_excel_to_json(excel_path: str, output_dir: str):
    """Convert Excel data to JSON files."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    
    # Get all week sheets (including playoff sheets with special names)
    playoff_sheet_names = {'Semi-Finals': 16, 'Championship': 17}
    week_sheets = []
    for name in wb.sheetnames:
        if name.startswith('Week '):
            week_sheets.append((int(name.replace('Week ', '')), name))
        elif name in playoff_sheet_names:
            week_sheets.append((playoff_sheet_names[name], name))
    week_sheets.sort(key=lambda x: x[0])
    week_sheets = [name for _, name in week_sheets]  # Extract just the names in order
    
    # Extract team info and rosters from the most recent week
    latest_week = week_sheets[-1] if week_sheets else None
    if not latest_week:
        print("No week sheets found!")
        return
    
    ws = wb[latest_week]
    
    # Build teams.json and rosters.json
    teams = []
    rosters = {}
    
    for col in TEAM_COLUMNS:
        team_name = ws.cell(row=2, column=col).value
        if not team_name:
            continue
        
        team_name = str(team_name).strip().strip('*')
        owner = str(ws.cell(row=3, column=col).value or "").strip()
        abbrev = str(ws.cell(row=4, column=col).value or "").strip()
        
        # Create owner_key from owner name
        owner_key = owner.lower().replace(' ', '_').replace('/', '_')
        owner_key = re.sub(r'[^a-z0-9_]', '', owner_key)
        
        teams.append({
            "abbrev": abbrev,
            "name": team_name,
            "owner": owner,
            "owner_key": owner_key
        })
        
        # Get roster
        roster = []
        for position, (header_row, player_rows) in POSITION_ROWS.items():
            for row in player_rows:
                player_cell = ws.cell(row=row, column=col)
                if player_cell.value:
                    player_name, nfl_team = parse_player_name(str(player_cell.value))
                    roster.append({
                        "name": player_name,
                        "nfl_team": nfl_team,
                        "position": position
                    })
        
        # Get taxi squad
        taxi_squad = []
        for pos_row, player_row in TAXI_ROWS:
            pos_cell = ws.cell(row=pos_row, column=col)
            player_cell = ws.cell(row=player_row, column=col)
            
            if pos_cell.value and player_cell.value:
                position = str(pos_cell.value).strip()
                player_name, nfl_team = parse_player_name(str(player_cell.value))
                if player_name:
                    taxi_squad.append({
                        "name": player_name,
                        "nfl_team": nfl_team,
                        "position": position
                    })
        
        rosters[abbrev] = {
            "roster": roster,
            "taxi_squad": taxi_squad
        }
    
    # Save teams.json
    teams_path = output_path / "teams.json"
    with open(teams_path, 'w') as f:
        json.dump({"teams": teams}, f, indent=2)
    print(f"Saved {teams_path}")
    
    # Save rosters.json
    rosters_path = output_path / "rosters.json"
    with open(rosters_path, 'w') as f:
        json.dump(rosters, f, indent=2)
    print(f"Saved {rosters_path}")
    
    # Export lineups for each week
    lineups_dir = output_path / "lineups" / "2025"
    lineups_dir.mkdir(parents=True, exist_ok=True)
    
    for week_name in week_sheets:
        week_num = int(week_name.replace('Week ', ''))
        ws = wb[week_name]
        
        week_lineups = {}
        
        for col in TEAM_COLUMNS:
            abbrev = str(ws.cell(row=4, column=col).value or "").strip()
            if not abbrev:
                continue
            
            starters = {pos: [] for pos in POSITION_ROWS.keys()}
            
            for position, (header_row, player_rows) in POSITION_ROWS.items():
                for row in player_rows:
                    player_cell = ws.cell(row=row, column=col)
                    if player_cell.value:
                        player_name, _ = parse_player_name(str(player_cell.value))
                        is_starter = player_cell.font.bold if player_cell.font else False
                        if is_starter:
                            starters[position].append(player_name)
            
            week_lineups[abbrev] = starters
        
        # Save week lineup
        week_path = lineups_dir / f"week_{week_num}.json"
        with open(week_path, 'w') as f:
            json.dump({
                "week": week_num,
                "lineups": week_lineups
            }, f, indent=2)
        print(f"Saved {week_path}")
    
    print("\nMigration complete!")
    print(f"  Teams: {len(teams)}")
    print(f"  Weeks: {len(week_sheets)}")


if __name__ == "__main__":
    import sys
    
    excel_path = sys.argv[1] if len(sys.argv) > 1 else "2025 Scores.xlsx"
    output_dir = sys.argv[2] if len(sys.argv) > 2 else "data"
    
    migrate_excel_to_json(excel_path, output_dir)

