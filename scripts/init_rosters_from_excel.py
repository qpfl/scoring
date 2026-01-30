#!/usr/bin/env python3
"""
Initialize rosters.json from a Rosters.xlsx file in QPFL format.

The Excel uses the standard QPFL column layout:
- Row 2: Team names
- Row 3: Owner names
- Row 4: Team abbreviations (GSA, CGK, etc.)
- Rows 6+: Players by position with headers (QB, RB, WR, etc.)
- Player format: "Player Name (NFL_TEAM)"

Usage:
    python scripts/init_rosters_from_excel.py
    python scripts/init_rosters_from_excel.py --excel "Rosters.xlsx"
"""

import json
import re
import sys
from pathlib import Path

# Add parent directory for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl

from qpfl.constants import POSITION_ROWS, TEAM_COLUMNS


def parse_player_cell(cell_value: str) -> tuple[str, str]:
    """Parse 'Player Name (TEAM)' format, return (name, nfl_team)."""
    if not cell_value:
        return '', ''

    cell_value = str(cell_value).strip()
    match = re.match(r'^(.+?)\s*\(([A-Z]{2,3})\)$', cell_value)
    if match:
        return match.group(1).strip(), match.group(2).strip()
    return cell_value, ''


def init_rosters_from_excel(excel_path: Path, output_path: Path, sheet_name: str = None):
    """
    Initialize rosters.json from QPFL-format Excel file.
    """
    if not excel_path.exists():
        print(f'Error: {excel_path} not found')
        return False

    wb = openpyxl.load_workbook(excel_path, read_only=True)

    # Use first sheet if not specified
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            print(f"Error: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
            wb.close()
            return False
        ws = wb[sheet_name]
    else:
        ws = wb.active
        sheet_name = ws.title

    print(f'Reading from sheet: {sheet_name}')

    # Get team abbreviations from row 4
    team_abbrevs = {}
    for col in TEAM_COLUMNS:
        abbrev = ws.cell(row=4, column=col).value
        if abbrev:
            team_abbrevs[col] = str(abbrev).strip()

    if not team_abbrevs:
        print('Error: No team abbreviations found in row 4')
        wb.close()
        return False

    print(f'Found teams: {list(team_abbrevs.values())}')

    # Parse rosters
    rosters = {abbrev: [] for abbrev in team_abbrevs.values()}

    for position, (_header_row, player_rows) in POSITION_ROWS.items():
        for col, abbrev in team_abbrevs.items():
            for row in player_rows:
                cell_value = ws.cell(row=row, column=col).value
                if not cell_value:
                    continue

                name, nfl_team = parse_player_cell(cell_value)
                if not name:
                    continue

                # For D/ST and HC, the "name" is the team
                if position in ('D/ST', 'HC'):
                    nfl_team = nfl_team or name  # Sometimes format is just "Team"

                rosters[abbrev].append(
                    {
                        'name': name,
                        'position': position,
                        'nfl_team': nfl_team,
                        'status': 'active',
                    }
                )

    wb.close()

    # Write output
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, 'w') as f:
        json.dump(rosters, f, indent=2)

    # Summary
    print(f'\nInitialized rosters from {excel_path}')
    print(f'Output: {output_path}')
    print(f'\nTeams ({len(rosters)}):')
    total = 0
    for team, players in sorted(rosters.items()):
        print(f'  {team}: {len(players)} players')
        total += len(players)
    print(f'\nTotal: {total} players')

    return True


def main():
    import argparse

    parser = argparse.ArgumentParser(description='Initialize rosters.json from QPFL Excel')
    parser.add_argument('--excel', '-e', default='Rosters.xlsx', help='Path to Excel file')
    parser.add_argument('--sheet', '-s', default=None, help='Sheet name (default: first sheet)')
    parser.add_argument('--output', '-o', default='data/rosters.json', help='Output JSON path')
    args = parser.parse_args()

    project_dir = Path(__file__).parent.parent
    excel_path = project_dir / args.excel
    output_path = project_dir / args.output

    init_rosters_from_excel(excel_path, output_path, args.sheet)


if __name__ == '__main__':
    main()
