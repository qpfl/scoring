#!/usr/bin/env python3
"""
Sync JSON lineup submissions to the Excel file.

This script updates the Excel file to mark the correct players as starters (bold)
based on JSON lineup submissions from the website.

This ensures the autoscorer will score the correct players for teams that
submitted their lineups via the website.
"""

import argparse
import json
import re
import sys
from pathlib import Path

# Add parent directory for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Font

from qpfl.constants import POSITION_ROWS, TEAM_COLUMNS


def parse_player_name(cell_value: str) -> str:
    """Extract player name from 'Player Name (TEAM)' format."""
    if not cell_value:
        return ''
    match = re.match(r'^(.+?)\s*\(([A-Z]{2,3})\)$', cell_value.strip())
    if match:
        return match.group(1).strip()
    return cell_value.strip()


def sync_lineups_to_excel(
    excel_path: str, lineup_file: str, sheet_name: str, *, write: bool = True
):
    """
    Sync JSON lineup data to Excel by updating bold formatting.

    Args:
        excel_path: Path to the Excel file
        lineup_file: Path to the JSON lineup file
        sheet_name: Name of the sheet to update
    """
    # Load JSON lineup
    try:
        with open(lineup_file) as f:
            lineup_data = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError) as e:
        print(f'Warning: Could not load lineup file {lineup_file}: {e}')
        return 0

    json_lineups = lineup_data.get('lineups', {})
    if not json_lineups:
        print('No lineups found in JSON file')
        return 0

    # Load Excel workbook
    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception as e:
        print(f'Error loading Excel file {excel_path}: {e}')
        return 0

    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' not found in {excel_path}")
        wb.close()
        return 0

    ws = wb[sheet_name]
    changes = 0

    # Build team abbrev to column mapping
    team_to_col = {}
    for col in TEAM_COLUMNS:
        abbrev_cell = ws.cell(row=4, column=col).value
        if abbrev_cell:
            team_to_col[str(abbrev_cell).strip()] = col

    # Process each team with a JSON lineup
    for abbrev, starters in json_lineups.items():
        if abbrev not in team_to_col:
            print(f'Warning: Team {abbrev} not found in Excel sheet')
            continue

        # Skip teams with empty lineups (they submit via Excel, not website)
        total_starters = sum(len(v) for v in starters.values())
        if total_starters == 0:
            print(f'Skipping {abbrev} - no website lineup (uses Excel)')
            continue

        col = team_to_col[abbrev]
        print(f'Syncing lineup for {abbrev} (column {col})...')

        # Process each position
        for position, (_header_row, player_rows) in POSITION_ROWS.items():
            position_starters = starters.get(position, [])

            for row in player_rows:
                cell = ws.cell(row=row, column=col)
                if not cell.value:
                    continue

                player_name = parse_player_name(str(cell.value))
                should_be_starter = player_name in position_starters

                # Check current bold status
                current_font = cell.font
                is_currently_bold = current_font.bold if current_font else False

                if should_be_starter != is_currently_bold:
                    # Update bold formatting
                    new_font = Font(
                        name=current_font.name or 'Arial',
                        size=current_font.size,
                        bold=should_be_starter,
                        italic=current_font.italic,
                        color=current_font.color,
                    )
                    cell.font = new_font

                    status = 'STARTER' if should_be_starter else 'bench'
                    print(f'  {position}: {player_name} -> {status}')
                    changes += 1

    if changes > 0 and write:
        wb.save(excel_path)
        print(f'\n✓ Saved {changes} changes to {excel_path}')
    elif changes > 0:
        print(f'\nDry run: {changes} changes would be written to {excel_path}')
    else:
        print('\n✓ No changes needed - Excel already matches JSON lineups')

    wb.close()
    return changes


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description='Apply one JSON lineup file to a legacy Excel scoring workbook.'
    )
    parser.add_argument('--season', type=int, required=True, help='Season containing the lineup')
    parser.add_argument('--week', type=int, required=True, choices=range(1, 18), help='Week number')
    parser.add_argument('--excel', type=Path, required=True, help='Excel workbook to update')
    parser.add_argument('--lineup-file', type=Path, help='Override the JSON lineup path')
    parser.add_argument('--sheet', help='Override the target sheet name')
    parser.add_argument(
        '--write',
        action='store_true',
        help='Save workbook changes; without this flag the command is a dry run',
    )
    args = parser.parse_args(argv)

    project_dir = Path(__file__).parent.parent
    excel_path = args.excel
    lineup_file = args.lineup_file or (
        project_dir / 'data' / 'lineups' / str(args.season) / f'week_{args.week}.json'
    )
    sheet_name = args.sheet or f'Week {args.week}'

    print(f'Syncing lineups for Week {args.week}...')
    print(f'  Excel: {excel_path}')
    print(f'  Lineups: {lineup_file}')
    print(f'  Sheet: {sheet_name}')
    print()

    if not lineup_file.exists():
        print(f'No lineup file found at {lineup_file}')
        return 1
    if not excel_path.exists():
        print(f'No Excel workbook found at {excel_path}')
        return 1

    sync_lineups_to_excel(str(excel_path), str(lineup_file), sheet_name, write=args.write)
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
