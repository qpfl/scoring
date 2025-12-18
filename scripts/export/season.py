#!/usr/bin/env python3
"""Export season-specific data.

Exports for a given season:
- meta.json - teams, schedule, trade_deadline
- standings.json - current standings
- rosters.json - full rosters by team
- draft_picks.json - pick ownership
- live.json - game_times, fa_pool, pending_trades (current season only)
- weeks/week_N.json - individual week data
"""

import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import nflreadpy as nfl
import openpyxl

from . import CURRENT_SEASON, DATA_DIR, PROJECT_DIR, SEASONS_DIR, ensure_dirs


# Team mappings
OWNER_TO_CODE = {
    'Griffin': 'GSA',
    'Bill': 'WJK',
    'Ryan': 'RPA',
    'Spencer/Tim': 'S/T',
    'Kaminska': 'CGK',
    'Anagh': 'AST',
    'Connor': 'CWR',
    'Joe/Joe': 'J/J',
    'Stephen': 'SLS',
    'Arnav': 'AYP',
}

ALL_TEAMS = ['GSA', 'WJK', 'RPA', 'S/T', 'CGK', 'AST', 'CWR', 'J/J', 'SLS', 'AYP']

TEAM_ALIASES = {
    'T/S': 'S/T',
    'SPY': 'AYP',
}

# 2025 Schedule
SCHEDULE_2025 = [
    [('Griffin', 'Bill'), ('Ryan', 'Spencer/Tim'), ('Kaminska', 'Anagh'), ('Connor', 'Joe/Joe'), ('Stephen', 'Arnav')],
    [('Griffin', 'Anagh'), ('Ryan', 'Kaminska'), ('Connor', 'Bill'), ('Stephen', 'Joe/Joe'), ('Spencer/Tim', 'Arnav')],
    [('Griffin', 'Joe/Joe'), ('Ryan', 'Arnav'), ('Kaminska', 'Bill'), ('Connor', 'Stephen'), ('Spencer/Tim', 'Anagh')],
    [('Griffin', 'Stephen'), ('Ryan', 'Joe/Joe'), ('Kaminska', 'Spencer/Tim'), ('Connor', 'Anagh'), ('Bill', 'Arnav')],
    [('Griffin', 'Ryan'), ('Connor', 'Kaminska'), ('Bill', 'Joe/Joe'), ('Arnav', 'Anagh'), ('Spencer/Tim', 'Stephen')],
    [('Griffin', 'Arnav'), ('Ryan', 'Anagh'), ('Kaminska', 'Joe/Joe'), ('Connor', 'Spencer/Tim'), ('Stephen', 'Bill')],
    [('Griffin', 'Kaminska'), ('Ryan', 'Stephen'), ('Connor', 'Arnav'), ('Spencer/Tim', 'Bill'), ('Joe/Joe', 'Anagh')],
    [('Griffin', 'Connor'), ('Ryan', 'Bill'), ('Kaminska', 'Arnav'), ('Stephen', 'Anagh'), ('Spencer/Tim', 'Joe/Joe')],
    [('Griffin', 'Spencer/Tim'), ('Ryan', 'Connor'), ('Kaminska', 'Stephen'), ('Joe/Joe', 'Arnav'), ('Anagh', 'Bill')],
    [('Griffin', 'Stephen'), ('Ryan', 'Kaminska'), ('Connor', 'Spencer/Tim'), ('Joe/Joe', 'Bill'), ('Anagh', 'Arnav')],
    [('Griffin', 'Connor'), ('Ryan', 'Arnav'), ('Kaminska', 'Bill'), ('Stephen', 'Joe/Joe'), ('Spencer/Tim', 'Anagh')],
    [('Griffin', 'Arnav'), ('Ryan', 'Anagh'), ('Kaminska', 'Connor'), ('Stephen', 'Bill'), ('Spencer/Tim', 'Joe/Joe')],
    [('Griffin', 'Ryan'), ('Kaminska', 'Joe/Joe'), ('Connor', 'Bill'), ('Stephen', 'Anagh'), ('Spencer/Tim', 'Arnav')],
    [('Griffin', 'Kaminska'), ('Ryan', 'Spencer/Tim'), ('Connor', 'Joe/Joe'), ('Stephen', 'Arnav'), ('Anagh', 'Bill')],
    [('Griffin', 'Bill'), ('Ryan', 'Stephen'), ('Kaminska', 'Spencer/Tim'), ('Connor', 'Arnav'), ('Joe/Joe', 'Anagh')],
]

PLAYOFF_STRUCTURE = {
    16: {
        'round': 'Semifinals',
        'matchups': [
            {'seed1': 1, 'seed2': 4, 'bracket': 'playoffs', 'game': 'semi_1'},
            {'seed1': 2, 'seed2': 3, 'bracket': 'playoffs', 'game': 'semi_2'},
            {'seed1': 5, 'seed2': 6, 'bracket': 'mid_bowl', 'game': 'mid_bowl_week1', 'two_week': True},
            {'seed1': 7, 'seed2': 10, 'bracket': 'sewer_series', 'game': 'sewer_1'},
            {'seed1': 8, 'seed2': 9, 'bracket': 'sewer_series', 'game': 'sewer_2'},
        ]
    },
    17: {
        'round': 'Finals',
        'matchups': [
            {'from_games': ['semi_1', 'semi_2'], 'take': 'winners', 'bracket': 'championship', 'game': 'championship'},
            {'from_games': ['semi_1', 'semi_2'], 'take': 'losers', 'bracket': 'consolation_cup', 'game': 'consolation_cup'},
            {'seed1': 5, 'seed2': 6, 'bracket': 'mid_bowl', 'game': 'mid_bowl_week2', 'two_week': True},
            {'from_games': ['sewer_1', 'sewer_2'], 'take': 'losers', 'bracket': 'toilet_bowl', 'game': 'toilet_bowl'},
        ]
    }
}

# Default position rows for 2024+ format
POSITION_ROWS_2024 = {
    'QB': (6, [7, 8, 9]),
    'RB': (11, [12, 13, 14, 15]),
    'WR': (17, [18, 19, 20, 21, 22]),
    'TE': (24, [25, 26, 27]),
    'K': (29, [30, 31]),
    'D/ST': (33, [34, 35]),
    'HC': (37, [38, 39]),
    'OL': (41, [42, 43]),
}

# Position rows for 2022-2023 format (no OL, different row offsets)
POSITION_ROWS_2022 = {
    'QB': (6, [7, 8, 9]),
    'RB': (11, [12, 13, 14, 15]),
    'WR': (17, [18, 19, 20, 21, 22]),
    'TE': (23, [24, 25, 26]),
    'K': (28, [29, 30]),
    'D/ST': (32, [33, 34]),
    'HC': (36, [37, 38]),
}

# Position rows for 2020-2021 format (8 teams, different layout)
# Note: Team name is in Row 3, owner in Row 4, abbrev in Row 5
POSITION_ROWS_2020 = {
    'QB': (7, [8, 9, 10]),
    'RB': (12, [13, 14, 15, 16]),
    'WR': (18, [19, 20, 21, 22]),
    'TE': (24, [25, 26, 27]),
    'K': (29, [30, 31]),
    'D/ST': (33, [34, 35]),  # Called "DEF" in the file
    'HC': (37, [38, 39]),    # Called "COACH" in the file
}

TAXI_ROWS_2024 = [(48, 49), (50, 51), (52, 53), (54, 55)]
TAXI_ROWS_2022 = [(43, 44), (45, 46), (47, 48), (49, 50)]
# 2020 has "TAXI" header on row 42, so taxi data starts at row 43
TAXI_ROWS_2020 = [(43, 44), (45, 46), (47, 48), (49, 50)]
# 2021 has no TAXI header, taxi data starts at row 42
TAXI_ROWS_2021 = [(42, 43), (44, 45), (46, 47), (48, 49)]


def get_position_rows(season: int) -> dict:
    """Get position row mapping for a season."""
    if season >= 2024:
        return POSITION_ROWS_2024
    elif season >= 2022:
        return POSITION_ROWS_2022
    return POSITION_ROWS_2020


def get_taxi_rows(season: int) -> list:
    """Get taxi squad row mapping for a season."""
    if season >= 2024:
        return TAXI_ROWS_2024
    elif season >= 2022:
        return TAXI_ROWS_2022
    elif season == 2021:
        return TAXI_ROWS_2021
    return TAXI_ROWS_2020


def get_team_columns(season: int) -> list[int]:
    """Get column indices for teams based on season."""
    if season <= 2021:
        return [1, 3, 5, 7, 9, 11, 13, 15]  # 8 teams
    return [1, 3, 5, 7, 9, 11, 13, 15, 17, 19]  # 10 teams


def get_team_info_rows(season: int) -> tuple[int, int, int]:
    """Get (team_name_row, owner_row, abbrev_row) for a season."""
    if season <= 2021:
        return (3, 4, 5)  # 2020/2021: Row 3 = team name, Row 4 = owner, Row 5 = abbrev
    return (2, 3, 4)  # 2022+: Row 2 = team name, Row 3 = owner, Row 4 = abbrev
TEAM_COLUMNS_10 = [1, 3, 5, 7, 9, 11, 13, 15, 17, 19]
TEAM_COLUMNS_8 = [1, 3, 5, 7, 9, 11, 13, 15]
TRADE_DEADLINE_WEEK = 12


def normalize_team_code(team: str) -> str:
    team = team.strip()
    return TEAM_ALIASES.get(team, team)


def parse_player_name(cell_value: str) -> tuple[str, str]:
    if not cell_value:
        return "", ""
    match = re.match(r'^(.+?)\s*\(([A-Z]{2,3})\)$', cell_value.strip())
    if match:
        return match.group(1).strip(), match.group(2)
    return cell_value.strip(), ""


# Cache for NFL roster lookups
_nfl_roster_cache: dict[int, dict] = {}


def _load_nfl_roster(season: int) -> dict:
    """Load and cache NFL roster data for player name lookups."""
    if season in _nfl_roster_cache:
        return _nfl_roster_cache[season]
    
    try:
        rosters = nfl.load_rosters(season)
        # Build lookup dict: (last_name_lower, position) -> (full_name, team)
        # Also store just by last_name for single-match cases
        lookup = {
            'by_last_name': {},  # last_name_lower -> [(full_name, team, position), ...]
            'by_name_position': {},  # (last_name_lower, position) -> (full_name, team)
        }
        
        for row in rosters.iter_rows(named=True):
            full_name = row.get('full_name', '')
            last_name = row.get('last_name', '')
            first_name = row.get('first_name', '')
            team = row.get('team', '')
            position = row.get('position', '')
            
            if not last_name or not full_name:
                continue
                
            last_lower = last_name.lower()
            
            # Store by last name
            if last_lower not in lookup['by_last_name']:
                lookup['by_last_name'][last_lower] = []
            lookup['by_last_name'][last_lower].append((full_name, team, position, first_name))
            
            # Store by (last_name, position)
            key = (last_lower, position)
            if key not in lookup['by_name_position']:
                lookup['by_name_position'][key] = (full_name, team)
        
        _nfl_roster_cache[season] = lookup
        return lookup
    except Exception as e:
        print(f"Warning: Could not load NFL roster for {season}: {e}")
        return {'by_last_name': {}, 'by_name_position': {}}


def expand_legacy_player_name(short_name: str, position: str, season: int) -> tuple[str, str]:
    """
    Expand legacy player name format (e.g., "P. Mahomes II") to full format.
    Returns (full_name, nfl_team).
    """
    if not short_name:
        return "", ""
    
    # Manual overrides for known issues - these persist across regeneration
    # Format: (short_name_pattern, position) -> (full_name, team)
    PLAYER_OVERRIDES = {
        # Player corrections
        ('J. Taylor', 'RB'): ('Jonathan Taylor', 'IND'),
        ('T. Tagovailoa', 'QB'): ('Tua Tagovailoa', 'MIA'),
        ('T. Tagavailoa', 'QB'): ('Tua Tagovailoa', 'MIA'),  # Common typo
        ('J. Williams', 'RB'): ('Jamaal Williams', 'DET'),  # Was with DET in 2021
        
        # Head Coaches - add first names
        ('Tomlin', 'HC'): ('Mike Tomlin', 'PIT'),
        ('M. Tomlin', 'HC'): ('Mike Tomlin', 'PIT'),
        ('Reid', 'HC'): ('Andy Reid', 'KC'),
        ('A. Reid', 'HC'): ('Andy Reid', 'KC'),
        ('Shanahan', 'HC'): ('Kyle Shanahan', 'SF'),
        ('K. Shanahan', 'HC'): ('Kyle Shanahan', 'SF'),
        ('Belichick', 'HC'): ('Bill Belichick', 'NE'),
        ('B. Belichick', 'HC'): ('Bill Belichick', 'NE'),
        ('McVay', 'HC'): ('Sean McVay', 'LAR'),
        ('S. McVay', 'HC'): ('Sean McVay', 'LAR'),
        ('Payton', 'HC'): ('Sean Payton', 'DEN'),
        ('S. Payton', 'HC'): ('Sean Payton', 'DEN'),
        ('Carroll', 'HC'): ('Pete Carroll', 'SEA'),
        ('P. Carroll', 'HC'): ('Pete Carroll', 'SEA'),
        ('Harbaugh', 'HC'): ('John Harbaugh', 'BAL'),
        ('J. Harbaugh', 'HC'): ('John Harbaugh', 'BAL'),
        ('Stefanski', 'HC'): ('Kevin Stefanski', 'CLE'),
        ('K. Stefanski', 'HC'): ('Kevin Stefanski', 'CLE'),
        ('McDermott', 'HC'): ('Sean McDermott', 'BUF'),
        ('S. McDermott', 'HC'): ('Sean McDermott', 'BUF'),
        ('Vrabel', 'HC'): ('Mike Vrabel', 'TEN'),
        ('M. Vrabel', 'HC'): ('Mike Vrabel', 'TEN'),
        ('LaFleur', 'HC'): ('Matt LaFleur', 'GB'),
        ('M. LaFleur', 'HC'): ('Matt LaFleur', 'GB'),
        ('Kingsbury', 'HC'): ('Kliff Kingsbury', 'ARI'),
        ('K. Kingsbury', 'HC'): ('Kliff Kingsbury', 'ARI'),
        ('Reich', 'HC'): ('Frank Reich', 'IND'),
        ('F. Reich', 'HC'): ('Frank Reich', 'IND'),
        ('Rhule', 'HC'): ('Matt Rhule', 'CAR'),
        ('M. Rhule', 'HC'): ('Matt Rhule', 'CAR'),
        ('Flores', 'HC'): ('Brian Flores', 'MIA'),
        ('B. Flores', 'HC'): ('Brian Flores', 'MIA'),
        ('Taylor', 'HC'): ('Zac Taylor', 'CIN'),
        ('Z. Taylor', 'HC'): ('Zac Taylor', 'CIN'),
        ('Staley', 'HC'): ('Brandon Staley', 'LAC'),
        ('B. Staley', 'HC'): ('Brandon Staley', 'LAC'),
        ('Arians', 'HC'): ('Bruce Arians', 'TB'),
        ('B. Arians', 'HC'): ('Bruce Arians', 'TB'),
        ('Meyer', 'HC'): ('Urban Meyer', 'JAC'),
        ('U. Meyer', 'HC'): ('Urban Meyer', 'JAC'),
        ('Judge', 'HC'): ('Joe Judge', 'NYG'),
        ('J. Judge', 'HC'): ('Joe Judge', 'NYG'),
        ('Saleh', 'HC'): ('Robert Saleh', 'NYJ'),
        ('R. Saleh', 'HC'): ('Robert Saleh', 'NYJ'),
        ('Sirianni', 'HC'): ('Nick Sirianni', 'PHI'),
        ('N. Sirianni', 'HC'): ('Nick Sirianni', 'PHI'),
        ('Campbell', 'HC'): ('Dan Campbell', 'DET'),
        ('D. Campbell', 'HC'): ('Dan Campbell', 'DET'),
        ('Rivera', 'HC'): ('Ron Rivera', 'WAS'),
        ('R. Rivera', 'HC'): ('Ron Rivera', 'WAS'),
        ('Zimmer', 'HC'): ('Mike Zimmer', 'MIN'),
        ('M. Zimmer', 'HC'): ('Mike Zimmer', 'MIN'),
        ('Gruden', 'HC'): ('Jon Gruden', 'LV'),
        ('J. Gruden', 'HC'): ('Jon Gruden', 'LV'),
        ('Fangio', 'HC'): ('Vic Fangio', 'DEN'),
        ('V. Fangio', 'HC'): ('Vic Fangio', 'DEN'),
        ('McCarthy', 'HC'): ('Mike McCarthy', 'DAL'),
        ('M. McCarthy', 'HC'): ('Mike McCarthy', 'DAL'),
        ('Lynn', 'HC'): ('Anthony Lynn', 'LAC'),
        ('A. Lynn', 'HC'): ('Anthony Lynn', 'LAC'),
        ('Pederson', 'HC'): ('Doug Pederson', 'PHI'),
        ('D. Pederson', 'HC'): ('Doug Pederson', 'PHI'),
        ('Nagy', 'HC'): ('Matt Nagy', 'CHI'),
        ('M. Nagy', 'HC'): ('Matt Nagy', 'CHI'),
        ('Gase', 'HC'): ('Adam Gase', 'NYJ'),
        ('A. Gase', 'HC'): ('Adam Gase', 'NYJ'),
        ('Patricia', 'HC'): ('Matt Patricia', 'DET'),
        ('M. Patricia', 'HC'): ('Matt Patricia', 'DET'),
        ('O\'Brien', 'HC'): ('Bill O\'Brien', 'HOU'),
        ('B. O\'Brien', 'HC'): ('Bill O\'Brien', 'HOU'),
        ('Marrone', 'HC'): ('Doug Marrone', 'JAC'),
        ('D. Marrone', 'HC'): ('Doug Marrone', 'JAC'),
        ('Quinn', 'HC'): ('Dan Quinn', 'ATL'),
        ('D. Quinn', 'HC'): ('Dan Quinn', 'ATL'),
    }
    
    # D/ST mappings - city/mascot to full name with team code
    DST_OVERRIDES = {
        'Chicago': ('Chicago Bears', 'CHI'),
        'Bears': ('Chicago Bears', 'CHI'),
        'Kansas City': ('Kansas City Chiefs', 'KC'),
        'Chiefs': ('Kansas City Chiefs', 'KC'),
        'San Francisco': ('San Francisco 49ers', 'SF'),
        '49ers': ('San Francisco 49ers', 'SF'),
        'Baltimore': ('Baltimore Ravens', 'BAL'),
        'Ravens': ('Baltimore Ravens', 'BAL'),
        'Pittsburgh': ('Pittsburgh Steelers', 'PIT'),
        'Steelers': ('Pittsburgh Steelers', 'PIT'),
        'Buffalo': ('Buffalo Bills', 'BUF'),
        'Bills': ('Buffalo Bills', 'BUF'),
        'New England': ('New England Patriots', 'NE'),
        'Patriots': ('New England Patriots', 'NE'),
        'Tampa Bay': ('Tampa Bay Buccaneers', 'TB'),
        'Buccaneers': ('Tampa Bay Buccaneers', 'TB'),
        'Los Angeles Rams': ('Los Angeles Rams', 'LAR'),
        'LA Rams': ('Los Angeles Rams', 'LAR'),
        'Rams': ('Los Angeles Rams', 'LAR'),
        'Los Angeles Chargers': ('Los Angeles Chargers', 'LAC'),
        'LA Chargers': ('Los Angeles Chargers', 'LAC'),
        'Chargers': ('Los Angeles Chargers', 'LAC'),
        'Indianapolis': ('Indianapolis Colts', 'IND'),
        'Colts': ('Indianapolis Colts', 'IND'),
        'Miami': ('Miami Dolphins', 'MIA'),
        'Dolphins': ('Miami Dolphins', 'MIA'),
        'Cleveland': ('Cleveland Browns', 'CLE'),
        'Browns': ('Cleveland Browns', 'CLE'),
        'Green Bay': ('Green Bay Packers', 'GB'),
        'Packers': ('Green Bay Packers', 'GB'),
        'Seattle': ('Seattle Seahawks', 'SEA'),
        'Seahawks': ('Seattle Seahawks', 'SEA'),
        'Arizona': ('Arizona Cardinals', 'ARI'),
        'Cardinals': ('Arizona Cardinals', 'ARI'),
        'New Orleans': ('New Orleans Saints', 'NO'),
        'Saints': ('New Orleans Saints', 'NO'),
        'Tennessee': ('Tennessee Titans', 'TEN'),
        'Titans': ('Tennessee Titans', 'TEN'),
        'Denver': ('Denver Broncos', 'DEN'),
        'Broncos': ('Denver Broncos', 'DEN'),
        'Dallas': ('Dallas Cowboys', 'DAL'),
        'Cowboys': ('Dallas Cowboys', 'DAL'),
        'Minnesota': ('Minnesota Vikings', 'MIN'),
        'Vikings': ('Minnesota Vikings', 'MIN'),
        'Atlanta': ('Atlanta Falcons', 'ATL'),
        'Falcons': ('Atlanta Falcons', 'ATL'),
        'Las Vegas': ('Las Vegas Raiders', 'LV'),
        'Raiders': ('Las Vegas Raiders', 'LV'),
        'Philadelphia': ('Philadelphia Eagles', 'PHI'),
        'Eagles': ('Philadelphia Eagles', 'PHI'),
        'New York Giants': ('New York Giants', 'NYG'),
        'Giants': ('New York Giants', 'NYG'),
        'New York Jets': ('New York Jets', 'NYJ'),
        'Jets': ('New York Jets', 'NYJ'),
        'Washington': ('Washington Commanders', 'WAS'),
        'Commanders': ('Washington Commanders', 'WAS'),
        'Football Team': ('Washington Football Team', 'WAS'),
        'Carolina': ('Carolina Panthers', 'CAR'),
        'Panthers': ('Carolina Panthers', 'CAR'),
        'Detroit': ('Detroit Lions', 'DET'),
        'Lions': ('Detroit Lions', 'DET'),
        'Jacksonville': ('Jacksonville Jaguars', 'JAC'),
        'Jaguars': ('Jacksonville Jaguars', 'JAC'),
        'Cincinnati': ('Cincinnati Bengals', 'CIN'),
        'Bengals': ('Cincinnati Bengals', 'CIN'),
        'Houston': ('Houston Texans', 'HOU'),
        'Texans': ('Houston Texans', 'HOU'),
    }
    
    # Check if already has a team code - extract it but still try to expand the name
    existing_team = ""
    match = re.match(r'^(.+?)\s*\(([A-Z]{2,3})\)$', short_name.strip())
    if match:
        short_name = match.group(1).strip()
        existing_team = match.group(2)
    
    # Parse legacy format: "F. Lastname" or "F. Lastname Jr." or "Lastname" (for DEF)
    short_name = short_name.strip()
    
    # Check player overrides first
    override_key = (short_name, position)
    if override_key in PLAYER_OVERRIDES:
        return PLAYER_OVERRIDES[override_key]
    
    # Handle defense names (just city/team name)
    if position in ('D/ST', 'DEF'):
        if short_name in DST_OVERRIDES:
            return DST_OVERRIDES[short_name]
        return short_name, existing_team
    
    # Handle coach names - they're already in "F. Lastname" format
    # Try to extract last name
    parts = short_name.split()
    if not parts:
        return short_name, ""
    
    # Find the main last name (ignoring suffixes like Jr., II, III)
    suffixes = {'jr.', 'jr', 'sr.', 'sr', 'ii', 'iii', 'iv', 'v'}
    last_name = None
    first_initial = None
    
    for i, part in enumerate(parts):
        if part.endswith('.') and len(part) == 2 and i == 0:
            first_initial = part[0].upper()
        elif part.lower() not in suffixes and not part.endswith('.'):
            last_name = part
    
    if not last_name and len(parts) > 0:
        # Fallback: use last part that's not a suffix
        for part in reversed(parts):
            if part.lower() not in suffixes:
                last_name = part
                break
    
    if not last_name:
        return short_name, existing_team
    
    # Look up in NFL roster
    roster = _load_nfl_roster(season)
    last_lower = last_name.lower()
    
    # Try to find by last name first, then filter by initial and position
    if last_lower in roster['by_last_name']:
        candidates = roster['by_last_name'][last_lower]
        
        # Filter by first initial if we have one
        if first_initial:
            matching = [c for c in candidates if c[3] and c[3][0].upper() == first_initial]
            if len(matching) == 1:
                return matching[0][0], matching[0][1] or existing_team
            elif len(matching) > 1:
                # Multiple matches with same initial - try to match position
                pos_match = [c for c in matching if c[2] == position]
                if pos_match:
                    return pos_match[0][0], pos_match[0][1] or existing_team
                return matching[0][0], matching[0][1] or existing_team
        
        # No first initial - try by position
        pos_match = [c for c in candidates if c[2] == position]
        if len(pos_match) == 1:
            return pos_match[0][0], pos_match[0][1] or existing_team
        
        # Multiple position matches or no position match - return first candidate
        if len(candidates) == 1:
            return candidates[0][0], candidates[0][1] or existing_team
        
        # Multiple matches - try position filter
        pos_match = [c for c in candidates if c[2] == position]
        if pos_match:
            return pos_match[0][0], pos_match[0][1] or existing_team
    
    # Could not expand - return original with existing team if available
    return short_name, existing_team


def load_teams() -> list[dict]:
    """Load canonical team info from teams.json."""
    teams_path = DATA_DIR / 'teams.json'
    if teams_path.exists():
        with open(teams_path) as f:
            return json.load(f).get("teams", [])
    return []


def load_rosters() -> dict[str, list[dict]]:
    """Load full rosters from rosters.json."""
    rosters_path = DATA_DIR / 'rosters.json'
    if rosters_path.exists():
        with open(rosters_path) as f:
            return json.load(f)
    return {}


def load_pending_trades() -> list[dict]:
    """Load pending trades from JSON file."""
    pending_path = DATA_DIR / 'pending_trades.json'
    if pending_path.exists():
        with open(pending_path) as f:
            return json.load(f).get("trades", [])
    return []


def load_fa_pool() -> list[dict]:
    """Load FA pool from JSON file."""
    fa_path = DATA_DIR / 'fa_pool.json'
    if fa_path.exists():
        with open(fa_path) as f:
            return json.load(f).get("players", [])
    return []


def get_current_nfl_week() -> int:
    return nfl.get_current_week()


def get_game_times(season: int) -> dict[int, dict[str, str]]:
    """Get game kickoff times for each team by week."""
    try:
        schedule = nfl.load_schedules(seasons=season)
        game_times = {}
        
        for week in range(1, 19):
            week_games = schedule.filter(schedule['week'] == week)
            if week_games.height == 0:
                continue
                
            game_times[week] = {}
            
            for row in week_games.iter_rows(named=True):
                game_date = row.get('gameday', '')
                game_time = row.get('gametime', '')
                
                if game_date and game_time:
                    try:
                        dt_str = f"{game_date} {game_time}"
                        dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M")
                        kickoff_iso = dt.strftime("%Y-%m-%dT%H:%M:00-05:00")
                        
                        home_team = row.get('home_team', '')
                        away_team = row.get('away_team', '')
                        
                        if home_team:
                            game_times[week][home_team] = kickoff_iso
                        if away_team:
                            game_times[week][away_team] = kickoff_iso
                    except (ValueError, TypeError):
                        pass
        
        return game_times
    except Exception as e:
        print(f"Warning: Could not load game times: {e}")
        return {}


def get_schedule_data(standings: list[dict] = None) -> list[dict]:
    """Convert schedule to JSON format."""
    schedule_data = []
    
    for week_num, matchups in enumerate(SCHEDULE_2025, 1):
        week_matchups = []
        for owner1, owner2 in matchups:
            week_matchups.append({
                'team1': OWNER_TO_CODE.get(owner1, owner1),
                'team2': OWNER_TO_CODE.get(owner2, owner2),
            })
        schedule_data.append({
            'week': week_num,
            'is_rivalry': week_num == 5,
            'is_playoffs': False,
            'matchups': week_matchups,
        })
    
    if standings:
        seed_to_team = {i + 1: team['abbrev'] for i, team in enumerate(standings)}
        
        for week_num in [16, 17]:
            playoff_info = PLAYOFF_STRUCTURE[week_num]
            week_matchups = []
            
            for game in playoff_info['matchups']:
                matchup = {'bracket': game['bracket'], 'game': game['game']}
                
                if 'seed1' in game:
                    matchup['team1'] = seed_to_team.get(game['seed1'], 'TBD')
                    matchup['team2'] = seed_to_team.get(game['seed2'], 'TBD')
                    matchup['seed1'] = game['seed1']
                    matchup['seed2'] = game['seed2']
                else:
                    matchup['team1'] = 'TBD'
                    matchup['team2'] = 'TBD'
                
                week_matchups.append(matchup)
            
            schedule_data.append({
                'week': week_num,
                'is_rivalry': False,
                'is_playoffs': True,
                'playoff_round': playoff_info['round'],
                'matchups': week_matchups,
            })
    
    return schedule_data


def parse_draft_picks(excel_path: str) -> dict[str, dict]:
    """Parse traded picks and calculate what picks each team owns."""
    DEFAULT_OFFSEASON = list(range(1, 7))
    DEFAULT_TAXI = list(range(1, 5))
    DEFAULT_WAIVER = list(range(1, 5))
    SEASONS = ['2026', '2027', '2028', '2029']
    
    picks = {}
    for team in ALL_TEAMS:
        picks[team] = {}
        for season in SEASONS:
            picks[team][season] = {
                'offseason': [(r, team) for r in DEFAULT_OFFSEASON],
                'offseason_taxi': [(r, team) for r in DEFAULT_TAXI],
                'waiver': [(r, team) for r in DEFAULT_WAIVER],
                'waiver_taxi': [(r, team) for r in DEFAULT_TAXI],
            }
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
    except Exception as e:
        print(f"Warning: Could not load traded picks: {e}")
        return picks
    
    pattern = r'([A-Z/]+)\s+holds\s+([A-Z/]+)\s+(\d+)(?:st|nd|rd|th)\s+(rounder|round taxi|round waiver)'
    season_cols = {1: '2026', 2: '2027', 3: '2028', 4: '2029'}
    draft_type = {1: 'offseason', 2: 'offseason', 3: 'offseason', 4: 'offseason'}
    
    trades = []
    
    for row_num in range(5, 50):
        for col in range(1, 5):
            cell_val = ws.cell(row=row_num, column=col).value
            if not cell_val:
                continue
            
            cell_str = str(cell_val).strip()
            
            if cell_str == 'Offseason Draft':
                draft_type[col] = 'offseason'
                continue
            elif cell_str == 'Waiver Draft':
                draft_type[col] = 'waiver'
                continue
            
            if cell_str.startswith('*'):
                continue
            
            match = re.search(pattern, cell_str, re.IGNORECASE)
            if match:
                holder = normalize_team_code(match.group(1))
                original = normalize_team_code(match.group(2))
                round_num = int(match.group(3))
                pick_type_str = match.group(4).lower()
                
                season = season_cols.get(col, '2026')
                
                if pick_type_str == 'rounder':
                    pick_type = draft_type[col]
                elif pick_type_str == 'round taxi':
                    pick_type = f'{draft_type[col]}_taxi'
                elif pick_type_str == 'round waiver':
                    pick_type = 'waiver'
                else:
                    continue
                
                trades.append((season, holder, original, round_num, pick_type))
    
    wb.close()
    
    for season, holder, original, round_num, pick_type in trades:
        if holder not in ALL_TEAMS or original not in ALL_TEAMS:
            continue
        if season not in SEASONS:
            continue
        
        original_picks = picks[original][season][pick_type]
        pick_to_remove = None
        for i, (r, owner) in enumerate(original_picks):
            if r == round_num and owner == original:
                pick_to_remove = i
                break
        if pick_to_remove is not None:
            original_picks.pop(pick_to_remove)
        
        picks[holder][season][pick_type].append((round_num, original))
    
    formatted = {}
    for team in ALL_TEAMS:
        formatted[team] = {}
        for season in SEASONS:
            formatted[team][season] = {}
            for draft_type_key in ['offseason', 'offseason_taxi', 'waiver', 'waiver_taxi']:
                team_picks = sorted(picks[team][season][draft_type_key], key=lambda x: (x[0], x[1]))
                # 'own' is True for all picks in this team's list since they currently hold them
                # 'from' indicates the original owner (for display purposes like "2026 Round 1 (from AST)")
                formatted[team][season][draft_type_key] = [
                    {'round': r, 'from': original_owner, 'own': True}
                    for r, original_owner in team_picks
                ]
    
    return formatted


def calculate_bench_scores(excel_path: str, sheet_name: str, season: int, week_num: int) -> dict:
    """Calculate scores for bench players using the scorer.
    
    Returns:
        Dict mapping (team_abbrev, player_name) -> score
        
    Note: Only works for 2022+ seasons due to different Excel formats in earlier years.
    """
    # Skip legacy seasons - the excel parser doesn't support their format
    if season < 2022:
        return {}
    
    import sys
    script_dir = Path(__file__).parent
    project_dir = script_dir.parent.parent
    if str(project_dir) not in sys.path:
        sys.path.insert(0, str(project_dir))
    
    try:
        from qpfl import QPFLScorer
        from qpfl.excel_parser import parse_roster_from_excel
    except ImportError:
        return {}
    
    try:
        teams = parse_roster_from_excel(excel_path, sheet_name)
        scorer = QPFLScorer(season, week_num)
        
        bench_scores = {}
        for team in teams:
            for position, players in team.players.items():
                for player_name, nfl_team, is_started in players:
                    if not is_started:  # Only calculate for bench players
                        try:
                            result = scorer.score_player(player_name, nfl_team, position)
                            bench_scores[(team.abbreviation, player_name)] = result.total_points
                        except Exception:
                            pass  # Skip if scoring fails
        
        return bench_scores
    except Exception as e:
        print(f"    Warning: Could not calculate bench scores for week {week_num}: {e}")
        return {}


def export_week(ws, week_num: int, season: int = 2025, bench_scores: dict = None) -> dict[str, Any]:
    """Export a single week's data from Excel.
    
    Args:
        ws: Excel worksheet
        week_num: Week number
        season: Season year (affects layout detection)
        bench_scores: Optional dict mapping (team_abbrev, player_name) -> score for bench players
    """
    teams_data = []
    position_rows = get_position_rows(season)
    taxi_rows = get_taxi_rows(season)
    team_columns = get_team_columns(season)
    team_name_row, owner_row, abbrev_row = get_team_info_rows(season)
    
    # For legacy seasons (2020/2021), we expand player names using NFL roster data
    use_legacy_expansion = season <= 2021
    
    for i, col in enumerate(team_columns):
        team_name = ws.cell(row=team_name_row, column=col).value
        if not team_name:
            continue
        
        team_name = str(team_name).strip().strip('*')
        owner = ws.cell(row=owner_row, column=col).value or ""
        abbrev = ws.cell(row=abbrev_row, column=col).value or ""
        
        roster = []
        total_score = 0.0
        
        for position, (header_row, player_rows) in position_rows.items():
            for row in player_rows:
                player_cell = ws.cell(row=row, column=col)
                score_cell = ws.cell(row=row, column=col + 1)
                
                if player_cell.value:
                    cell_value = str(player_cell.value)
                    
                    # Parse player name based on season format
                    if use_legacy_expansion:
                        player_name, nfl_team = expand_legacy_player_name(cell_value, position, season)
                    else:
                        player_name, nfl_team = parse_player_name(cell_value)
                    is_starter = player_cell.font.bold if player_cell.font else False
                    try:
                        excel_score = float(score_cell.value) if score_cell.value else 0.0
                    except (ValueError, TypeError):
                        excel_score = 0.0  # Handle "BYE" or other non-numeric values
                    
                    # For bench players, use calculated score if available
                    if is_starter:
                        score = excel_score
                    elif bench_scores and (abbrev, player_name) in bench_scores:
                        score = bench_scores[(abbrev, player_name)]
                    else:
                        score = excel_score
                    
                    roster.append({
                        'name': player_name,
                        'nfl_team': nfl_team,
                        'position': position,
                        'score': score,
                        'starter': is_starter,
                    })
                    
                    if is_starter:
                        total_score += excel_score  # Always use Excel score for total
        
        taxi_squad = []
        for pos_row, player_row in taxi_rows:
            pos_cell = ws.cell(row=pos_row, column=col)
            player_cell = ws.cell(row=player_row, column=col)
            
            if pos_cell.value and player_cell.value:
                taxi_position = str(pos_cell.value).strip()
                cell_value = str(player_cell.value)
                
                if use_legacy_expansion:
                    player_name, nfl_team = expand_legacy_player_name(cell_value, taxi_position, season)
                else:
                    player_name, nfl_team = parse_player_name(cell_value)
                
                if player_name:
                    taxi_squad.append({
                        'name': player_name,
                        'nfl_team': nfl_team,
                        'position': taxi_position,
                    })
        
        teams_data.append({
            'name': team_name,
            'owner': owner,
            'abbrev': abbrev,
            'roster': roster,
            'taxi_squad': taxi_squad,
            'total_score': round(total_score, 1),
        })
    
    sorted_by_score = sorted(teams_data, key=lambda t: t['total_score'], reverse=True)
    for rank, team in enumerate(sorted_by_score, 1):
        team['score_rank'] = rank
    
    matchups = []
    
    # Playoff bracket assignments by matchup index
    # Based on standard QPFL playoff structure
    if season == 2020:
        # 2020: 8-team structure with Jamboree instead of Sewer Series
        # Week 15 = Semifinals + Jamboree Week 1
        # Week 16 = Finals + Jamboree Week 2
        # Jamboree is a 2-week total points contest for non-playoff teams
        playoff_brackets = {
            15: ['playoffs', 'playoffs', 'jamboree', 'jamboree'],
            16: ['championship', 'consolation_cup', 'jamboree', 'jamboree'],
        }
    elif season == 2021:
        # 8-team playoff structure (2021)
        # Week 15 = Semifinals (Playoffs, Week 1 / Playoffs, Round 1)
        # Week 16 = Finals (Super Bowl Week)
        playoff_brackets = {
            15: ['playoffs', 'playoffs', 'sewer_series', 'sewer_series'],
            16: ['championship', 'consolation_cup', 'toilet_bowl', None],
        }
    else:
        # 10-team playoff structure (2022+)
        # Week 16 = Semifinals
        # Week 17 = Finals
        playoff_brackets = {
            16: ['playoffs', 'playoffs', 'mid_bowl', 'sewer_series', 'sewer_series'],
            17: ['championship', 'consolation_cup', 'mid_bowl', 'toilet_bowl', None],
        }
    
    for i in range(0, len(teams_data), 2):
        if i + 1 < len(teams_data):
            matchup = {
                'team1': teams_data[i],
                'team2': teams_data[i + 1],
            }
            
            # Add bracket info for playoff weeks
            if week_num in playoff_brackets:
                matchup_idx = i // 2
                brackets = playoff_brackets[week_num]
                if matchup_idx < len(brackets) and brackets[matchup_idx]:
                    matchup['bracket'] = brackets[matchup_idx]
            
            matchups.append(matchup)
    
    has_scores = any(t['total_score'] > 0 for t in teams_data)
    
    return {
        'week': week_num,
        'matchups': matchups,
        'teams': teams_data,
        'has_scores': has_scores,
    }


def calculate_standings(weeks: list[dict], max_week: int = None, season: int = None) -> list[dict]:
    """Calculate standings from week data."""
    standings = {}
    
    # Determine regular season length based on season
    # 2020/2021: 14 regular season weeks (playoffs weeks 15-16)
    # 2022+: 15 regular season weeks (playoffs weeks 16-17)
    if season and season <= 2021:
        regular_season_weeks = 14
    else:
        regular_season_weeks = 15
    
    for week_data in weeks:
        if not week_data.get('has_scores', False):
            continue
        
        if max_week and week_data['week'] >= max_week:
            continue
        
        # Skip playoff weeks
        if week_data['week'] > regular_season_weeks:
            continue
        
        for matchup in week_data['matchups']:
            t1, t2 = matchup['team1'], matchup['team2']
            
            for team in [t1, t2]:
                abbrev = team['abbrev']
                if abbrev not in standings:
                    standings[abbrev] = {
                        'name': team['name'],
                        'owner': team['owner'],
                        'abbrev': abbrev,
                        'rank_points': 0.0,
                        'wins': 0,
                        'losses': 0,
                        'ties': 0,
                        'top_half': 0,
                        'points_for': 0.0,
                        'points_against': 0.0,
                    }
                else:
                    standings[abbrev]['name'] = team['name']
                    standings[abbrev]['owner'] = team['owner']
            
            s1 = t1['total_score']
            s2 = t2['total_score']
            
            standings[t1['abbrev']]['points_for'] += s1
            standings[t1['abbrev']]['points_against'] += s2
            standings[t2['abbrev']]['points_for'] += s2
            standings[t2['abbrev']]['points_against'] += s1
            
            if s1 > s2:
                standings[t1['abbrev']]['rank_points'] += 1.0
                standings[t1['abbrev']]['wins'] += 1
                standings[t2['abbrev']]['losses'] += 1
            elif s2 > s1:
                standings[t2['abbrev']]['rank_points'] += 1.0
                standings[t2['abbrev']]['wins'] += 1
                standings[t1['abbrev']]['losses'] += 1
            else:
                standings[t1['abbrev']]['rank_points'] += 0.5
                standings[t2['abbrev']]['rank_points'] += 0.5
                standings[t1['abbrev']]['ties'] += 1
                standings[t2['abbrev']]['ties'] += 1
        
        teams_by_score = sorted(week_data['teams'], key=lambda x: x['total_score'], reverse=True)
        
        # Top half cutoff: 4 for 8-team leagues, 5 for 10-team leagues
        num_teams = len(teams_by_score)
        top_half_cutoff = num_teams // 2
        
        current_rank = 1
        i = 0
        while i < len(teams_by_score):
            current_score = teams_by_score[i]['total_score']
            tied_teams = []
            while i < len(teams_by_score) and teams_by_score[i]['total_score'] == current_score:
                tied_teams.append(teams_by_score[i])
                i += 1
            
            tied_positions = list(range(current_rank, current_rank + len(tied_teams)))
            positions_in_top_half = [p for p in tied_positions if p <= top_half_cutoff]
            
            if positions_in_top_half:
                # Top half rank points formula changed:
                # 2020-2021: 1 RP per top-half finish
                # 2022+: 0.5 RP per top-half finish
                top_half_rp_multiplier = 1.0 if season and season <= 2021 else 0.5
                points_per_team = (top_half_rp_multiplier * len(positions_in_top_half)) / len(tied_teams)
                
                for team in tied_teams:
                    standings[team['abbrev']]['rank_points'] += points_per_team
                    standings[team['abbrev']]['top_half'] += len(positions_in_top_half) / len(tied_teams)
            
            current_rank += len(tied_teams)
    
    return sorted(
        standings.values(),
        key=lambda x: (x['rank_points'], x['points_for']),
        reverse=True
    )


def calculate_jamboree(weeks: list[dict]) -> list[dict]:
    """Calculate 2020 Jamboree standings (2-week total points for non-playoff teams).
    
    The Jamboree was a 2-week total points contest in weeks 15-16 for the 4 teams 
    that didn't make the playoffs. MPA (KiloCorp) won.
    """
    # Find week 15 and 16 data
    week_15 = next((w for w in weeks if w['week'] == 15), None)
    week_16 = next((w for w in weeks if w['week'] == 16), None)
    
    if not week_15 or not week_16:
        return []
    
    # Get jamboree teams (those with 'jamboree' bracket in week 15)
    jamboree_teams = set()
    for matchup in week_15.get('matchups', []):
        if matchup.get('bracket') == 'jamboree':
            jamboree_teams.add(matchup['team1']['abbrev'])
            jamboree_teams.add(matchup['team2']['abbrev'])
    
    if not jamboree_teams:
        return []
    
    # Calculate total points for each jamboree team over weeks 15-16
    team_totals = {}
    for week_data in [week_15, week_16]:
        for team in week_data.get('teams', []):
            abbrev = team.get('abbrev')
            if abbrev in jamboree_teams:
                if abbrev not in team_totals:
                    team_totals[abbrev] = {
                        'abbrev': abbrev,
                        'name': team.get('name', ''),
                        'owner': team.get('owner', ''),
                        'week_15': 0.0,
                        'week_16': 0.0,
                        'total': 0.0,
                    }
                score = team.get('total_score', 0.0)
                if week_data['week'] == 15:
                    team_totals[abbrev]['week_15'] = score
                else:
                    team_totals[abbrev]['week_16'] = score
                team_totals[abbrev]['total'] += score
    
    # Sort by total points (highest first)
    standings = sorted(team_totals.values(), key=lambda x: x['total'], reverse=True)
    
    # Add placement
    for i, team in enumerate(standings):
        team['place'] = i + 1
    
    return standings


# Mapping of owner names to team abbreviations for adjusting standings
OWNER_TO_ABBREV = {
    'Griffin Ansel': 'GSA',
    'Griff': 'GSA',
    'Connor Reardon': 'CGK',
    'Connor Kaminska': 'CGK',
    'Kaminska': 'CGK',
    'Redacted Kaminska': 'CGK',
    'Connor': 'CGK',
    'Ryan Ansel': 'RPA',
    'Ryan': 'RPA',
    'Stephen Schmidt': 'SLS',
    'Stephen': 'SLS',
    'Joe Ward': 'JRW',
    'Joe Kuhl': 'JRW',
    'Censored Ward': 'JRW',
    'Spencer/Tim': 'S/T',
    'Tim/Spencer': 'S/T',
    'Arnav Patel': 'AYP',
    'Arnav': 'AYP',
    'Anagh Talasila': 'AST',
    'Anagh': 'AST',
    'Bill Kuhl': 'WJK',
    'Bill': 'WJK',
    'Miles Agus': 'MPA',
    'Miles': 'MPA',
    'Joe/Joe': 'J/J',
    'Joe/Censored': 'J/J',
}


def adjust_standings_for_playoffs(standings: list[dict], season: int, weeks: list[dict]) -> list[dict]:
    """Adjust historical standings to reflect final playoff positions.
    
    For historical seasons, reorder standings based on playoff results:
    - 1st: Championship winner
    - 2nd: Championship loser
    - 3rd: Consolation cup winner
    - 4th: Consolation cup loser
    - 5th+: Based on regular season (non-playoff teams)
    """
    if not weeks:
        return standings
    
    # Find the final week (championship week)
    final_week_num = 16 if season <= 2021 else 17
    final_week = next((w for w in weeks if w['week'] == final_week_num), None)
    
    if not final_week:
        return standings
    
    # Find championship and consolation cup matchups
    championship_matchup = None
    consolation_matchup = None
    
    for matchup in final_week.get('matchups', []):
        bracket = matchup.get('bracket', '')
        if bracket == 'championship':
            championship_matchup = matchup
        elif bracket == 'consolation_cup':
            consolation_matchup = matchup
    
    if not championship_matchup:
        return standings
    
    # Determine winners and losers
    t1 = championship_matchup['team1']
    t2 = championship_matchup['team2']
    champ_abbrev = t1['abbrev'] if t1['total_score'] > t2['total_score'] else t2['abbrev']
    runner_up_abbrev = t2['abbrev'] if t1['total_score'] > t2['total_score'] else t1['abbrev']
    
    third_abbrev = None
    fourth_abbrev = None
    if consolation_matchup:
        t1 = consolation_matchup['team1']
        t2 = consolation_matchup['team2']
        third_abbrev = t1['abbrev'] if t1['total_score'] > t2['total_score'] else t2['abbrev']
        fourth_abbrev = t2['abbrev'] if t1['total_score'] > t2['total_score'] else t1['abbrev']
    
    # Create a mapping of abbrev to desired position
    playoff_positions = {}
    playoff_positions[champ_abbrev] = 1
    playoff_positions[runner_up_abbrev] = 2
    if third_abbrev:
        playoff_positions[third_abbrev] = 3
    if fourth_abbrev:
        playoff_positions[fourth_abbrev] = 4
    
    # Separate playoff teams from non-playoff teams
    playoff_teams = [s for s in standings if s['abbrev'] in playoff_positions]
    non_playoff_teams = [s for s in standings if s['abbrev'] not in playoff_positions]
    
    # Sort playoff teams by their playoff position
    playoff_teams.sort(key=lambda x: playoff_positions.get(x['abbrev'], 999))
    
    # Combine: playoff teams first (in playoff order), then non-playoff teams (in regular season order)
    adjusted = playoff_teams + non_playoff_teams
    
    return adjusted


def export_season(season: int, excel_path: str = None):
    """Export all data for a season."""
    ensure_dirs()
    
    season_dir = SEASONS_DIR / str(season)
    weeks_dir = season_dir / "weeks"
    season_dir.mkdir(parents=True, exist_ok=True)
    weeks_dir.mkdir(parents=True, exist_ok=True)
    
    is_current = season == CURRENT_SEASON
    
    if excel_path is None:
        if is_current:
            excel_path = PROJECT_DIR / f"{season} Scores.xlsx"
        else:
            excel_path = PROJECT_DIR / "previous_seasons" / f"{season} Scores.xlsx"
    
    if not Path(excel_path).exists():
        print(f"Error: Excel file not found: {excel_path}")
        return
    
    print(f"Exporting season {season}...")
    
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    
    # Find all week sheets
    week_sheets = []
    
    # Map for written-out week numbers
    word_to_num = {
        'one': 1, 'two': 2, 'three': 3, 'four': 4, 'five': 5,
        'six': 6, 'seven': 7, 'eight': 8, 'nine': 9, 'ten': 10,
        'eleven': 11, 'twelve': 12, 'thirteen': 13, 'fourteen': 14, 'fifteen': 15
    }
    
    # Special playoff sheet names vary by season
    if season <= 2021:
        # 2020/2021: 8 teams, playoffs weeks 14-15
        playoff_sheet_names = {
            'Playoffs, Week 1': 15,
            'Playoffs, Round 1': 15,
            'Super Bowl Week': 16,
        }
    else:
        # 2022+: 10 teams, playoffs weeks 16-17
        playoff_sheet_names = {
            'Semi-Finals': 16, 
            'Championship': 17,
            'Championship Week': 17,  # May be overridden by 2.0 version
        }
        
        # "Championship Week 2.0" is the redo after the Bengals/Bills game was aborted
        # It should take precedence over "Championship Week"
        if 'Championship Week 2.0' in wb.sheetnames:
            playoff_sheet_names = {k: v for k, v in playoff_sheet_names.items() if k != 'Championship Week'}
            playoff_sheet_names['Championship Week 2.0'] = 17
    
    for sheet_name in wb.sheetnames:
        # Check playoff sheets first (before general matching)
        if sheet_name in playoff_sheet_names:
            week_sheets.append((playoff_sheet_names[sheet_name], sheet_name))
            continue
        
        # Try "Week 6" format
        match = re.match(r'^Week (\d+)$', sheet_name)
        if match:
            week_sheets.append((int(match.group(1)), sheet_name))
            continue
            
        # Try "Week One", "Week Two" format
        match = re.match(r'^Week (\w+)$', sheet_name, re.IGNORECASE)
        if match:
            word = match.group(1).lower()
            if word in word_to_num:
                week_sheets.append((word_to_num[word], sheet_name))
                continue
        
        # Try "Rivalry Week 5" or similar formats
        match = re.search(r'Week (\d+)', sheet_name)
        if match:
            week_sheets.append((int(match.group(1)), sheet_name))
            continue
    
    week_sheets.sort(key=lambda x: x[0])
    
    # Export each week
    weeks = []
    for week_num, sheet_name in week_sheets:
        print(f"  - week_{week_num}.json", end="")
        ws = wb[sheet_name]
        
        # Calculate bench scores for this week
        bench_scores = calculate_bench_scores(str(excel_path), sheet_name, season, week_num)
        if bench_scores:
            print(f" ({len(bench_scores)} bench scores)")
        else:
            print()
        
        week_data = export_week(ws, week_num, season, bench_scores)
        weeks.append(week_data)
        
        with open(weeks_dir / f"week_{week_num}.json", 'w') as f:
            json.dump(week_data, f, indent=2)
    
    wb.close()
    
    # Calculate standings
    if is_current:
        current_nfl_week = get_current_nfl_week()
        standings = calculate_standings(weeks, max_week=current_nfl_week, season=season)
    else:
        standings = calculate_standings(weeks, season=season)
        # For historical seasons, adjust standings based on playoff results
        standings = adjust_standings_for_playoffs(standings, season, weeks)
    
    # Export standings
    print("  - standings.json")
    with open(season_dir / "standings.json", 'w') as f:
        json.dump({
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "standings": standings
        }, f, indent=2)
    
    # Export meta
    print("  - meta.json")
    teams = load_teams() if is_current else []
    if not teams and weeks:
        teams = [{'name': t['name'], 'owner': t['owner'], 'abbrev': t['abbrev']} 
                 for t in weeks[-1].get('teams', [])]
    
    schedule = get_schedule_data(standings) if is_current else []
    
    meta = {
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "season": season,
        "is_current": is_current,
        "is_historical": not is_current,
        "current_week": get_current_nfl_week() if is_current else (weeks[-1]['week'] if weeks else 17),
        "trade_deadline_week": TRADE_DEADLINE_WEEK,
        "teams": teams,
        "schedule": schedule,
        "weeks_available": [w['week'] for w in weeks],
    }
    
    # 2020 Jamboree: Calculate 2-week total points for non-playoff teams
    if season == 2020:
        jamboree_standings = calculate_jamboree(weeks)
        if jamboree_standings:
            meta["jamboree"] = jamboree_standings
    
    with open(season_dir / "meta.json", 'w') as f:
        json.dump(meta, f, indent=2)
    
    # Current season only: rosters, draft picks, live data
    if is_current:
        print("  - rosters.json")
        rosters = load_rosters()
        with open(season_dir / "rosters.json", 'w') as f:
            json.dump({
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "rosters": rosters
            }, f, indent=2)
        
        traded_picks_path = PROJECT_DIR / "Traded Picks.xlsx"
        if traded_picks_path.exists():
            print("  - draft_picks.json")
            draft_picks = parse_draft_picks(str(traded_picks_path))
            with open(season_dir / "draft_picks.json", 'w') as f:
                json.dump({
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                    "picks": draft_picks
                }, f, indent=2)
        
        print("  - live.json")
        live_data = {
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "game_times": get_game_times(season),
            "fa_pool": load_fa_pool(),
            "pending_trades": load_pending_trades(),
        }
        with open(season_dir / "live.json", 'w') as f:
            json.dump(live_data, f, indent=2)
    
    print(f"Season {season} exported!")


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        season = int(sys.argv[1])
    else:
        season = CURRENT_SEASON
    
    export_season(season)

