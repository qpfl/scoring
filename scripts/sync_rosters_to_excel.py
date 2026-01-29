#!/usr/bin/env python3
"""
Sync JSON roster/transaction data to Excel (2026+).

This script updates the Rosters.xlsx file with any roster changes
from the transaction log. The Excel serves as a backup and is NOT
used for scoring (JSON is the source of truth for 2026+).

No bold formatting or scores are written - just player names.
"""

import json
from pathlib import Path

import openpyxl


def load_json(path: Path) -> dict | list:
    """Load JSON file."""
    if not path.exists():
        return {}
    with open(path) as f:
        return json.load(f)


def sync_rosters_to_excel(
    rosters_json: Path,
    excel_path: Path,
    transaction_log: Path = None,
):
    """
    Sync rosters from JSON to Excel.

    Args:
        rosters_json: Path to data/rosters.json
        excel_path: Path to Rosters.xlsx
        transaction_log: Optional path to transaction_log.json for logging
    """
    if not rosters_json.exists():
        print(f"No rosters.json found at {rosters_json}")
        return

    rosters = load_json(rosters_json)
    if not rosters:
        print("Empty rosters.json")
        return

    # Load or create Excel workbook
    if excel_path.exists():
        wb = openpyxl.load_workbook(excel_path)
        print(f"Updating existing {excel_path}")
    else:
        wb = openpyxl.Workbook()
        wb.active.title = "Rosters"
        print(f"Creating new {excel_path}")

    # Get or create the Rosters sheet
    if "Rosters" in wb.sheetnames:
        ws = wb["Rosters"]
    else:
        ws = wb.active
        ws.title = "Rosters"

    # Clear existing content (except header row)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.value = None

    # Write header row
    headers = ["Team", "Position", "Player", "NFL Team", "Status"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Write roster data
    row = 2
    for team_abbrev, players in sorted(rosters.items()):
        for player in players:
            ws.cell(row=row, column=1, value=team_abbrev)
            ws.cell(row=row, column=2, value=player.get("position", ""))
            ws.cell(row=row, column=3, value=player.get("name", ""))
            ws.cell(row=row, column=4, value=player.get("nfl_team", ""))
            ws.cell(row=row, column=5, value=player.get("status", "active"))
            row += 1

    # Save
    wb.save(excel_path)
    print(f"Saved {row - 2} players to {excel_path}")

    # Log the sync
    if transaction_log and transaction_log.exists():
        log_data = load_json(transaction_log)
        # Just note that we synced - actual transactions are in the log
        print(f"Transaction log has {len(log_data.get('transactions', []))} entries")


def main():
    """Main entry point."""
    import argparse

    parser = argparse.ArgumentParser(description="Sync rosters from JSON to Excel")
    parser.add_argument("--rosters", "-r", default="data/rosters.json", help="Path to rosters.json")
    parser.add_argument("--excel", "-e", default="Rosters.xlsx", help="Path to output Excel file")
    parser.add_argument("--transactions", "-t", default="data/transaction_log.json", help="Path to transaction log")
    args = parser.parse_args()

    project_dir = Path(__file__).parent.parent

    rosters_json = project_dir / args.rosters
    excel_path = project_dir / args.excel
    transaction_log = project_dir / args.transactions

    print("Syncing rosters to Excel...")
    print(f"  Source: {rosters_json}")
    print(f"  Target: {excel_path}")

    sync_rosters_to_excel(rosters_json, excel_path, transaction_log)

    print("Done!")


if __name__ == "__main__":
    main()

